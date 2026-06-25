"""
BOUN Web — backend FastAPI.
Reutiliza la misma lógica de la app de escritorio (database.py, ml_scraper,
ml_fees, scoring) y los mismos datos en Supabase. Expone una API REST y
sirve el frontend carbón.
"""
import os
import io
import time
import json
import base64
import threading
import secrets
from datetime import datetime, timezone, timedelta
import hmac as _hmac
import hashlib as _hashlib
from fastapi import FastAPI, HTTPException, Depends, Header, Request
from fastapi.staticfiles import StaticFiles
from fastapi.responses import (FileResponse, JSONResponse, Response,
                               RedirectResponse, HTMLResponse, StreamingResponse)
from pydantic import BaseModel
from typing import Optional, List

import database as db

app = FastAPI(title="BOUN Análisis ML")

# ── Sesiones simples en memoria (token → user) ───────────────────────────────
_SESSIONS = {}


def _current_user(authorization: Optional[str] = Header(None)) -> dict:
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "No autenticado")
    tok = authorization.split(" ", 1)[1]
    u = _SESSIONS.get(tok)
    if not u:
        raise HTTPException(401, "Sesión expirada")
    return u


def _admin(user: dict = Depends(_current_user)) -> dict:
    if user.get("role") != "admin":
        raise HTTPException(403, "Solo administradores")
    return user


# ── Auth ─────────────────────────────────────────────────────────────────────

class LoginIn(BaseModel):
    username: str
    password: str


class PwChangeIn(BaseModel):
    new_password: str


@app.get("/api/health")
def health():
    return {"ok": True, "users": db.users_count()}


@app.post("/api/login")
def login(data: LoginIn):
    r = db.verify_login(data.username, data.password)
    if not r.get("ok"):
        raise HTTPException(401, r.get("error", "Credenciales incorrectas"))
    tok = secrets.token_urlsafe(32)
    _SESSIONS[tok] = r["user"]
    return {"token": tok, "user": r["user"]}


@app.get("/api/admin-login")
def admin_login(k: str = ""):
    """Auto-login SOLO del administrador mediante un token secreto
    (ADMIN_AUTOLOGIN_TOKEN). Permite al creador entrar sin contraseña con un
    link marcado como favorito. Los demás usuarios siguen con login normal."""
    token = os.environ.get("ADMIN_AUTOLOGIN_TOKEN", "")
    if not token or k != token:
        raise HTTPException(401, "No autorizado")
    admins = [u for u in (db.list_users() or []) if u.get("role") == "admin"]
    if not admins:
        raise HTTPException(500, "Sin administrador")
    user = {"username": admins[0]["username"], "role": "admin",
            "must_change": False}
    tok = secrets.token_urlsafe(32)
    _SESSIONS[tok] = user
    return {"token": tok, "user": user}


@app.post("/api/logout")
def logout(user: dict = Depends(_current_user),
           authorization: str = Header(None)):
    _SESSIONS.pop(authorization.split(" ", 1)[1], None)
    return {"ok": True}


@app.post("/api/change-password")
def change_password(data: PwChangeIn, user: dict = Depends(_current_user)):
    if len(data.new_password) < 6:
        raise HTTPException(400, "Mínimo 6 caracteres")
    r = db.set_password(user["username"], data.new_password, must_change=False)
    if not r.get("ok"):
        raise HTTPException(400, r.get("error", "No se pudo actualizar"))
    return {"ok": True}


# ── Colaboradores (admin) ────────────────────────────────────────────────────

class UserIn(BaseModel):
    username: str
    password: str


@app.get("/api/users")
def list_users(user: dict = Depends(_admin)):
    return db.list_users()


@app.post("/api/users")
def create_user(data: UserIn, user: dict = Depends(_admin)):
    r = db.create_user(data.username, data.password, role="colaborador",
                       must_change=True)
    if not r.get("ok"):
        raise HTTPException(400, r.get("error", "No se pudo crear"))
    return {"ok": True}


@app.delete("/api/users/{username}")
def delete_user(username: str, user: dict = Depends(_admin)):
    return {"ok": db.delete_user(username)}


class ActiveIn(BaseModel):
    active: bool


@app.patch("/api/users/{username}/active")
def set_active(username: str, data: ActiveIn, user: dict = Depends(_admin)):
    return {"ok": db.set_user_active(username, data.active)}


class ResetIn(BaseModel):
    new_password: str


@app.post("/api/users/{username}/reset")
def reset_pw(username: str, data: ResetIn, user: dict = Depends(_admin)):
    r = db.set_password(username, data.new_password, must_change=True)
    if not r.get("ok"):
        raise HTTPException(400, r.get("error", "No se pudo"))
    return {"ok": True}


# ── Inventario ───────────────────────────────────────────────────────────────

@app.get("/api/inventory")
def inventory(user: dict = Depends(_current_user)):
    return db.inv_list_products()


class InvProductIn(BaseModel):
    code: str
    name: str
    cost_product: float = 0
    cost_shipping: float = 0


@app.post("/api/inventory")
def inv_create(data: InvProductIn, user: dict = Depends(_current_user)):
    r = db.inv_create_product(data.code, data.name,
                              created_by=user.get("username", ""),
                              cost_product=data.cost_product,
                              cost_shipping=data.cost_shipping)
    if not r.get("ok"):
        raise HTTPException(400, r.get("error", "No se pudo crear"))
    return r


class InvUpdateIn(BaseModel):
    code: Optional[str] = None
    name: Optional[str] = None
    cost_product: Optional[float] = None
    cost_shipping: Optional[float] = None
    qty_bogota: Optional[float] = None
    qty_yopal: Optional[float] = None
    qty_transit: Optional[float] = None
    owner: Optional[str] = None       # 'BOUN' | 'MARIA_JOSE'
    mj_qty: Optional[float] = None     # unidades de María José (0/None = todas)
    mj_anchor: Optional[str] = None    # 'YYYY-MM-DD' desde cuándo cuentan sus ventas
    mj_consumed: Optional[float] = None  # unidades ya vendidas de ella (lo pone el sync)


@app.patch("/api/inventory/{pid}")
def inv_update(pid: int, data: InvUpdateIn,
               user: dict = Depends(_current_user)):
    fields = {k: v for k, v in data.dict().items() if v is not None}
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    # Solo el administrador puede editar el SKU/código del producto.
    if "code" in fields and user.get("role") != "admin":
        raise HTTPException(403, "Solo el administrador puede editar el SKU.")
    # Las bodegas (Bogotá/Yopal) solo las edita el admin directamente; los
    # demás usuarios cargan stock con el botón "Ingreso de mercancía".
    if user.get("role") != "admin" and ("qty_bogota" in fields or
                                        "qty_yopal" in fields):
        raise HTTPException(403, "Solo el administrador puede editar las "
                                 "bodegas. Usa el botón «Ingreso de mercancía».")
    ok = db.inv_update_product(pid, fields)
    # Si se cambió el dueño (marcar/desmarcar como de María José), limpiar sus
    # ventas en la liquidación: al DESMARCAR desaparece de la sección; al
    # RE-MARCAR se parte de cero y el sync la repuebla desde la nueva fecha.
    if ok and "owner" in fields:
        try:
            db._sb_delete("mj_ventas", "product_id=eq.%d" % pid)
        except Exception:
            pass
        if fields.get("owner") == "BOUN":
            try:
                db._sb_patch("inventory_products", "id=eq.%d" % pid,
                             {"mj_consumed": 0})
            except Exception:
                pass
        _MJ_CACHE["ts"] = 0     # fuerza re-sync en la próxima carga de la sección
    return {"ok": ok}


class IngresoIn(BaseModel):
    bodega: str                       # "bogota" | "yopal"
    cantidad: float                   # > 0 (unidades que LLEGARON)
    nota: Optional[str] = ""


@app.post("/api/inventory/{pid}/ingreso")
def inv_ingreso(pid: int, data: IngresoIn, user: dict = Depends(_current_user)):
    """Ingreso de mercancía: SUMA unidades a una bodega (Bogotá o Yopal) sobre
    el valor ACTUAL del sistema; nunca reemplaza el total. Así no se borran los
    descuentos que el motor ya aplicó por ventas. Deja registro en
    movimiento_stock con delta positivo. Los combos no aplican (su stock se
    deriva de los componentes)."""
    bod = (data.bodega or "").strip().lower()
    if bod not in ("bogota", "yopal"):
        raise HTTPException(400, "bodega debe ser 'bogota' o 'yopal'")
    cant = int(data.cantidad or 0)
    if cant <= 0:
        raise HTTPException(400, "La cantidad debe ser mayor a 0")
    rows = db._sb_get("inventory_products?id=eq.%d&select=id,code,name,"
                      "qty_bogota,qty_yopal" % pid) or []
    if not rows:
        raise HTTPException(404, "Producto no encontrado")
    p = rows[0]
    codigo = p.get("code")
    if _combo_components(codigo):
        raise HTTPException(400, "Es un combo: su stock se calcula desde los "
                                 "componentes. Ingresa la mercancía en ellos.")
    col = "qty_bogota" if bod == "bogota" else "qty_yopal"
    actual = int(p.get(col) or 0)
    nuevo = actual + cant
    db._sb_patch("inventory_products", "id=eq.%d" % pid, {col: nuevo})
    nota = (data.nota or "").strip()
    db._sb_post("movimiento_stock", {
        "codigo_boun": codigo, "delta": cant,
        "motivo": "ingreso_%s%s" % (bod, (" (" + nota + ")") if nota else ""),
        "canal": "manual", "order_id": ""})
    return {"ok": True, "bodega": bod, "anterior": actual,
            "ingresado": cant, "nuevo": nuevo, "code": codigo}


@app.delete("/api/inventory/{pid}")
def inv_delete(pid: int, user: dict = Depends(_admin)):
    return {"ok": db.inv_delete_product(pid)}


class AssignIn(BaseModel):
    items: list        # [[ext_id,title,thumb,sold,qty,logistic,price,net,
                       #   margin,roas,acos,sold60,inv_id,upid,channel], …]
    channels: list = None   # canales que el diálogo administra (reemplazo
                            # selectivo). None = deducir de los items.


@app.post("/api/inventory/{pid}/links")
def inv_assign(pid: int, data: AssignIn, user: dict = Depends(_current_user)):
    return {"ok": db.inv_set_links(pid, data.items, data.channels)}


# ── Excel: descargar / subir inventario ──────────────────────────────────────
# Columnas EDITABLES (las que el import lee, mapeadas por encabezado).
_XLS_EDIT = [
    ("Código", "code"),               # llave — NO se modifica, solo empareja
    ("Producto", "name"),
    ("Costo producto", "cost_product"),
    ("Costo envío", "cost_shipping"),
    ("Bodega Bogotá", "qty_bogota"),
    ("Bodega Yopal", "qty_yopal"),
    ("En tránsito", "qty_transit"),
]
# Columnas de SOLO LECTURA (referencia; el import las ignora).
_XLS_READONLY = ["ML Full", "Inventario total", "Vendidas 60d",
                 "Publicaciones asignadas"]

_CH_LABEL = {"mercadolibre": "ML", "falabella": "Falabella",
             "shopify_boun": "Shopify BOUN", "shopify_kat": "Shopify KAT"}


def _xls_num(v):
    """Convierte una celda (número o texto '1.234' / '$ 1.234,5') a float.
    Formato colombiano: '.' = miles, ',' = decimales."""
    if v is None or v == "":
        return None
    if isinstance(v, (int, float)):
        return float(v)
    import re as _re
    s = str(v).strip().replace("$", "").replace(" ", "")
    if not s:
        return None
    if "," in s:
        # la coma es el decimal; los puntos son miles
        s = s.replace(".", "").replace(",", ".")
    elif _re.fullmatch(r"-?\d{1,3}(\.\d{3})+", s):
        # patrón de agrupación de miles con puntos (1.234 / 1.234.567) → entero
        s = s.replace(".", "")
    # en otros casos el '.' se interpreta como decimal (p.ej. '1.5')
    try:
        return float(s)
    except ValueError:
        return None


def _links_summary(p):
    """Texto legible de las publicaciones asignadas de un producto."""
    parts = []
    for l in p.get("links", []):
        ch = _CH_LABEL.get(l.get("channel") or "mercadolibre", l.get("channel"))
        ext = l.get("ml_item_id") or ""
        title = (l.get("ml_title") or "").strip()
        qty = int(float(l.get("ml_qty") or 0))
        sg = l.get("share_group")
        tag = (" •%s" % sg) if sg else ""
        parts.append("[%s] %s — %s (%du)%s" % (ch, ext, title, qty, tag))
    return "\n".join(parts)


@app.get("/api/inventory/export.xlsx")
def inv_export_xlsx(k: str = "", authorization: Optional[str] = Header(None)):
    """Descarga el inventario completo como Excel (.xlsx). Hoja 'Inventario'
    (una fila por producto, con costos/bodegas editables y un resumen de las
    publicaciones asignadas) + hoja 'Publicaciones' (una fila por publicación
    vinculada, para referencia). El archivo es el mismo formato que acepta el
    import: editar y volver a subir."""
    # Auth: header Bearer normal o ?k=<token> (para descarga directa por link).
    _current_user(authorization if authorization else (("Bearer " + k) if k else None))
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
    except Exception:
        raise HTTPException(500, "Falta la librería openpyxl en el servidor.")

    prods = db.inv_list_products()
    wb = Workbook()
    ws = wb.active
    ws.title = "Inventario"

    headers = [h for h, _ in _XLS_EDIT] + _XLS_READONLY
    ws.append(headers)
    hdr_fill = PatternFill("solid", fgColor="252427")
    ro_fill = PatternFill("solid", fgColor="3A3A3D")
    bold_w = Font(bold=True, color="F2ECE0")
    n_edit = len(_XLS_EDIT)
    for ci, _h in enumerate(headers, start=1):
        c = ws.cell(row=1, column=ci)
        c.font = bold_w
        c.fill = hdr_fill if ci <= n_edit else ro_fill
        c.alignment = Alignment(vertical="center")

    for p in prods:
        u = (int(p.get("qty_bogota") or 0) + int(p.get("qty_yopal") or 0)
             + int(p.get("qty_full") or 0) + int(p.get("qty_transit") or 0))
        ws.append([
            p.get("code", ""), p.get("name", ""),
            float(p.get("cost_product") or 0), float(p.get("cost_shipping") or 0),
            int(p.get("qty_bogota") or 0), int(p.get("qty_yopal") or 0),
            int(p.get("qty_transit") or 0),
            int(p.get("qty_full") or 0), u,
            int(p.get("sold60_total") or 0), _links_summary(p),
        ])

    widths = [18, 40, 14, 12, 14, 13, 11, 9, 14, 12, 60]
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w
    ws.freeze_panes = "A2"

    # Hoja de detalle de publicaciones
    ws2 = wb.create_sheet("Publicaciones")
    ws2.append(["Código producto", "Producto", "Canal", "ID publicación",
                "Título", "Stock", "Logística", "Comparte stock"])
    for ci in range(1, 9):
        ws2.cell(row=1, column=ci).font = bold_w
        ws2.cell(row=1, column=ci).fill = hdr_fill
    for p in prods:
        for l in p.get("links", []):
            ws2.append([
                p.get("code", ""), p.get("name", ""),
                _CH_LABEL.get(l.get("channel") or "mercadolibre",
                              l.get("channel")),
                l.get("ml_item_id") or "", (l.get("ml_title") or ""),
                int(float(l.get("ml_qty") or 0)), l.get("ml_logistic") or "",
                l.get("share_group") or "",
            ])
    for col, w in zip("ABCDEFGH", [18, 36, 13, 18, 46, 9, 14, 14]):
        ws2.column_dimensions[col].width = w
    ws2.freeze_panes = "A2"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    fname = "BOUN_inventario_%s.xlsx" % datetime.now().strftime("%Y-%m-%d")
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument."
                   "spreadsheetml.sheet",
        headers={"Content-Disposition": 'attachment; filename="%s"' % fname})


@app.post("/api/inventory/import")
async def inv_import_xlsx(request: Request,
                          user: dict = Depends(_admin)):
    """Sube un Excel (el mismo que entrega export.xlsx) y actualiza los
    productos existentes emparejando por la columna 'Código'. Solo modifica
    campos editables: Producto, Costo producto, Costo envío, Bodega Bogotá,
    Bodega Yopal, En tránsito. NO crea ni borra productos, NO toca las
    publicaciones asignadas (esas se editan con «Asignar publicaciones»).
    Los combos no actualizan bodegas (su stock se deriva de los componentes).
    Solo administrador."""
    try:
        from openpyxl import load_workbook
    except Exception:
        raise HTTPException(500, "Falta la librería openpyxl en el servidor.")
    raw = await request.body()
    if not raw:
        raise HTTPException(400, "Archivo vacío.")
    try:
        wb = load_workbook(io.BytesIO(raw), data_only=True, read_only=True)
    except Exception as e:
        raise HTTPException(400, "No se pudo leer el Excel: %s" % e)
    ws = wb["Inventario"] if "Inventario" in wb.sheetnames else wb.worksheets[0]

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(400, "La hoja está vacía.")
    header = [str(h).strip() if h is not None else "" for h in rows[0]]
    col_idx = {h: i for i, h in enumerate(header)}
    if "Código" not in col_idx:
        raise HTTPException(400, "El Excel no tiene la columna 'Código'. "
                                 "Usa el archivo descargado con «Excel».")

    # Mapa code → producto (para id + saber si es combo y valores actuales).
    prods = {str(p.get("code", "")).strip(): p for p in db.inv_list_products()}

    numeric = {"cost_product", "cost_shipping", "qty_bogota", "qty_yopal",
               "qty_transit"}
    qty_fields = {"qty_bogota", "qty_yopal", "qty_transit"}
    updated, unchanged, not_found, errors = [], [], [], []

    for r in rows[1:]:
        code = r[col_idx["Código"]] if col_idx["Código"] < len(r) else None
        code = ("" if code is None else str(code)).strip()
        if not code or code.upper().startswith("TOTAL"):
            continue
        p = prods.get(code)
        if not p:
            not_found.append(code)
            continue
        is_combo = bool(_combo_components(code))
        fields = {}
        for hdr, key in _XLS_EDIT:
            if key == "code" or hdr not in col_idx:
                continue
            idx = col_idx[hdr]
            if idx >= len(r):
                continue
            val = r[idx]
            if key in numeric:
                nv = _xls_num(val)
                if nv is None:
                    continue
                if is_combo and key in qty_fields:
                    continue  # combo: stock derivado, no tocar bodegas
                nv = round(nv, 2) if key.startswith("cost") else int(nv)
                cur = p.get(key)
                cur = (round(float(cur or 0), 2) if key.startswith("cost")
                       else int(float(cur or 0)))
                if nv != cur:
                    fields[key] = nv
            else:  # name
                sv = ("" if val is None else str(val)).strip()
                if sv and sv != (p.get("name") or "").strip():
                    fields[key] = sv
        if not fields:
            unchanged.append(code)
            continue
        try:
            db.inv_update_product(p["id"], fields)
            updated.append({"code": code, "changes": fields})
        except Exception as e:
            errors.append({"code": code, "error": str(e)})

    return {"ok": True, "updated": updated, "n_updated": len(updated),
            "unchanged": len(unchanged), "not_found": not_found,
            "errors": errors}


def _fetch_channel_items() -> tuple:
    """Recoge las publicaciones VIVAS de los 4 canales (MercadoLibre, Falabella,
    Shopify BOUN y Shopify KAT) en un formato común. Devuelve (items, ch_status).

    Cada item trae `channel`, `item_id`, `sku`, `title`, `thumbnail`,
    `inventory`, `price`, etc. `ch_status` informa por canal si su catálogo cargó
    (para que quien consuma no borre vínculos de un canal que no respondió).
    Lo usan tanto /api/inventory/items (asignación) como la sección Mapeo."""
    items = []
    ch_status = {}

    # ── MercadoLibre ──
    try:
        from ml_scraper import get_my_items_basic
        r = get_my_items_basic()
        if r.get("ok"):
            for it in r["items"]:
                it["channel"] = "mercadolibre"
                items.append(it)
            ch_status["mercadolibre"] = {"ok": True, "n": len(r["items"])}
        else:
            ch_status["mercadolibre"] = {"ok": False, "error": r.get("error")}
    except Exception as e:
        ch_status["mercadolibre"] = {"ok": False, "error": str(e)[:160]}

    # ── Falabella (SellerSku = código BOUN) ──
    try:
        import falabella as fb
        prods = fb.get_products_list()
        n = 0
        for x in prods or []:
            st = (x.get("Status") or "").lower()
            if st and st in ("deleted", "inactive"):
                continue
            sku = x.get("SellerSku") or ""
            if not sku:
                continue
            items.append({
                "channel": "falabella", "item_id": sku, "sku": sku,
                "title": x.get("Name") or sku,
                "thumbnail": x.get("Image") or "",
                "inventory": int(float(x.get("Quantity") or 0)),
                "price": float(x.get("Price") or 0),
                "sold_total": 0, "logistic_type": "",
                "net_unit": 0, "margin_pct": 0, "ad_roas": 0,
                "ad_acos": 0, "sold_60d": 0, "inventory_id": "", "upid": "",
            })
            n += 1
        ch_status["falabella"] = {"ok": True, "n": n}
    except Exception as e:
        ch_status["falabella"] = {"ok": False, "error": str(e)[:160]}

    # ── Shopify BOUN + KAT (SKU = código BOUN) ──
    for ckey, shop in _SHOPIFY_SHOPS.items():
        try:
            vs = _shopify_products_list(shop)
            for v in vs:
                items.append({
                    "channel": ckey, "item_id": v["item_id"],
                    "sku": v.get("sku") or "", "title": v.get("title") or "",
                    "thumbnail": v.get("thumbnail") or "",
                    "inventory": int(v.get("inventory") or 0),
                    "price": float(v.get("price") or 0),
                    "sold_total": 0, "logistic_type": "", "net_unit": 0,
                    "margin_pct": 0, "ad_roas": 0, "ad_acos": 0,
                    "sold_60d": 0, "inventory_id": "", "upid": "",
                })
            ch_status[ckey] = {"ok": True, "n": len(vs)}
        except Exception as e:
            ch_status[ckey] = {"ok": False, "error": str(e)[:160]}

    return items, ch_status


@app.get("/api/inventory/items")
def inventory_items(user: dict = Depends(_current_user)):
    """Publicaciones de los 4 canales para asignar al inventario + vínculos.

    Cada item trae un campo `channel`. `channels` informa, por canal, si su
    catálogo cargó bien (para que el front conserve los canales que no
    respondieron y no borre sus vínculos al guardar)."""
    items, ch_status = _fetch_channel_items()
    links = db.inv_get_links()
    any_ok = any(v.get("ok") for v in ch_status.values())
    return {"ok": any_ok, "items": items, "links": links,
            "channels": ch_status}


# ── Mis Productos ────────────────────────────────────────────────────────────

# ── Caché en servidor de Mis Productos (refresco cada 20 min) ────────────────
_MP_CACHE = {}        # days -> {"ts": epoch, "data": result}
_MP_LOCK = threading.Lock()
_MP_TTL = 20 * 60     # 20 minutos


def _fetch_my_products(days):
    from ml_scraper import get_my_products
    r = get_my_products(days=days)
    if r.get("ok"):
        try:
            inv = db.inv_links_map()
            for pr in r.get("products", []):
                pr["inv_code"] = inv.get(pr.get("item_id"), "")
        except Exception:
            pass
        # refrescar también el stock/inventory_id de los vínculos del
        # inventario (igual que el refresco de la app de escritorio)
        try:
            sm = {pr.get("item_id"): (
                pr.get("inventory", 0) or 0, pr.get("logistic_type", "") or "",
                pr.get("sold_total", 0) or 0, pr.get("price", 0) or 0,
                pr.get("net_unit", 0) or 0, pr.get("margin_pct", 0) or 0,
                pr.get("ad_roas", 0) or 0, pr.get("ad_acos", 0) or 0,
                pr.get("sold_60d", 0) or 0, pr.get("inventory_id", "") or "",
                pr.get("upid", "") or "")
                for pr in r.get("products", []) if pr.get("item_id")}
            db.inv_refresh_link_stock(sm)
        except Exception:
            pass
    return r


@app.get("/api/my-products")
def my_products(days: int = 60, force: bool = False,
                user: dict = Depends(_current_user)):
    now = time.time()
    c = _MP_CACHE.get(days)
    if c and not force and (now - c["ts"]) < _MP_TTL:
        out = dict(c["data"])
        out["cache_age_min"] = int((now - c["ts"]) / 60)
        return out
    # refrescar (con lock para no duplicar trabajo en peticiones simultáneas)
    with _MP_LOCK:
        c = _MP_CACHE.get(days)
        if c and not force and (time.time() - c["ts"]) < _MP_TTL:
            out = dict(c["data"]); out["cache_age_min"] = int((time.time()-c["ts"])/60); return out
        r = _fetch_my_products(days)
        if r.get("ok"):
            _MP_CACHE[days] = {"ts": time.time(), "data": r}
        out = dict(r); out["cache_age_min"] = 0
        return out


def _mp_background_refresh():
    """Refresca el periodo de 60 días cada 20 min, en segundo plano,
    para que la página abra al instante con datos recientes."""
    while True:
        # precalentar primero el mapa de ventas de 7 días (lo usa el export)
        try:
            _sold7_map(force=True)
        except Exception:
            pass
        try:
            r = _fetch_my_products(60)
            if r.get("ok"):
                _MP_CACHE[60] = {"ts": time.time(), "data": r}
        except Exception:
            pass
        time.sleep(_MP_TTL)


@app.on_event("startup")
def _start_bg():
    t = threading.Thread(target=_mp_background_refresh, daemon=True)
    t.start()


@app.get("/api/product-trends")
def product_trends(item_id: str = "", title: str = "", days: int = 60,
                   user: dict = Depends(_current_user)):
    """Google Trends + visitas ML de UNA publicación (panel desplegable)."""
    out = {"trends": [], "visits": {}}
    try:
        from ml_scraper import trends_for_name, get_item_visits
        out["trends"] = trends_for_name(title) or []
    except Exception:
        pass
    try:
        from ml_scraper import get_item_visits
        out["visits"] = get_item_visits(item_id, days) or {}
    except Exception:
        pass
    return out


# ── Productos para comprar (catálogo guardado) ───────────────────────────────

@app.get("/api/products")
def products(user: dict = Depends(_current_user)):
    return db.get_all_products("viability_score DESC")


@app.get("/api/products/{pid}")
def product_get(pid: int, user: dict = Depends(_current_user)):
    p = db.get_product(pid)
    if not p:
        raise HTTPException(404, "No encontrado")
    return p


class ProductPatchIn(BaseModel):
    purchase_price: Optional[float] = None
    sale_price: Optional[float] = None
    shipping_cost: Optional[float] = None
    ml_commission_total: Optional[float] = None
    total_costs: Optional[float] = None
    net_profit: Optional[float] = None
    profit_margin_pct: Optional[float] = None
    viability_score: Optional[float] = None


@app.patch("/api/products/{pid}")
def product_patch(pid: int, data: ProductPatchIn,
                  user: dict = Depends(_current_user)):
    fields = {k: v for k, v in data.dict().items() if v is not None}
    if not fields:
        raise HTTPException(400, "Nada que actualizar")
    db.update_product(pid, fields)
    return {"ok": True}


@app.delete("/api/products/{pid}")
def product_delete(pid: int, user: dict = Depends(_admin)):
    db.delete_product(pid)
    return {"ok": True}


# ── Agregar producto (analizar link ML) ─────────────────────────────────────

class AnalyzeIn(BaseModel):
    url: str
    cost: float = 0


@app.post("/api/analyze")
def analyze(data: AnalyzeIn, user: dict = Depends(_current_user)):
    from ml_scraper import analyze_ml_url
    r = analyze_ml_url(data.url, data.cost)
    if not r.get("ok"):
        raise HTTPException(400, r.get("error", "No se pudo analizar"))
    return r


class RecalcIn(BaseModel):
    sale_price: float
    cost: float
    category: str = "Otro / General"
    commission_rate: float = 0
    advertising_pct: float = 0
    competitor_count: int = 0
    search_level: int = 0


@app.post("/api/recalc")
def recalc(data: RecalcIn, user: dict = Depends(_current_user)):
    from ml_fees import calculate_fees
    from scoring import calculate_score
    from ml_scraper import _ml_shipping_cost
    ship = _ml_shipping_cost(data.sale_price)
    fees = calculate_fees(
        sale_price=data.sale_price, purchase_price=data.cost,
        category=data.category, listing_type="Clásica",
        shipping_cost=ship, advertising_pct=(data.advertising_pct or 0) / 100.0,
        commission_rate=(data.commission_rate or None))
    lvl = data.search_level or 0
    score = calculate_score(
        profit_margin_pct=fees["profit_margin_pct"],
        monthly_sales=int(lvl * 5), competitor_count=data.competitor_count,
        search_volume=int(lvl * 50), avg_rating=4.0)
    return {"fees": fees, "shipping": ship, "score": score}


class SaveProductIn(BaseModel):
    name: str
    category: str = "Otro / General"
    purchase_price: float = 0
    sale_price: float = 0
    ml_competitor_count: int = 0
    ml_category_commission: float = 0
    ml_monthly_sales: int = 0
    ml_search_volume: int = 0
    shipping_cost: float = 0
    advertising_pct: float = 0
    ml_commission_total: float = 0
    total_costs: float = 0
    viability_score: float = 0
    profit_margin_pct: float = 0
    net_profit: float = 0
    permalink: str = ""
    image_url: str = ""


@app.post("/api/products")
def product_create(data: SaveProductIn, user: dict = Depends(_current_user)):
    d = data.dict()
    d["pdf_filename"] = d.pop("permalink", "") or ""
    d["image_path"] = d.pop("image_url", "") or ""
    d["created_by"] = user.get("username", "")
    pid = db.insert_product(d)
    return {"ok": True, "id": pid}


# ── Configuración / estado ML ────────────────────────────────────────────────

@app.get("/api/ml-status")
def ml_status(user: dict = Depends(_current_user)):
    try:
        from ml_scraper import is_connected, get_connected_username
        return {"connected": is_connected(),
                "username": get_connected_username()}
    except Exception:
        return {"connected": False, "username": ""}


# Información de la empresa + credenciales (admin)
@app.get("/api/settings")
def get_settings(user: dict = Depends(_current_user)):
    g = db.get_setting
    return {
        "company_name": g("company_name", "BOUN"),
        "company_nit": g("company_nit", ""),
        "default_user": g("default_user", "Admin"),
        "currency": g("currency", "COP"),
        "ml_app_id": g("ml_app_id", ""),
        "ml_redirect_uri": g("ml_redirect_uri", "https://boun.com.co/oauth"),
        "has_secret": bool(g("ml_client_secret", "")),
    }


class SettingsIn(BaseModel):
    company_name: Optional[str] = None
    company_nit: Optional[str] = None
    default_user: Optional[str] = None
    currency: Optional[str] = None


@app.post("/api/settings")
def save_settings(data: SettingsIn, user: dict = Depends(_admin)):
    for k, v in data.dict().items():
        if v is not None:
            db.set_setting(k, str(v))
    return {"ok": True}


class CredsIn(BaseModel):
    ml_app_id: str = ""
    ml_client_secret: str = ""
    ml_redirect_uri: str = ""


@app.post("/api/ml/credentials")
def ml_creds(data: CredsIn, user: dict = Depends(_admin)):
    if data.ml_app_id:
        db.set_setting("ml_app_id", data.ml_app_id.strip())
    if data.ml_client_secret:
        db.set_setting("ml_client_secret", data.ml_client_secret.strip())
    if data.ml_redirect_uri:
        db.set_setting("ml_redirect_uri", data.ml_redirect_uri.strip())
    return {"ok": True}


@app.get("/api/ml/auth-url")
def ml_auth_url(user: dict = Depends(_admin)):
    import urllib.parse
    from ml_scraper import ML_AUTH_URL, OAUTH_REDIRECT_URI
    app_id = db.get_setting("ml_app_id", "").strip()
    secret = db.get_setting("ml_client_secret", "").strip()
    if not app_id or not secret:
        raise HTTPException(400, "Configura primero el APP ID y Client Secret.")
    redirect = db.get_setting("ml_redirect_uri", OAUTH_REDIRECT_URI).strip() or OAUTH_REDIRECT_URI
    url = (ML_AUTH_URL + "?response_type=code&client_id="
           + urllib.parse.quote(app_id) + "&redirect_uri="
           + urllib.parse.quote(redirect))
    return {"url": url}


class ExchangeIn(BaseModel):
    code: str


@app.post("/api/ml/exchange")
def ml_exchange(data: ExchangeIn, user: dict = Depends(_admin)):
    import time as _t
    from ml_scraper import (_extract_code, _get_session, get_ml_username,
                            ML_API, OAUTH_REDIRECT_URI)
    app_id = db.get_setting("ml_app_id", "").strip()
    secret = db.get_setting("ml_client_secret", "").strip()
    redirect = db.get_setting("ml_redirect_uri", OAUTH_REDIRECT_URI).strip() or OAUTH_REDIRECT_URI
    code = _extract_code(data.code)
    if not code:
        raise HTTPException(400, "Código inválido. Copia solo el código de la URL.")
    try:
        sx = _get_session()
        r = sx.post(ML_API + "/oauth/token", data={
            "grant_type": "authorization_code", "client_id": app_id,
            "client_secret": secret, "code": code, "redirect_uri": redirect},
            headers={"Content-Type": "application/x-www-form-urlencoded"}, timeout=15)
    except Exception as e:
        raise HTTPException(400, "Error de red: %s" % e)
    if r.status_code != 200:
        raise HTTPException(400, "Error al obtener token: %s" % r.text[:200])
    td = r.json()
    db.set_setting("ml_access_token", td.get("access_token", ""))
    db.set_setting("ml_refresh_token", td.get("refresh_token", ""))
    db.set_setting("ml_token_ts", str(_t.time()))
    db.set_setting("ml_user_id", str(td.get("user_id", "")))
    uname = get_ml_username(td.get("access_token", ""), str(td.get("user_id", "")))
    db.set_setting("ml_username", uname)
    return {"ok": True, "username": uname}


@app.post("/api/ml/disconnect")
def ml_disconnect(user: dict = Depends(_admin)):
    for k in ("ml_access_token", "ml_refresh_token", "ml_username",
              "ml_user_id", "ml_token_ts"):
        db.set_setting(k, "")
    return {"ok": True}


# ── Dashboard ────────────────────────────────────────────────────────────────

@app.get("/api/stats")
def stats(user: dict = Depends(_current_user)):
    return db.get_stats()


# ── Ventas diarias (MercadoLibre + Falabella + total) ────────────────────────

def _ml_ads_daily(date_list: list) -> dict:
    """{fecha → {roas, acos}} de Product Ads de ML, consultando día por día
    (en paralelo). ROAS = ingresos por ads / gasto; ACOS = gasto / ingresos.
    """
    out = {}
    try:
        from concurrent.futures import ThreadPoolExecutor
        from ml_scraper import _ml_session_auth, ML_API
        s, uid = _ml_session_auth()
        if not s:
            return out
        s.headers["Api-Version"] = "1"
        adv = s.get(f"{ML_API}/advertising/advertisers?product_id=PADS",
                    timeout=12)
        adv_id = None
        if adv.status_code == 200:
            arr = adv.json().get("advertisers", [])
            if arr:
                adv_id = arr[0].get("advertiser_id")
        if not adv_id:
            return out

        def _one(day):
            cost = rev = 0.0
            try:
                r = s.get(f"{ML_API}/advertising/advertisers/{adv_id}"
                          f"/product_ads/campaigns?date_from={day}&date_to={day}"
                          f"&metrics=cost,total_amount&limit=100", timeout=15)
                if r.status_code == 200:
                    for c in r.json().get("results", []):
                        m = c.get("metrics", {}) or {}
                        cost += m.get("cost", 0) or 0
                        rev += m.get("total_amount", 0) or 0
            except Exception:
                pass
            roas = round(rev / cost, 2) if cost > 0 else None
            acos = round(cost / rev * 100, 1) if rev > 0 else None
            return day, {"roas": roas, "acos": acos}

        with ThreadPoolExecutor(max_workers=8) as ex:
            for day, v in ex.map(_one, date_list):
                out[day] = v
    except Exception:
        pass
    return out


def _ml_thumbs(s, item_ids: list) -> dict:
    """{item_id → miniatura https} vía multiget de /items."""
    out = {}
    from ml_scraper import ML_API
    for i in range(0, len(item_ids), 20):
        ch = [x for x in item_ids[i:i + 20] if x]
        if not ch:
            continue
        try:
            r = s.get(f"{ML_API}/items?ids={','.join(ch)}"
                      f"&attributes=id,secure_thumbnail,thumbnail", timeout=15)
            if r.status_code == 200:
                for e in r.json():
                    bd = e.get("body", {}) or {}
                    u = bd.get("secure_thumbnail") or bd.get("thumbnail") or ""
                    if u.startswith("http://"):
                        u = "https://" + u[7:]
                    if bd.get("id"):
                        out[bd["id"]] = u
        except Exception:
            pass
    return out


def _ml_daily_sales(days: int = 14, date_from: str = None,
                    date_to: str = None) -> dict:
    """Ventas diarias de ML por fecha {ordenes, unidades, ingresos}.

    Rango personalizado date_from/date_to ('YYYY-MM-DD') si se pasan;
    si no, los últimos `days` días.
    """
    try:
        import datetime as _dt
        from ml_scraper import _ml_session_auth, ML_API
        s, uid = _ml_session_auth()
        if not s:
            return {"ok": False, "error": "MercadoLibre sin conexión"}
        co = _dt.timezone(_dt.timedelta(hours=-5))
        d_from = d_to = None
        if date_from:
            d1 = _dt.date.fromisoformat(date_from[:10])
            d2 = (_dt.date.fromisoformat(date_to[:10]) if date_to
                  else _dt.datetime.now(co).date())
            if d2 < d1:
                d1, d2 = d2, d1
            d_from, d_to = d1.isoformat(), d2.isoformat()
            from_d = _dt.datetime.combine(d1, _dt.time(0, 0, 0), co)
            to_d = _dt.datetime.combine(d2, _dt.time(23, 59, 59), co)
        else:
            to_d = _dt.datetime.now(co)
            from_d = (to_d - _dt.timedelta(days=days)).replace(
                hour=0, minute=0, second=0, microsecond=0)
        since = from_d.strftime("%Y-%m-%dT%H:%M:%S.000-05:00")
        until = to_d.strftime("%Y-%m-%dT%H:%M:%S.000-05:00")
        # Estados de ML que NO son una venta concretada y por tanto no deben
        # contar en el tablero: pago pendiente/en proceso, canceladas e
        # inválidas.  Además se deduplica por order_id porque orders/search
        # puede devolver la misma orden más de una vez (carrito / órdenes
        # gemelas sin pagar), lo que inflaba el conteo —p.ej. el arenero
        # marcaba 2 unidades con una sola venta real.
        _NO_VENTA = {"cancelled", "invalid", "payment_required",
                     "payment_in_process"}
        by, offset, seen = {}, 0, set()
        while True:
            r = s.get(f"{ML_API}/orders/search?seller={uid}"
                      f"&order.date_created.from={since}"
                      f"&order.date_created.to={until}"
                      f"&sort=date_desc&limit=50&offset={offset}", timeout=20)
            if r.status_code != 200:
                break
            d = r.json()
            results = d.get("results", [])
            for od in results:
                oid = str(od.get("id") or "")
                if oid and oid in seen:
                    continue            # orden ya contada (duplicada por la API)
                if oid:
                    seen.add(oid)
                if (od.get("status") or "") in _NO_VENTA:
                    continue            # no es una venta concretada
                # Agrupar por la fecha en hora Colombia (-05:00), no por el
                # timestamp crudo (que viene en UTC): así una venta de la
                # noche cae en su día correcto y no se corre al siguiente.
                dc = od.get("date_created") or ""
                try:
                    fecha = _dt.datetime.fromisoformat(
                        dc.replace("Z", "+00:00")).astimezone(co).date().isoformat()
                except Exception:
                    fecha = dc[:10]
                if not fecha:
                    continue
                if d_from and (fecha < d_from or fecha > d_to):
                    continue
                b = by.setdefault(fecha, {"fecha": fecha, "ordenes": 0,
                                          "unidades": 0, "ingresos": 0.0,
                                          "_prod": {}, "roas": None,
                                          "acos": None})
                b["ordenes"] += 1
                b["ingresos"] += float(od.get("total_amount") or 0)
                for oi in od.get("order_items", []):
                    q = int(oi.get("quantity") or 0)
                    b["unidades"] += q
                    it = oi.get("item") or {}
                    iid = it.get("id")
                    nm = (it.get("title") or "").strip()
                    if iid:
                        e = b["_prod"].setdefault(
                            iid, {"nombre": nm, "unidades": 0})
                        e["unidades"] += q
                        if nm and not e["nombre"]:
                            e["nombre"] = nm
            total = d.get("paging", {}).get("total", 0)
            offset += 50
            if offset >= total or not results:
                break
        dias = sorted(by.values(), key=lambda x: x["fecha"])
        all_ids = set()
        for x in dias:
            x["ingresos"] = round(x["ingresos"], 2)
            # TODOS los productos del día, ordenados por unidades (top primero).
            top = sorted(x.pop("_prod").items(),
                         key=lambda y: -y[1]["unidades"])
            x["top"] = [{"item_id": iid, "nombre": v["nombre"],
                         "unidades": v["unidades"]} for iid, v in top]
            for t in x["top"]:
                all_ids.add(t["item_id"])
        # miniaturas de los productos top
        thumbs = _ml_thumbs(s, list(all_ids)) if all_ids else {}
        for x in dias:
            for t in x["top"]:
                t["img"] = thumbs.get(t.pop("item_id"), "")
        # ROAS/ACOS por día desde Product Ads (si hay publicidad activa)
        try:
            ads = _ml_ads_daily([x["fecha"] for x in dias])
            for x in dias:
                a = ads.get(x["fecha"])
                if a:
                    x["roas"] = a.get("roas")
                    x["acos"] = a.get("acos")
        except Exception:
            pass
        return {"ok": True, "dias": dias}
    except Exception as e:
        return {"ok": False, "error": "MercadoLibre: %s" % str(e)[:120]}


def _shopify_orders(shop: str, token: str, since_iso: str,
                    until_iso: str) -> list:
    """Órdenes de una tienda Shopify en el rango (paginado por Link header)."""
    import requests as _rq
    import re as _re
    out = []
    url = "https://%s/admin/api/2025-01/orders.json" % shop
    params = {"status": "any", "created_at_min": since_iso,
              "created_at_max": until_iso, "limit": 250,
              "fields": "created_at,total_price,line_items,financial_status"}
    headers = {"X-Shopify-Access-Token": token}
    for _ in range(15):
        r = _rq.get(url, params=params, headers=headers, timeout=20)
        if r.status_code != 200:
            raise Exception("HTTP %d" % r.status_code)
        out.extend(r.json().get("orders", []) or [])
        nxt = None
        for part in (r.headers.get("Link", "") or "").split(","):
            if 'rel="next"' in part:
                m = _re.search(r"<([^>]+)>", part)
                if m:
                    nxt = m.group(1)
        if not nxt:
            break
        url, params = nxt, None   # la URL "next" ya trae page_info+limit
    return out


def _shopify_product_images(shop: str, token: str, product_ids) -> dict:
    """{str(product_id): url_imagen destacada} para productos Shopify.
    Las órdenes no traen la imagen, así que se consulta /products.json?ids=."""
    import requests as _rq
    out = {}
    ids = [str(p) for p in product_ids if p]
    if not ids:
        return out
    headers = {"X-Shopify-Access-Token": token}
    url = "https://%s/admin/api/2025-01/products.json" % shop
    for i in range(0, len(ids), 250):
        chunk = ids[i:i + 250]
        try:
            r = _rq.get(url, params={"ids": ",".join(chunk),
                                     "fields": "id,image", "limit": 250},
                        headers=headers, timeout=20)
            if r.status_code != 200:
                continue
            for p in r.json().get("products", []) or []:
                src = (p.get("image") or {}).get("src") or ""
                if src:
                    out[str(p.get("id"))] = src
        except Exception:
            continue
    return out


def _shopify_daily_sales(days: int = 14, date_from: str = None,
                         date_to: str = None) -> dict:
    """Ventas diarias combinadas de las tiendas Shopify (BOUN + KAT)."""
    import datetime as _dt
    co = _dt.timezone(_dt.timedelta(hours=-5))
    d_from = d_to = None
    if date_from:
        d1 = _dt.date.fromisoformat(date_from[:10])
        d2 = (_dt.date.fromisoformat(date_to[:10]) if date_to
              else _dt.datetime.now(co).date())
        if d2 < d1:
            d1, d2 = d2, d1
        d_from, d_to = d1.isoformat(), d2.isoformat()
        from_d = _dt.datetime.combine(d1, _dt.time(0, 0, 0), co)
        to_d = _dt.datetime.combine(d2, _dt.time(23, 59, 59), co)
    else:
        to_d = _dt.datetime.now(co)
        from_d = (to_d - _dt.timedelta(days=days)).replace(
            hour=0, minute=0, second=0, microsecond=0)
    since_iso, until_iso = from_d.isoformat(), to_d.isoformat()
    by, ok_any, errors, store_imgs = {}, False, [], {}
    for ckey, shop in _SHOPIFY_SHOPS.items():
        tok = db.get_setting("shopify_token::%s" % shop, "")
        if not tok:
            continue
        try:
            orders = _shopify_orders(shop, tok, since_iso, until_iso)
            ok_any = True
        except Exception as e:
            errors.append("%s: %s" % (ckey.replace("shopify_", ""),
                                      str(e)[:80]))
            continue
        tienda = ckey.replace("shopify_", "").upper()
        pids = set()
        for od in orders:
            ca = od.get("created_at") or ""
            try:
                dt = _dt.datetime.fromisoformat(
                    ca.replace("Z", "+00:00")).astimezone(co)
            except Exception:
                continue
            fecha = dt.strftime("%Y-%m-%d")
            if d_from and (fecha < d_from or fecha > d_to):
                continue
            b = by.setdefault(fecha, {"fecha": fecha, "ordenes": 0,
                                      "unidades": 0, "ingresos": 0.0,
                                      "_prod": {}, "roas": None, "acos": None})
            b["ordenes"] += 1
            b["ingresos"] += float(od.get("total_price") or 0)
            for li in od.get("line_items", []):
                q = int(li.get("quantity") or 0)
                b["unidades"] += q
                nm = (li.get("title") or "").strip()
                pid = li.get("product_id")
                if pid:
                    pids.add(pid)
                k = "%s|%s" % (tienda, li.get("sku") or nm or pid)
                e = b["_prod"].setdefault(
                    k, {"nombre": "[%s] %s" % (tienda, nm), "unidades": 0,
                        "pid": pid, "tienda": tienda})
                e["unidades"] += q
        # Imágenes de los productos vendidos de esta tienda.
        try:
            store_imgs[tienda] = _shopify_product_images(shop, tok, pids)
        except Exception:
            store_imgs[tienda] = {}
    dias = sorted(by.values(), key=lambda x: x["fecha"])
    for d in dias:
        d["ingresos"] = round(d["ingresos"], 2)
        top = sorted(d.pop("_prod").values(), key=lambda v: -v["unidades"])
        d["top"] = [{"nombre": t["nombre"], "unidades": t["unidades"],
                     "img": store_imgs.get(t.get("tienda"), {}).get(
                         str(t.get("pid")), "")} for t in top]
    if not ok_any:
        return {"ok": False,
                "error": "Shopify: " + ("; ".join(errors) or "sin tiendas")}
    return {"ok": True, "dias": dias}


_SALES_CACHE = {}        # days -> {"ts": epoch, "data": ...}
_SALES_TTL = 10 * 60
_FAL_LAST_GOOD = {}      # rango -> {"ts": epoch, "dias": [...]} última lectura ok


def _build_sales(days: int, date_from: str = None, date_to: str = None) -> dict:
    ml = _ml_daily_sales(days, date_from, date_to)
    try:
        import falabella as fb
        fa = fb.daily_sales(days, date_from, date_to)
    except Exception as e:
        fa = {"ok": False, "error": "Falabella: %s" % str(e)[:120]}
    try:
        sh = _shopify_daily_sales(days, date_from, date_to)
    except Exception as e:
        sh = {"ok": False, "error": "Shopify: %s" % str(e)[:120]}
    # Degradación elegante: si Falabella cae (503), mostrar la última lectura
    # buena de ese mismo rango marcada como "stale", en vez de quedar vacío.
    _fal_key = "%s|%s|%s" % (days, date_from or "", date_to or "")
    fal_ok = bool(fa.get("ok"))
    fal_error = fa.get("error", "")
    fal_stale, fal_as_of = False, 0
    if fal_ok:
        _FAL_LAST_GOOD[_fal_key] = {"ts": time.time(),
                                    "dias": fa.get("dias", [])}
    else:
        lg = _FAL_LAST_GOOD.get(_fal_key)
        if lg:
            fa = {"ok": True, "dias": lg["dias"]}
            fal_stale = True
            fal_as_of = int((time.time() - lg["ts"]) / 60)
    combo = {}

    _empty = lambda: {"ordenes": 0, "unidades": 0, "ingresos": 0,
                      "roas": None, "acos": None, "top": []}

    def _add(src, key):
        if src.get("ok"):
            for d in src.get("dias", []):
                b = combo.setdefault(d["fecha"], {
                    "fecha": d["fecha"], "ml": _empty(),
                    "falabella": _empty(), "shopify": _empty()})
                b[key] = {"ordenes": d["ordenes"], "unidades": d["unidades"],
                          "ingresos": d["ingresos"], "roas": d.get("roas"),
                          "acos": d.get("acos"), "top": d.get("top", [])}
    _add(ml, "ml")
    _add(fa, "falabella")
    _add(sh, "shopify")
    dias = []
    for f in sorted(combo):
        b = combo[f]
        b.setdefault("shopify", _empty())
        b["total"] = {
            "ordenes": b["ml"]["ordenes"] + b["falabella"]["ordenes"]
            + b["shopify"]["ordenes"],
            "unidades": b["ml"]["unidades"] + b["falabella"]["unidades"]
            + b["shopify"]["unidades"],
            "ingresos": round(b["ml"]["ingresos"] + b["falabella"]["ingresos"]
                              + b["shopify"]["ingresos"], 2)}
        dias.append(b)
    return {"ok": True, "days": days, "dias": dias,
            "date_from": date_from or "", "date_to": date_to or "",
            "ml_ok": bool(ml.get("ok")), "ml_error": ml.get("error", ""),
            "fal_ok": fal_ok, "fal_error": fal_error,
            "fal_stale": fal_stale, "fal_as_of": fal_as_of,
            "shop_ok": bool(sh.get("ok")), "shop_error": sh.get("error", "")}


@app.get("/api/sales")
def sales(days: int = 14, date_from: str = "", date_to: str = "",
          force: bool = False, user: dict = Depends(_current_user)):
    now = time.time()
    key = ("range:%s:%s" % (date_from, date_to)) if date_from else ("days:%d" % days)
    c = _SALES_CACHE.get(key)
    if c and not force and (now - c["ts"]) < _SALES_TTL:
        out = dict(c["data"]); out["cache_age_min"] = int((now - c["ts"]) / 60)
        return out
    data = _build_sales(days, date_from or None, date_to or None)
    _SALES_CACHE[key] = {"ts": time.time(), "data": data}
    out = dict(data); out["cache_age_min"] = 0
    return out


# ── Exportación de inventario (solo lectura, protegida por token) ─────────────
# Ruta PÚBLICA (sin sesión de usuario) pensada para que un agente externo lea
# el inventario en tiempo real pasando solo una URL con ?key=...  No toca el
# login ni las demás rutas: es puramente aditiva.

# Caché del nº de ventas por publicación en los últimos 7 días.  Se obtiene de
# ML con get_my_products(days=7) (el campo sold_60d ahí trae el conteo de la
# ventana pedida) y NO se escribe en la base, para no pisar el dato de 60 días.
_SOLD7_CACHE = {"ts": 0.0, "map": {}}
_SOLD7_TTL = 20 * 60


def _sold7_map(force: bool = False) -> dict:
    """item_id → unidades vendidas en los últimos 7 días (cacheado 20 min).

    Consulta SOLO el endpoint de órdenes de ML (ligero), no el scrape pesado
    de get_my_products — así es rápido y fiable en Render free.
    """
    now = time.time()
    if (not force and _SOLD7_CACHE["map"]
            and (now - _SOLD7_CACHE["ts"]) < _SOLD7_TTL):
        return _SOLD7_CACHE["map"]
    try:
        import datetime as _dt
        from ml_scraper import _ml_session_auth, ML_API
        s, uid = _ml_session_auth()
        if not s:
            return _SOLD7_CACHE["map"]
        to_d = _dt.date.today()
        from_d = to_d - _dt.timedelta(days=7)
        since = from_d.strftime("%Y-%m-%dT00:00:00.000-00:00")
        until = to_d.strftime("%Y-%m-%dT23:59:59.000-00:00")
        m = {}
        offset = 0
        while True:
            r = s.get(f"{ML_API}/orders/search?seller={uid}"
                      f"&order.date_created.from={since}"
                      f"&order.date_created.to={until}"
                      f"&sort=date_desc&limit=50&offset={offset}", timeout=20)
            if r.status_code != 200:
                break
            d = r.json()
            results = d.get("results", [])
            for od in results:
                for oi in od.get("order_items", []):
                    iid = (oi.get("item") or {}).get("id")
                    if iid:
                        m[iid] = m.get(iid, 0) + (oi.get("quantity", 0) or 0)
            total = d.get("paging", {}).get("total", 0)
            offset += 50
            if offset >= total or not results:
                break
        _SOLD7_CACHE["ts"] = now
        _SOLD7_CACHE["map"] = m
    except Exception:
        pass
    return _SOLD7_CACHE["map"]


_EXPORT_CORS = {
    "Access-Control-Allow-Origin": "*",
    "Access-Control-Allow-Methods": "GET, OPTIONS",
    "Access-Control-Allow-Headers": "*",
    "Cache-Control": "no-store",
}


def _img_full(thumb: str) -> str:
    """Sube la miniatura de ML a resolución original (-I.jpg → -O.jpg)."""
    t = thumb or ""
    for s in ("-I.jpg", "-I.webp", "-I.png"):
        if t.endswith(s):
            return t[: -len(s)] + s.replace("-I", "-O")
    return t


def _permalink_from_mco(mco: str) -> str:
    """Reconstruye un permalink usable desde el id de publicación (MCO…)."""
    mco = (mco or "").strip()
    if mco.startswith("MCO") and mco[3:].isdigit():
        return "https://articulo.mercadolibre.com.co/MCO-" + mco[3:]
    return ""


@app.options("/api/export/inventario.json")
def export_inventario_preflight():
    return Response(status_code=204, headers=_EXPORT_CORS)


@app.get("/api/export/inventario.json")
def export_inventario(key: str = ""):
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    if not token or key != token:
        return JSONResponse({"error": "unauthorized"}, status_code=401,
                            headers=_EXPORT_CORS)

    prods = db.inv_list_products()
    # ventas 7 días: SOLO lectura de caché (no bloquear la petición con un
    # fetch a ML). El refresco en segundo plano la mantiene caliente.
    s7 = _SOLD7_CACHE["map"]
    # ¿ya se calculó al menos una vez? Si sí, una publicación que no aparece
    # en el mapa = 0 ventas reales; si aún no, vendidos_7d = null.
    s7_warm = _SOLD7_CACHE["ts"] > 0
    productos = []
    for p in prods:
        pubs = []
        for l in p.get("links", []):
            # Este export es ML (permalinks MCO, ventas 7d de ML): los demás
            # canales no aplican aquí.
            if (l.get("channel") or "mercadolibre") != "mercadolibre":
                continue
            mco = l.get("ml_item_id") or ""
            thumb = l.get("ml_thumb") or ""
            imagenes = [_img_full(thumb)] if thumb else []
            roas_v = float(l.get("ml_roas") or 0)
            pub = {
                "mco": mco,
                "titulo": l.get("ml_title") or "",
                "permalink": _permalink_from_mco(mco),
                # Estadísticas de ventas por publicación (para rankear)
                "vendidos_60d": int(l.get("ml_sold60") or 0),
                "vendidos_7d": (s7.get(mco, 0) if s7_warm else None),
                "roas": (round(roas_v, 2) if roas_v > 0 else None),
                "imagenes": imagenes,
            }
            # Marca de stock compartido (A, B, …) si la app la detectó
            if l.get("share_group"):
                pub["comparte_stock_grupo"] = l["share_group"]
            if l.get("ml_logistic"):
                pub["logistica"] = l["ml_logistic"]
            pubs.append(pub)

        mg = float(p.get("avg_margin") or 0)
        ro = float(p.get("avg_roas") or 0)
        ac = float(p.get("avg_acos") or 0)
        productos.append({
            "codigo": p.get("code") or "",
            "nombre": p.get("name") or "",
            # La app no tiene un campo dedicado de SKU Falabella: se toma del
            # campo "Notas" del producto (queda vacío si no se ha llenado).
            "sku_falabella": (p.get("notes") or "").strip(),
            "bodega_bogota": int(p.get("qty_bogota") or 0),
            "bodega_yopal": int(p.get("qty_yopal") or 0),
            "ml_full": int(p.get("qty_full") or 0),
            "en_camino": int(p.get("qty_transit") or 0),
            "vendidos_60d": int(p.get("sold60_total") or 0),
            "margen": (f"{mg:.1f}%" if mg else "—"),
            "roas": (f"{ro:.2f}x" if ro else "—"),
            "acos": (f"{ac:.1f}%" if ac else "—"),
            # La app no calcula un sugerido de compra para el inventario.
            "sugerido_compra": None,
            "publicaciones": pubs,
        })

    out = {
        "generado": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
        "total": len(productos),
        "productos": productos,
    }
    return JSONResponse(out, headers=_EXPORT_CORS)


# ── API Falabella (ventas / catálogo / stock) — protegida por token ──────────
# Mismas reglas que el export: pública (sin sesión), ?key= vs BOUN_EXPORT_TOKEN,
# CORS abierto, no-store. Pensada para un agente externo sin internet abierto.

def _falabella_guard(key: str):
    """Devuelve una JSONResponse de error si no pasa auth/credenciales; si todo
    bien, devuelve None."""
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    if not token or key != token:
        return JSONResponse({"error": "unauthorized"}, status_code=401,
                            headers=_EXPORT_CORS)
    import falabella as fb
    if not fb.is_connected():
        return JSONResponse({"error": "missing_falabella_credentials"},
                            status_code=500, headers=_EXPORT_CORS)
    return None


@app.options("/api/falabella/ventas")
@app.options("/api/falabella/productos")
@app.options("/api/falabella/set-stock")
def falabella_preflight():
    return Response(status_code=204, headers=_EXPORT_CORS)


@app.get("/api/falabella/ventas")
def falabella_ventas(key: str = "", dias: int = 1):
    g = _falabella_guard(key)
    if g:
        return g
    import falabella as fb
    try:
        return JSONResponse(fb.ventas_por_sku(dias), headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"error": str(e)[:200]}, status_code=502,
                            headers=_EXPORT_CORS)


@app.get("/api/falabella/productos")
def falabella_productos(key: str = ""):
    g = _falabella_guard(key)
    if g:
        return g
    import falabella as fb
    try:
        return JSONResponse(fb.get_products_list(), headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"error": str(e)[:200]}, status_code=502,
                            headers=_EXPORT_CORS)


@app.get("/api/falabella/set-stock")
def falabella_set_stock(key: str = "", sku: str = "", cantidad: str = "",
                        dry: str = ""):
    g = _falabella_guard(key)
    if g:
        return g
    if not sku:
        return JSONResponse({"error": "bad_request"}, status_code=400,
                            headers=_EXPORT_CORS)
    try:
        c = int(cantidad)
        if c < 0:
            raise ValueError()
    except (ValueError, TypeError):
        return JSONResponse({"error": "bad_request"}, status_code=400,
                            headers=_EXPORT_CORS)
    import falabella as fb
    try:
        if dry == "1":
            return JSONResponse(fb.set_stock(sku, c, dry=True),
                                headers=_EXPORT_CORS)
        res = fb.set_stock(sku, c, dry=False)
        return JSONResponse({"ok": True, "sku": sku, "cantidad": c,
                             "respuesta": res}, headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)[:200]},
                            status_code=502, headers=_EXPORT_CORS)


# ── Publicador SEO multicanal (genera publicaciones) — protegido por token ──
# Mismas reglas que el export: ?key= vs BOUN_EXPORT_TOKEN, CORS abierto.
# Reúne el motor SEO (seo_publisher) + creación en Falabella (falabella).

def _pub_guard(key: str, authorization: Optional[str] = None):
    """Acepta ?key=BOUN_EXPORT_TOKEN O una sesión Bearer válida (como la SPA),
    así la página publicador.html puede llamar logueada sin exponer el token."""
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    if token and key == token:
        return None
    if authorization and authorization.startswith("Bearer "):
        if _SESSIONS.get(authorization.split(" ", 1)[1]):
            return None
    return JSONResponse({"error": "unauthorized"}, status_code=401,
                        headers=_EXPORT_CORS)


@app.options("/api/publisher/keywords")
@app.options("/api/publisher/package")
@app.options("/api/publisher/falabella/categories")
@app.options("/api/publisher/falabella/attributes")
@app.options("/api/publisher/falabella/create")
def publisher_preflight():
    return Response(status_code=204, headers=_EXPORT_CORS)


@app.get("/api/publisher/keywords")
def publisher_keywords(key: str = "", ref: str = "", family: str = "",
                       authorization: Optional[str] = Header(None)):
    g = _pub_guard(key, authorization)
    if g:
        return g
    import seo_publisher as sp
    try:
        return JSONResponse(sp.analyze_reference(ref, family or None),
                            headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"error": str(e)[:200]}, status_code=500,
                            headers=_EXPORT_CORS)


@app.get("/api/publisher/package")
def publisher_package(key: str = "", sku: str = "", ref: str = "",
                      family: str = "", color: str = "", price: int = 0,
                      list_price: int = 0, parents: str = "",
                      authorization: Optional[str] = Header(None)):
    g = _pub_guard(key, authorization)
    if g:
        return g
    import seo_publisher as sp
    fam = family or sp.detect_family(ref or sku or "")
    ar = sp.analyze_reference(ref, fam)
    ranked = ar["keywords"]
    titles = sp.title_variants(fam, color=color, ranked=ranked)
    long_t = sp.build_long_title(fam, color=color, ranked=ranked)
    out = {
        "sku": sku, "familia": fam, "color": color,
        "precio": price, "precio_lista": list_price,
        "titulo_recomendado": titles[0] if titles else "",
        "titulos": titles,
        "titulo_largo": long_t,
        "descripcion": sp.build_description(fam, color=color, ranked=ranked),
        "caracteristicas": sp.build_features(fam, color=color),
        "atributos": sp.FAMILY_SPECS.get(fam, sp.FAMILY_SPECS["sleep"]),
        "keywords": ranked[:12],
        "fotos_reglas": {"mercadolibre": sp.ML_PHOTO_RULES},
        "ml": {"parent_ids": [p.strip() for p in (parents or "").split(",")
                              if p.strip()]},
        "falabella": {"seller_sku": sku},
    }
    return JSONResponse(out, headers=_EXPORT_CORS)


@app.get("/api/publisher/falabella/categories")
def publisher_fala_categories(key: str = "", q: str = "",
                              authorization: Optional[str] = Header(None)):
    g = _pub_guard(key, authorization)
    if g:
        return g
    import falabella as fb
    if not fb.is_connected():
        return JSONResponse({"error": "missing_falabella_credentials"},
                            status_code=500, headers=_EXPORT_CORS)
    try:
        return JSONResponse({"categorias": fb.search_categories(q)[:80]},
                            headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"error": str(e)[:200]}, status_code=502,
                            headers=_EXPORT_CORS)


@app.get("/api/publisher/falabella/attributes")
def publisher_fala_attrs(key: str = "", category: str = "",
                         authorization: Optional[str] = Header(None)):
    g = _pub_guard(key, authorization)
    if g:
        return g
    import falabella as fb
    if not fb.is_connected():
        return JSONResponse({"error": "missing_falabella_credentials"},
                            status_code=500, headers=_EXPORT_CORS)
    try:
        return JSONResponse({"atributos": fb.get_category_attributes(category)},
                            headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"error": str(e)[:200]}, status_code=502,
                            headers=_EXPORT_CORS)


class FalaCreateIn(BaseModel):
    sku: str
    title: str
    description: str = ""
    brand: str = "BOUN"
    category: str
    price: int = 0
    images: list = []
    family: str = "sleep"
    color: str = ""
    dry: bool = True


@app.post("/api/publisher/falabella/create")
def publisher_fala_create(data: FalaCreateIn, key: str = "",
                          authorization: Optional[str] = Header(None)):
    g = _pub_guard(key, authorization)
    if g:
        return g
    import seo_publisher as sp
    pkg = {"sku": data.sku, "titulo_recomendado": data.title,
           "descripcion": data.description, "precio": data.price,
           "atributos": sp.FAMILY_SPECS.get(data.family,
                                            sp.FAMILY_SPECS["sleep"]),
           "fotos": data.images}
    if data.dry:
        return JSONResponse(sp.publish_falabella(pkg, category_id=data.category,
                            brand=data.brand, dry=True), headers=_EXPORT_CORS)
    import falabella as fb
    if not fb.is_connected():
        return JSONResponse({"error": "missing_falabella_credentials"},
                            status_code=500, headers=_EXPORT_CORS)
    try:
        res = sp.publish_falabella(pkg, category_id=data.category,
                                   brand=data.brand, dry=False)
        return JSONResponse({"ok": True, "respuesta": res},
                            headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)[:300]},
                            status_code=502, headers=_EXPORT_CORS)


# ── API MercadoLibre (leer/fijar SELLER_SKU) — protegida por token ───────────
# Reutiliza el token de ML del backend (mismo que /api/inventory). El access
# token NUNCA se devuelve. Refresca y reintenta una vez si ML responde 401.

def _ml_guard(key: str):
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    if not token or key != token:
        return JSONResponse({"error": "unauthorized"}, status_code=401,
                            headers=_EXPORT_CORS)
    from ml_scraper import is_connected
    if not is_connected():
        return JSONResponse({"error": "ml_not_connected"}, status_code=500,
                            headers=_EXPORT_CORS)
    return None


def _ml_request(method: str, path: str, json_body=None, timeout: int = 20,
                headers=None):
    """Petición a la API de ML con el token del backend; si responde 401,
    refresca el token y reintenta una vez. `headers` mergea headers extra
    (p. ej. x-version para /user-products/.../stock). Devuelve el Response o None."""
    from ml_scraper import _ml_session_auth, _try_refresh, ML_API
    s, _uid = _ml_session_auth()
    if not s:
        return None
    url = ML_API + path
    r = s.request(method, url, json=json_body, timeout=timeout, headers=headers)
    if r.status_code == 401:
        tok = _try_refresh()
        if tok:
            s.headers["Authorization"] = "Bearer " + tok
            r = s.request(method, url, json=json_body, timeout=timeout,
                          headers=headers)
    return r


def _seller_sku(attrs) -> Optional[str]:
    for a in (attrs or []):
        if a.get("id") == "SELLER_SKU":
            return a.get("value_name")
    return None


@app.options("/api/ml/item")
@app.options("/api/ml/set-sku")
def ml_sku_preflight():
    return Response(status_code=204, headers=_EXPORT_CORS)


@app.get("/api/ml/item")
def ml_item(key: str = "", item_id: str = ""):
    g = _ml_guard(key)
    if g:
        return g
    if not item_id:
        return JSONResponse({"ok": False, "error": "bad_request"},
                            status_code=400, headers=_EXPORT_CORS)
    try:
        r = _ml_request("GET", "/items/%s" % item_id)
        if r is None:
            return JSONResponse({"ok": False, "error": "ml_not_connected"},
                                status_code=502, headers=_EXPORT_CORS)
        if r.status_code != 200:
            return JSONResponse({"ok": False, "error": r.text[:200],
                                 "ml_status": r.status_code}, status_code=502,
                                headers=_EXPORT_CORS)
        d = r.json()
        out = {"ok": True, "item_id": item_id,
               "seller_sku": _seller_sku(d.get("attributes")),
               "variations": [
                   {"id": v.get("id"),
                    "seller_sku": (v.get("seller_custom_field")
                                   or _seller_sku(v.get("attributes")))}
                   for v in (d.get("variations") or [])]}
        return JSONResponse(out, headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)[:200]},
                            status_code=502, headers=_EXPORT_CORS)


@app.get("/api/ml/set-sku")
def ml_set_sku(key: str = "", item_id: str = "", sku: str = "", dry: str = ""):
    g = _ml_guard(key)
    if g:
        return g
    if not item_id or not sku:
        return JSONResponse({"ok": False, "error": "bad_request"},
                            status_code=400, headers=_EXPORT_CORS)
    try:
        import json as _json
        # leer el ítem: estado, catálogo y variaciones
        r = _ml_request("GET", "/items/%s" % item_id)
        if r is None:
            return JSONResponse({"ok": False, "error": "ml_not_connected"},
                                status_code=502, headers=_EXPORT_CORS)
        if r.status_code != 200:
            return JSONResponse({"ok": False, "error": r.text[:300],
                                 "ml_status": r.status_code}, status_code=502,
                                headers=_EXPORT_CORS)
        item = r.json()
        status = item.get("status")
        catalog = bool(item.get("catalog_listing"))
        variations = item.get("variations") or []
        # Publicaciones CERRADAS: ML no las deja modificar y no necesitan SKU.
        if status == "closed":
            return JSONResponse({"ok": False, "skip": True, "reason": "closed",
                                 "item_id": item_id, "status": status,
                                 "catalog_listing": catalog},
                                headers=_EXPORT_CORS)
        # Con variaciones → el SKU por variación va en seller_custom_field
        # (las variaciones no usan el atributo SELLER_SKU). Se envían TODAS.
        if variations:
            body = {"variations": [
                {"id": v.get("id"), "seller_custom_field": sku}
                for v in variations]}
            applied = "variations"
        else:
            body = {"attributes": [{"id": "SELLER_SKU", "value_name": sku}]}
            applied = "item"
        if dry == "1":
            return JSONResponse({"ok": True, "item_id": item_id, "set": sku,
                                 "applied_to": applied, "status": status,
                                 "catalog_listing": catalog, "dry_run": True,
                                 "body": body}, headers=_EXPORT_CORS)
        pr = _ml_request("PUT", "/items/%s" % item_id, json_body=body)
        if pr is None:
            return JSONResponse({"ok": False, "error": "ml_not_connected"},
                                status_code=502, headers=_EXPORT_CORS)
        if pr.status_code in (200, 201):
            return JSONResponse({"ok": True, "item_id": item_id, "set": sku,
                                 "applied_to": applied,
                                 "ml_status": pr.status_code},
                                headers=_EXPORT_CORS)
        # error: devolver el error CRUDO de ML (cause/code/message)
        try:
            err = pr.json()
        except Exception:
            err = pr.text[:500]
        txt = _json.dumps(err) if isinstance(err, (dict, list)) else str(err)
        # no modificable (catálogo/cerrada) → skip, no es un error real
        if "not_modifiable" in txt or "catalog_listing" in txt:
            return JSONResponse({"ok": False, "skip": True,
                                 "reason": "closed/catalog", "item_id": item_id,
                                 "ml_status": pr.status_code, "error": err},
                                headers=_EXPORT_CORS)
        return JSONResponse({"ok": False, "error": err,
                             "ml_status": pr.status_code}, status_code=502,
                            headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"ok": False, "error": str(e)[:200]},
                            status_code=502, headers=_EXPORT_CORS)


@app.options("/api/ml/set-stock")
def ml_set_stock_preflight():
    return Response(status_code=204, headers=_EXPORT_CORS)


def _ml_set_stock_one(item_id: str, cantidad: int, dry: bool = False,
                      max_delta=None, reactivate: bool = False) -> dict:
    """Fija available_quantity de UNA publicación ML al valor ABSOLUTO `cantidad`.
    Devuelve un dict plano (lo usan el endpoint y el motor de propagación).

    Seguridad:
    - Salta Full, cerradas y catálogo (skip, no error).
    - Escritura absoluta → idempotente: reaplicar el mismo plan no descuadra.
    - `max_delta`: si |nuevo-actual| supera el tope, NO escribe (skip=delta_guard).
    - Snapshot del stock `actual` para auditoría y reversión.
    - `reactivate=True`: si la publicación está PAUSADA por falta de stock
      (sub_status `out_of_stock`) y hay cantidad>0, la vuelve a poner ACTIVA al
      escribir el stock. Las pausadas por otra razón se saltan (no se reactivan).
    """
    try:
        c = int(cantidad)
        if c < 0:
            return {"ok": False, "error": "bad_request", "item_id": item_id}
        r = _ml_request(
            "GET", "/items/%s?attributes=id,status,sub_status,available_quantity,"
                   "variations,shipping,catalog_listing,user_product_id" % item_id)
        if r is None:
            return {"ok": False, "error": "ml_not_connected", "item_id": item_id}
        if r.status_code != 200:
            return {"ok": False, "error": r.text[:300],
                    "ml_status": r.status_code, "item_id": item_id}
        item = r.json()
        status = item.get("status")
        sub = item.get("sub_status") or []
        logistic = (item.get("shipping") or {}).get("logistic_type")
        variations = item.get("variations") or []
        # Guardrails: no tocar Full ni cerradas/catálogo (no contar como error).
        if logistic == "fulfillment":
            return {"ok": False, "skip": True, "reason": "full",
                    "item_id": item_id}
        if status == "closed":
            return {"ok": False, "skip": True, "reason": "closed",
                    "item_id": item_id}
        if item.get("catalog_listing") is True:
            # Catálogo: la API /items rechaza stock (not_modifiable). Se escribe
            # por user_product (selling_address); ML lo propaga a TODAS las
            # publicaciones ligadas a ese upid (clásica sincronizada + catálogo).
            upid = item.get("user_product_id") or ""
            if not upid:
                return {"ok": False, "skip": True, "reason": "catalog_no_upid",
                        "item_id": item_id}
            return _ml_up_stock_one(upid, c, dry=dry, max_delta=max_delta,
                                    item_id=item_id, reactivate=reactivate)
        # Pausadas: solo se tocan si reactivate=True Y están pausadas por falta de
        # stock Y hay algo que cargar. Las pausadas a propósito quedan intactas.
        reactivar = False
        if status == "paused":
            if reactivate and ("out_of_stock" in sub) and c > 0:
                reactivar = True
            else:
                return {"ok": False, "skip": True, "reason": "paused",
                        "item_id": item_id}
        # Con variaciones → fija available_quantity en CADA variación (mismo N
        # por defecto). Sin variaciones → en el ítem. `actual` = snapshot previo.
        if variations:
            applied_to = "variations"
            actual = sum(int(v.get("available_quantity") or 0)
                         for v in variations)
            body = {"variations": [{"id": v.get("id"), "available_quantity": c}
                                   for v in variations if v.get("id")]}
        else:
            applied_to = "item"
            actual = int(item.get("available_quantity") or 0)
            body = {"available_quantity": c}
        if reactivar:
            body["status"] = "active"   # revive la publicación agotada
        # Guardia de salto: evita escrituras desproporcionadas por un cálculo raro.
        if max_delta is not None and abs(c - actual) > max_delta:
            return {"ok": False, "skip": True, "reason": "delta_guard",
                    "item_id": item_id, "actual": actual, "target": c,
                    "max_delta": max_delta, "applied_to": applied_to}
        if dry:
            return {"ok": True, "dry_run": True, "item_id": item_id, "set": c,
                    "actual": actual, "applied_to": applied_to, "reactivar": reactivar,
                    "status": status, "logistic": logistic, "body": body}
        pr = _ml_request("PUT", "/items/%s" % item_id, json_body=body)
        if pr is None:
            return {"ok": False, "error": "ml_not_connected", "item_id": item_id}
        if pr.status_code in (200, 201):
            return {"ok": True, "item_id": item_id, "set": c, "actual": actual,
                    "applied_to": applied_to, "reactivar": reactivar,
                    "ml_status": pr.status_code}
        txt = (pr.text or "").lower()
        if "not_modifiable" in txt or "catalog_listing" in txt:
            return {"ok": False, "skip": True, "reason": "closed/catalog",
                    "item_id": item_id}
        try:
            err = pr.json()
        except Exception:
            err = pr.text[:500]
        return {"ok": False, "error": err, "ml_status": pr.status_code,
                "item_id": item_id}
    except Exception as e:
        return {"ok": False, "error": str(e)[:200], "item_id": item_id}


def _ml_up_stock_one(upid, c, dry=False, max_delta=None, item_id="",
                     reactivate=True, loc_type="selling_address"):
    """Stock de depósito vendedor (selling_address) de un user_product vía API
    oficial. GET /user-products/{upid}/stock (lee x-version + actual); PUT a
    /user-products/{upid}/stock/type/selling_address con {"quantity": c} y el
    x-version. ML propaga a todas las publicaciones del upid."""
    try:
        c = int(c)
    except (ValueError, TypeError):
        return {"ok": False, "error": "bad_request", "upid": upid}
    g = _ml_request("GET", "/user-products/%s/stock" % upid)
    if g is None:
        return {"ok": False, "error": "ml_not_connected", "upid": upid}
    if g.status_code != 200:
        return {"ok": False, "error": (g.text or "")[:300],
                "ml_status": g.status_code, "upid": upid, "reason": "up_get"}
    xver = g.headers.get("x-version") or g.headers.get("X-Version") or ""
    try:
        data = g.json()
    except Exception:
        data = {}
    actual = 0
    for loc in (data.get("locations") or []):
        if loc.get("type") == "selling_address":
            actual = int(loc.get("quantity") or 0)
            break
    if max_delta is not None and abs(c - actual) > max_delta:
        return {"ok": False, "skip": True, "reason": "delta_guard", "upid": upid,
                "actual": actual, "target": c, "max_delta": max_delta,
                "applied_to": "user_product", "item_id": item_id}
    if dry:
        return {"ok": True, "dry_run": True, "upid": upid, "item_id": item_id,
                "set": c, "actual": actual, "x_version": xver,
                "applied_to": "user_product"}
    hdrs = {"x-version": str(xver)} if xver else None
    body = {"quantity": c}
    pr = _ml_request("PUT", "/user-products/%s/stock/type/%s" % (upid, loc_type),
                     json_body=body, headers=hdrs)
    if pr is None:
        return {"ok": False, "error": "ml_not_connected", "upid": upid}
    if pr.status_code in (200, 201, 204):
        return {"ok": True, "upid": upid, "item_id": item_id, "set": c,
                "actual": actual, "applied_to": "user_product",
                "ml_status": pr.status_code}
    # ML bloquea la escritura por API en algunas cuentas (selling address
    # blocked). No es un error real del motor: lo tratamos como SALTADO para no
    # ensuciar el scan; la publicacion de catalogo se stockea por otra via.
    return {"ok": False, "skip": True, "reason": "catalog_blocked",
            "detail": (pr.text or "")[:300], "ml_status": pr.status_code,
            "upid": upid, "item_id": item_id, "sent_body": body}


@app.get("/api/ml/up-stock")
def ml_up_stock(key: str = "", upid: str = "", item_id: str = "",
                cantidad: str = "", dry: str = "1", loc_type: str = "selling_address",
                via: str = "", authorization: Optional[str] = Header(None)):
    """DIAGNÓSTICO/escritura de stock por user_product (catálogo). Auth por
    sesión (Bearer) o key. Sin `cantidad` → solo GET (stock + x-version). Con
    `cantidad`: dry=1 calcula; dry=0 escribe a /stock/type/{loc_type}."""
    try:
        _current_user(authorization)
    except Exception:
        gg = _ml_guard(key)
        if gg:
            return gg
    up = upid
    if not up and item_id:
        r = _ml_request("GET", "/items/%s?attributes=user_product_id,"
                        "catalog_listing" % item_id)
        if r is not None and r.status_code == 200:
            up = (r.json() or {}).get("user_product_id") or ""
    if not up:
        return JSONResponse({"ok": False, "error": "missing upid/item_id"},
                            status_code=400, headers=_EXPORT_CORS)
    # Info del item (catalog_listing/logística) para diagnóstico
    item_info = {}
    if item_id:
        ri = _ml_request("GET", "/items/%s?attributes=id,status,catalog_listing,"
                         "user_product_id,shipping" % item_id)
        if ri is not None and ri.status_code == 200:
            ij = ri.json() or {}
            item_info = {"catalog_listing": ij.get("catalog_listing"),
                         "status": ij.get("status"),
                         "user_product_id": ij.get("user_product_id"),
                         "logistic": (ij.get("shipping") or {}).get("logistic_type")}
    if cantidad == "":
        s = _ml_request("GET", "/user-products/%s/stock" % up)
        if s is None:
            return JSONResponse({"ok": False, "error": "ml_not_connected"},
                                status_code=502, headers=_EXPORT_CORS)
        try:
            body = s.json()
        except Exception:
            body = (s.text or "")[:1000]
        return JSONResponse({"ok": s.status_code == 200,
                             "ml_status": s.status_code,
                             "x_version": s.headers.get("x-version"),
                             "upid": up, "item": item_info, "stock": body},
                            status_code=200, headers=_EXPORT_CORS)
    try:
        c = int(cantidad)
        if c < 0:
            raise ValueError()
    except (ValueError, TypeError):
        return JSONResponse({"ok": False, "error": "bad_request"},
                            status_code=400, headers=_EXPORT_CORS)
    if via == "items" and item_id:
        # Probar escribir por /items (ruta de publicación normal). Para no-catálogo
        # ML acepta available_quantity; ML propaga al catálogo sincronizado.
        res = _ml_set_stock_one(item_id, c, dry=(dry == "1"), reactivate=True)
        res["item_info"] = item_info
        code = 200 if (res.get("ok") or res.get("skip")) else 502
        return JSONResponse(res, status_code=code, headers=_EXPORT_CORS)
    res = _ml_up_stock_one(up, c, dry=(dry == "1"), item_id=item_id, loc_type=loc_type)
    res["item_info"] = item_info
    code = 200 if (res.get("ok") or res.get("skip")) else 502
    return JSONResponse(res, status_code=code, headers=_EXPORT_CORS)


@app.get("/api/ml/set-stock")
def ml_set_stock(key: str = "", item_id: str = "", cantidad: str = "",
                 dry: str = ""):
    """Fija available_quantity de una publicación ML (no toca Full ni cerradas).
    dry=1 devuelve el body sin enviar."""
    g = _ml_guard(key)
    if g:
        return g
    if not item_id:
        return JSONResponse({"ok": False, "error": "bad_request"},
                            status_code=400, headers=_EXPORT_CORS)
    try:
        c = int(cantidad)
        if c < 0:
            raise ValueError()
    except (ValueError, TypeError):
        return JSONResponse({"ok": False, "error": "bad_request"},
                            status_code=400, headers=_EXPORT_CORS)
    res = _ml_set_stock_one(item_id, c, dry=(dry == "1"))
    code = 200 if (res.get("ok") or res.get("skip")) else 502
    return JSONResponse(res, status_code=code, headers=_EXPORT_CORS)


# ── Motor de sincronización — helpers de canales + pipeline (DRY-RUN) ────────

_SHOPIFY_SHOPS = {"shopify_boun": "uapngf-er.myshopify.com",
                  "shopify_kat": "2kp2p9-qu.myshopify.com"}


def _shopify_variants_by_sku(shop: str, sku: str) -> list:
    """Variantes activas de una tienda Shopify con ese SKU (= código BOUN)."""
    tok = db.get_setting("shopify_token::%s" % shop, "")
    if not tok:
        return []
    q = ('query($q:String!){productVariants(first:20,query:$q){edges{node{'
         'id sku inventoryQuantity product{status} inventoryItem{id}}}}}')
    try:
        import requests as _rq
        r = _rq.post("https://%s/admin/api/2025-01/graphql.json" % shop,
                     json={"query": q, "variables": {"q": "sku:%s" % sku}},
                     headers={"X-Shopify-Access-Token": tok}, timeout=15)
        if r.status_code != 200:
            return []
        edges = (((r.json().get("data") or {}).get("productVariants") or {})
                 .get("edges") or [])
        out = []
        for e in edges:
            n = e.get("node") or {}
            if (n.get("product") or {}).get("status") == "ACTIVE":
                out.append({"key": n.get("id"), "ventas": 0,
                            "actual": n.get("inventoryQuantity"),
                            "inv_item": (n.get("inventoryItem") or {}).get("id")})
        return out
    except Exception:
        return []


def _shopify_products_list(shop: str, limit_pages: int = 8) -> list:
    """Variantes ACTIVAS de una tienda Shopify (para el diálogo de asignar).
    Devuelve [{item_id(gid de variante), sku, title, inventory, price,
    thumbnail}, …]. Paginado y acotado para no colgar la petición."""
    tok = db.get_setting("shopify_token::%s" % shop, "")
    if not tok:
        return []
    q = ('query($cursor:String){products(first:100,after:$cursor,'
         'query:"status:active"){pageInfo{hasNextPage endCursor}'
         'edges{node{title featuredImage{url} variants(first:25){edges{node{'
         'id title sku inventoryQuantity price}}}}}}}')
    out, cursor, pages = [], None, 0
    import requests as _rq
    while pages < limit_pages:
        try:
            r = _rq.post("https://%s/admin/api/2025-01/graphql.json" % shop,
                         json={"query": q, "variables": {"cursor": cursor}},
                         headers={"X-Shopify-Access-Token": tok}, timeout=20)
            if r.status_code != 200:
                break
            data = (r.json().get("data") or {}).get("products") or {}
        except Exception:
            break
        for e in (data.get("edges") or []):
            n = e.get("node") or {}
            ptitle = n.get("title") or ""
            thumb = (n.get("featuredImage") or {}).get("url") or ""
            for ve in ((n.get("variants") or {}).get("edges") or []):
                vn = ve.get("node") or {}
                vt = vn.get("title") or ""
                full = ptitle if (not vt or vt == "Default Title") \
                    else "%s · %s" % (ptitle, vt)
                out.append({
                    "item_id": vn.get("id") or "",
                    "sku": vn.get("sku") or "",
                    "title": full,
                    "thumbnail": thumb,
                    "inventory": vn.get("inventoryQuantity") or 0,
                    "price": float(vn.get("price") or 0),
                })
        pi = data.get("pageInfo") or {}
        pages += 1
        if not pi.get("hasNextPage"):
            break
        cursor = pi.get("endCursor")
    return out


_SHOP_LOC = {}


def _shopify_location_id(shop: str, tok: str):
    """Primera location activa de la tienda (cacheada). Necesaria para fijar
    inventario con inventorySetQuantities."""
    if shop in _SHOP_LOC:
        return _SHOP_LOC[shop]
    if not tok:
        return None
    q = "query{locations(first:10){edges{node{id name isActive}}}}"
    try:
        import requests as _rq
        r = _rq.post("https://%s/admin/api/2025-01/graphql.json" % shop,
                     json={"query": q},
                     headers={"X-Shopify-Access-Token": tok}, timeout=15)
        if r.status_code != 200:
            return None
        edges = (((r.json().get("data") or {}).get("locations") or {})
                 .get("edges") or [])
        for e in edges:
            n = e.get("node") or {}
            if n.get("isActive"):
                _SHOP_LOC[shop] = n.get("id")
                return _SHOP_LOC[shop]
    except Exception:
        pass
    return None


def _shopify_set_inventory(shop: str, tok: str, inv_item: str, location: str,
                           cantidad: int, actual=None, dry: bool = False,
                           max_delta=None) -> dict:
    """Fija el inventario 'available' (valor ABSOLUTO → idempotente) de un
    inventoryItem en una location, vía inventorySetQuantities. Respeta max_delta."""
    c = int(cantidad)
    if max_delta is not None and actual is not None and abs(c - actual) > max_delta:
        return {"ok": False, "skip": True, "reason": "delta_guard",
                "inv_item": inv_item, "actual": actual, "target": c,
                "max_delta": max_delta}
    if dry:
        return {"ok": True, "dry_run": True, "shop": shop, "inv_item": inv_item,
                "actual": actual, "set": c}
    m = ("mutation($input:InventorySetQuantitiesInput!){"
         "inventorySetQuantities(input:$input){userErrors{field message}}}")
    variables = {"input": {"name": "available", "reason": "correction",
                           "ignoreCompareQuantity": True,
                           "quantities": [{"inventoryItemId": inv_item,
                                           "locationId": location,
                                           "quantity": c}]}}
    try:
        import requests as _rq
        r = _rq.post("https://%s/admin/api/2025-01/graphql.json" % shop,
                     json={"query": m, "variables": variables},
                     headers={"X-Shopify-Access-Token": tok}, timeout=20)
        if r.status_code != 200:
            return {"ok": False, "error": r.text[:200], "http": r.status_code,
                    "inv_item": inv_item}
        data = (r.json().get("data") or {}).get("inventorySetQuantities") or {}
        errs = data.get("userErrors") or []
        if errs:
            return {"ok": False, "error": errs, "inv_item": inv_item}
        return {"ok": True, "shop": shop, "inv_item": inv_item, "actual": actual,
                "set": c}
    except Exception as e:
        return {"ok": False, "error": str(e)[:200], "inv_item": inv_item}


def _ml_active_pubs(product_id: int, include_paused_oos: bool = False) -> tuple:
    """(activas[{key,ventas}], excluidas[]) de ML para un producto (no-Full).
    Si include_paused_oos=True, también mete en 'activas' las publicaciones
    pausadas por falta de stock (sub_status out_of_stock) para que el escaneo
    las reactive; las pausadas por otra razón siguen excluidas."""
    links = db._sb_get("inventory_links?product_id=eq.%d&%sselect=ml_item_id,"
                       "ml_sold60,ml_logistic"
                       % (product_id, db._ml_only_filter())) or []
    activas, excl = [], []
    for l in links:
        iid = l.get("ml_item_id")
        if not iid:
            continue
        r = _ml_request("GET", "/items/%s?attributes=id,status,sub_status,"
                        "shipping,catalog_listing,user_product_id" % iid)
        st = lg = None; sub = []; up = ""; cat = False
        if r is not None and r.status_code == 200:
            jd = r.json(); st = jd.get("status")
            lg = (jd.get("shipping") or {}).get("logistic_type")
            sub = jd.get("sub_status") or []
            up = jd.get("user_product_id") or ""
            cat = bool(jd.get("catalog_listing"))
        if lg == "fulfillment":
            excl.append({"item_id": iid, "motivo": "full"})
        elif st == "active":
            activas.append({"key": iid, "ventas": int(l.get("ml_sold60") or 0),
                            "upid": up, "catalog": cat})
        elif include_paused_oos and st == "paused" and "out_of_stock" in sub:
            # agotada → entra al reparto para que el escritor la reactive
            activas.append({"key": iid, "ventas": int(l.get("ml_sold60") or 0),
                            "upid": up, "catalog": cat})
        else:
            excl.append({"item_id": iid, "motivo": st or "?"})
    return activas, excl


def _reparto_por_userproduct(disp: int, pubs: list) -> dict:
    """Reparte `disp` entre publicaciones ML AGRUPANDO por user_product: las que
    comparten upid comparten stock, así que cuentan como UN solo grupo y se
    escriben UNA vez. Por grupo elige un representante ESCRIBIBLE (no-catálogo,
    mayor ventas) para mandarle el write por /items; ML propaga al resto del upid
    (incluida la de catálogo). Los grupos solo-catálogo igual reciben un rep
    (su write se saltará/bloqueará y se reporta)."""
    import sync as _sync
    groups = {}
    for p in pubs:
        gk = p.get("upid") or ("__solo__:" + str(p.get("key")))
        groups.setdefault(gk, []).append(p)
    reps = []
    for gk, members in groups.items():
        escribibles = [x for x in members if not x.get("catalog")]
        pool = escribibles or members
        rep = max(pool, key=lambda x: int(x.get("ventas") or 0))
        reps.append({"key": rep["key"],
                     "ventas": sum(int(x.get("ventas") or 0) for x in members)})
    return _sync.reparto(disp, reps)


def _compute_plan(codigo: str, disponible: int, reactivate: bool = False) -> dict:
    """Reparto del disponible entre publicaciones activas de los 4 canales.
    Si reactivate=True, ML incluye también las pausadas por falta de stock.
    REGLA TEMPORAL `ml_solo_bogota`: cuando está activa, MercadoLibre reparte
    SOLO el stock de Bogotá (disponible − Yopal); los demás canales siguen
    usando el disponible completo (ambas bodegas)."""
    import sync as _sync
    out = {}
    prows = db._sb_get("inventory_products?code=eq.%s&select=id,qty_yopal"
                       % _q_(codigo)) or []
    pid = prows[0]["id"] if prows else None
    disp_ml = disponible
    if prows and _ml_solo_bogota():
        if _combo_components(codigo):
            disp_ml = _combo_armable(codigo)[0]   # combo: solo armables en Bogotá
        else:
            disp_ml = max(0, disponible - int(prows[0].get("qty_yopal") or 0))
    # MercadoLibre (usa disp_ml: solo Bogotá si la regla temporal está activa)
    if pid:
        ml_act, ml_excl = _ml_active_pubs(pid, include_paused_oos=reactivate)
        out["mercadolibre"] = {"reparto": _reparto_por_userproduct(disp_ml, ml_act),
                               "excluidas": ml_excl}
    else:
        out["mercadolibre"] = {"reparto": {}, "excluidas": []}
    # Falabella: SKUs del CSV histórico + las publicaciones MAPEADAS en
    # inventory_links (sección Mapeo). Antes el reparto usaba SOLO el CSV
    # estático, así que una publicación de Falabella mapeada por Mapeo y que no
    # estuviera en el CSV (p. ej. PA001BAG-LILA) nunca recibía stock. Ahora se
    # unen ambas fuentes (dedup por seller_sku) para no dejar ninguna afuera.
    fal_skus = {f["seller_sku"] for f in _sync.falabella_skus(codigo)}
    if pid and db.channel_supported():
        try:
            for _l in (db._sb_get(
                    "inventory_links?product_id=eq.%d&channel=eq.falabella"
                    "&select=ml_item_id" % pid) or []):
                _ext = (_l.get("ml_item_id") or "").strip()
                if _ext:
                    fal_skus.add(_ext)
        except Exception:
            pass
    out["falabella"] = {"reparto": _sync.reparto(
        disponible, [{"key": sk, "ventas": 0} for sk in sorted(fal_skus)])}
    # Shopify ×2
    for ckey, shop in _SHOPIFY_SHOPS.items():
        vs = _shopify_variants_by_sku(shop, codigo)
        out[ckey] = {"reparto": _sync.reparto(disponible, vs),
                     "variantes": len(vs)}
    return out


def _q_(v):
    import urllib.parse as _u
    return _u.quote(str(v), safe="")


def _pending_cola(codigo: str) -> int:
    rows = db._sb_get("cola_bodega?codigo_boun=eq.%s&estado=eq.pendiente&"
                      "select=cantidad" % _q_(codigo)) or []
    return sum(int(r.get("cantidad") or 0) for r in rows)


# ── COMBOS (kits) ────────────────────────────────────────────────────────────
# Un combo es un producto publicado cuyo stock se calcula de sus componentes y
# que, al venderse, descuenta del inventario los productos que lo componen.
# Definición guardada como JSON en el setting `combos_def`:
#   {"PACK-DUO": [{"codigo": "PA002", "cant": 1}, {"codigo": "PA003", "cant": 1}]}

def _combos_def() -> dict:
    raw = db.get_setting("combos_def", "")
    if not raw:
        return {}
    try:
        d = json.loads(raw)
        return d if isinstance(d, dict) else {}
    except Exception:
        return {}


def _combo_components(codigo: str):
    """[{codigo, cant}, …] de un combo, o None si el código NO es un combo."""
    c = _combos_def().get(codigo)
    return c if (isinstance(c, list) and c) else None


def _combo_armable(codigo: str):
    """(armables_en_Bogotá, armables_en_Yopal). Un combo se arma con componentes
    de UNA sola bodega → se calcula el mínimo por bodega (no se mezclan)."""
    comps = _combo_components(codigo)
    if not comps:
        return (0, 0)
    bb = yy = None
    for comp in comps:
        ccod = str(comp.get("codigo") or "").strip()
        cq = max(1, int(comp.get("cant") or 1))
        if not ccod:
            continue
        rows = db._sb_get("inventory_products?code=eq.%s&select=qty_bogota,"
                          "qty_yopal" % _q_(ccod)) or []
        if not rows:
            return (0, 0)   # falta un componente → no se puede armar en ninguna
        b = int(rows[0].get("qty_bogota") or 0) // cq
        y = int(rows[0].get("qty_yopal") or 0) // cq
        bb = b if bb is None else min(bb, b)
        yy = y if yy is None else min(yy, y)
    return (max(0, bb or 0), max(0, yy or 0))


def _combo_disponible(codigo: str) -> int:
    """Total de combos armables = armables en Bogotá + armables en Yopal (cada
    combo se arma con componentes de una sola bodega; no se mezclan bodegas)."""
    b, y = _combo_armable(codigo)
    return b + y


def _sync_apply_channels() -> set:
    """Lista blanca de canales habilitados para ESCRITURA real. Separada de
    `sync_enabled` (que solo controla la ingesta de ventas). Vacío = nadie
    escribe → DRY-RUN puro. Ej.: setting `sync_apply_channels='mercadolibre'`.
    Kill-switch = vaciar el setting. Acepta coma o punto y coma."""
    raw = (db.get_setting("sync_apply_channels", "")
           or os.environ.get("SYNC_APPLY_CHANNELS", ""))
    return {c.strip() for c in raw.replace(";", ",").split(",") if c.strip()}


def _sync_apply_max_delta():
    """Tope de salto absoluto por publicación (guardia anti-cálculo-raro).
    setting `sync_apply_max_delta` o env SYNC_APPLY_MAX_DELTA. 0/vacío = sin tope."""
    try:
        v = int(db.get_setting("sync_apply_max_delta", "")
                or os.environ.get("SYNC_APPLY_MAX_DELTA", "") or "0")
        return v if v > 0 else None
    except Exception:
        return None


def _apply_plan(codigo: str, plan: dict, order_id: str = "",
                dry: bool = False, force=None, ignore_delta: bool = False,
                reactivate: bool = False) -> dict:
    """Aplica el plan de reparto SOLO a los canales en la lista blanca
    (`_sync_apply_channels`), o a `force` si se pasa (para previsualizar).

    - Hoy implementa MercadoLibre (escritor probado, valor absoluto =
      idempotente). Falabella/Shopify quedan listos para sumarse después.
    - Cada escritura se audita best-effort en la tabla `sync_aplicacion`
      (si no existe aún, `_sb_post` falla en silencio y no rompe el flujo).
    - `dry=True` calcula y devuelve lo que escribiría, sin enviar nada.
    - `ignore_delta=True` desactiva el tope de salto (lo usa el escaneo de
      reconciliación para poder rellenar agotadas grandes 0→N).
    """
    permitidos = set(force) if force else _sync_apply_channels()
    max_delta = None if ignore_delta else _sync_apply_max_delta()
    out = {"permitidos": sorted(permitidos), "dry": dry, "canales": {}}
    if not permitidos:
        return out  # nadie habilitado → no escribe (DRY-RUN puro)
    # ── MercadoLibre ──────────────────────────────────────────────────────────
    if "mercadolibre" in permitidos:
        ml_res = []
        reparto = (plan.get("mercadolibre", {}) or {}).get("reparto", {}) or {}
        for item_id, cant in reparto.items():
            r = _ml_set_stock_one(str(item_id), int(cant), dry=dry,
                                  max_delta=max_delta, reactivate=reactivate)
            ml_res.append(r)
            if not dry:
                _safe(db._sb_post, "sync_aplicacion", {
                    "codigo_boun": codigo, "canal": "mercadolibre",
                    "ref": str(item_id), "order_id": str(order_id),
                    "actual": r.get("actual"), "objetivo": r.get("set"),
                    "ok": bool(r.get("ok")),
                    "detalle": str(r.get("reason") or r.get("error") or "")[:200]})
        out["canales"]["mercadolibre"] = ml_res
    # ── Falabella ─────────────────────────────────────────────────────────────
    # fb.set_stock fija el valor absoluto (ProductUpdate) y _post ya reintenta los
    # 503/429 con backoff. No hay snapshot barato del stock actual → sin delta_guard.
    if "falabella" in permitidos:
        import falabella as fb
        fal_res = []
        reparto = (plan.get("falabella", {}) or {}).get("reparto", {}) or {}
        for sku, cant in reparto.items():
            try:
                if dry:
                    r = {"ok": True, "dry_run": True, "sku": sku, "set": int(cant)}
                else:
                    resp = fb.set_stock(sku, int(cant), dry=False)
                    r = {"ok": True, "sku": sku, "set": int(cant), "respuesta": resp}
            except Exception as e:
                r = {"ok": False, "sku": sku, "set": int(cant),
                     "error": str(e)[:200]}
            fal_res.append(r)
            if not dry:
                _safe(db._sb_post, "sync_aplicacion", {
                    "codigo_boun": codigo, "canal": "falabella", "ref": str(sku),
                    "order_id": str(order_id), "actual": None,
                    "objetivo": int(cant), "ok": bool(r.get("ok")),
                    "detalle": str(r.get("error") or "")[:200]})
        out["canales"]["falabella"] = fal_res
    # ── Shopify (BOUN y/o KAT) ────────────────────────────────────────────────
    for ckey in ("shopify_boun", "shopify_kat"):
        if ckey not in permitidos:
            continue
        shop = _SHOPIFY_SHOPS[ckey]
        tok = db.get_setting("shopify_token::%s" % shop, "")
        loc = _shopify_location_id(shop, tok)
        vmap = {v["key"]: v for v in _shopify_variants_by_sku(shop, codigo)}
        sh_res = []
        reparto = (plan.get(ckey, {}) or {}).get("reparto", {}) or {}
        for vid, cant in reparto.items():
            v = vmap.get(vid) or {}
            inv = v.get("inv_item")
            if not inv or not loc:
                r = {"ok": False, "skip": True,
                     "reason": "sin_inv_item_o_location", "ref": vid}
            else:
                r = _shopify_set_inventory(shop, tok, inv, loc, int(cant),
                                           actual=v.get("actual"), dry=dry,
                                           max_delta=max_delta)
                r["ref"] = vid
            sh_res.append(r)
            if not dry:
                _safe(db._sb_post, "sync_aplicacion", {
                    "codigo_boun": codigo, "canal": ckey, "ref": str(vid),
                    "order_id": str(order_id), "actual": v.get("actual"),
                    "objetivo": int(cant), "ok": bool(r.get("ok")),
                    "detalle": str(r.get("reason") or r.get("error") or "")[:200]})
        out["canales"][ckey] = sh_res
    return out


def _descontar_una(canal, order_id, p, cantidad, es_full):
    """Registra la venta de UN producto y asigna/descuenta bodega (misma lógica
    del motor: Full / regla ML-Bogotá / ambas→pendiente / una→auto / sin stock).
    Devuelve (asignado, bog, yop) con las bodegas ya actualizadas en memoria."""
    pid = p["id"]
    codigo = p.get("code")
    bog = int(p.get("qty_bogota") or 0)
    yop = int(p.get("qty_yopal") or 0)
    mov = db._sb_post("movimiento_stock", {
        "codigo_boun": codigo, "delta": -cantidad,
        "motivo": "venta_%s%s" % (canal, "_full" if es_full else ""),
        "canal": canal, "order_id": str(order_id)})
    cola = {"movimiento_id": (mov or {}).get("id"), "codigo_boun": codigo,
            "nombre": p.get("name"), "cantidad": cantidad,
            "canal": canal, "order_id": str(order_id)}
    if es_full is True:
        cola.update({"estado": "full"})
        db._sb_post("cola_bodega", cola)
        asignado = "full"
    elif canal == "mercadolibre" and _ml_solo_bogota() and bog > 0:
        bog = max(0, bog - cantidad)
        db._sb_patch("inventory_products", "id=eq.%d" % pid, {"qty_bogota": bog})
        db._sb_post("movimiento_stock", {
            "codigo_boun": codigo, "delta": -cantidad,
            "motivo": "asignacion_bodega_bogota_ml", "canal": canal,
            "order_id": str(order_id)})
        cola.update({"estado": "confirmado", "bodega_asignada": "bogota",
                     "auto": True})
        db._sb_post("cola_bodega", cola)
        asignado = "bogota(auto-ml)"
    elif bog > 0 and yop > 0:
        cola.update({"estado": "pendiente"})
        db._sb_post("cola_bodega", cola)
        asignado = "pendiente"
    elif bog > 0:
        bog = max(0, bog - cantidad)
        db._sb_patch("inventory_products", "id=eq.%d" % pid, {"qty_bogota": bog})
        db._sb_post("movimiento_stock", {
            "codigo_boun": codigo, "delta": -cantidad,
            "motivo": "asignacion_bodega_bogota", "canal": canal,
            "order_id": str(order_id)})
        cola.update({"estado": "confirmado", "bodega_asignada": "bogota",
                     "auto": True})
        db._sb_post("cola_bodega", cola)
        asignado = "bogota(auto)"
    elif yop > 0:
        yop = max(0, yop - cantidad)
        db._sb_patch("inventory_products", "id=eq.%d" % pid, {"qty_yopal": yop})
        db._sb_post("movimiento_stock", {
            "codigo_boun": codigo, "delta": -cantidad,
            "motivo": "asignacion_bodega_yopal", "canal": canal,
            "order_id": str(order_id)})
        cola.update({"estado": "confirmado", "bodega_asignada": "yopal",
                     "auto": True})
        db._sb_post("cola_bodega", cola)
        asignado = "yopal(auto)"
    else:
        cola.update({"estado": "pendiente"})
        db._sb_post("cola_bodega", cola)
        asignado = "pendiente(sin stock)"
    return asignado, bog, yop


def _propagar(codigo, disp, order_id):
    """Calcula el plan para un código y lo aplica a sus canales (lista blanca)."""
    plan = _compute_plan(codigo, disp)
    aplicado = _apply_plan(codigo, plan, order_id=order_id)
    return plan, aplicado


def _combo_vender(canal, order_id, codigo, cantidad, es_full):
    """Venta de un COMBO: descuenta sus componentes desde la MISMA bodega (un
    combo solo se arma si todos sus componentes están juntos). Decide cuántos
    salen de Bogotá y cuántos de Yopal; descuenta y propaga cada componente y el
    propio combo (= armables totales)."""
    comps = _combo_components(codigo) or []
    items = []
    for comp in comps:
        ccod = str(comp.get("codigo") or "").strip()
        cq = max(1, int(comp.get("cant") or 1))
        rows = db._sb_get("inventory_products?code=eq.%s&select=id,code,name,"
                          "qty_bogota,qty_yopal" % _q_(ccod)) or []
        items.append({"codigo": ccod, "cant": cq,
                      "row": (rows[0] if rows else None)})
    # ── Decidir de qué bodega(s) sale el combo (componentes de la MISMA) ──────
    plan_bog = plan_yop = 0
    if es_full is not True:
        def _cap(wh):
            best = None
            for it in items:
                if not it["row"]:
                    return 0
                mk = int(it["row"].get(wh) or 0) // it["cant"]
                best = mk if best is None else min(best, mk)
            return max(0, best or 0)
        cap_bog = _cap("qty_bogota")
        cap_yop = 0 if (canal == "mercadolibre" and _ml_solo_bogota()) \
            else _cap("qty_yopal")
        if cap_bog >= cantidad:           # cabe todo en Bogotá
            plan_bog = cantidad
        elif cap_yop >= cantidad:         # cabe todo en Yopal
            plan_yop = cantidad
        else:                             # ninguna sola alcanza → divide lo posible
            plan_bog = min(cantidad, cap_bog)
            plan_yop = min(cantidad - plan_bog, cap_yop)
    faltante = cantidad - (plan_bog + plan_yop)   # >0 = sobreventa (no había stock junto)
    # ── Descontar cada componente de la(s) bodega(s) decidida(s) ─────────────
    res_comp = []
    for it in items:
        ccod = it["codigo"]; cq = it["cant"]; row = it["row"]; ccant = cantidad * cq
        if not row:
            res_comp.append({"componente": ccod, "skip": True,
                             "motivo": "no_en_inventario"})
            continue
        pid = row["id"]
        db._sb_post("movimiento_stock", {
            "codigo_boun": ccod, "delta": -ccant,
            "motivo": "venta_combo_%s" % canal, "canal": canal,
            "order_id": str(order_id)})
        if es_full is True:
            res_comp.append({"componente": ccod, "cantidad": ccant, "bodega": "full"})
            continue
        nb = int(row.get("qty_bogota") or 0); ny = int(row.get("qty_yopal") or 0)
        from_bog = plan_bog * cq; from_yop = plan_yop * cq
        if from_bog:
            nb = max(0, nb - from_bog)
            db._sb_patch("inventory_products", "id=eq.%d" % pid, {"qty_bogota": nb})
            db._sb_post("movimiento_stock", {
                "codigo_boun": ccod, "delta": -from_bog,
                "motivo": "asignacion_bodega_bogota_combo", "canal": canal,
                "order_id": str(order_id)})
        if from_yop:
            ny = max(0, ny - from_yop)
            db._sb_patch("inventory_products", "id=eq.%d" % pid, {"qty_yopal": ny})
            db._sb_post("movimiento_stock", {
                "codigo_boun": ccod, "delta": -from_yop,
                "motivo": "asignacion_bodega_yopal_combo", "canal": canal,
                "order_id": str(order_id)})
        cdisp = max(0, nb + ny - _pending_cola(ccod))
        _propagar(ccod, cdisp, order_id)
        res_comp.append({"componente": ccod, "cantidad": ccant,
                         "bogota": from_bog, "yopal": from_yop, "disponible": cdisp})
    db._sb_post("movimiento_stock", {
        "codigo_boun": codigo, "delta": -cantidad,
        "motivo": "venta_combo_%s" % canal, "canal": canal,
        "order_id": str(order_id)})
    combo_disp = _combo_disponible(codigo)
    plan, aplicado = _propagar(codigo, combo_disp, order_id)
    return {"codigo": codigo, "combo": True, "cantidad": cantidad,
            "componentes": res_comp, "disponible_combo": combo_disp,
            "reparto_bodega": {"bogota": plan_bog, "yopal": plan_yop,
                               "faltante": max(0, faltante)},
            "plan": plan, "aplicado": aplicado}


def _process_sale(canal: str, order_id: str, items: list,
                  payload=None) -> dict:
    """Pipeline (DRY-RUN): idempotencia → descuento central (cola_bodega) →
    recalcula disponible → plan de reparto a los 4 canales. NO escribe en
    los canales. items = [(codigo_boun, cantidad), …]. Si el código vendido es
    un COMBO, descuenta sus componentes en vez de la bodega del combo."""
    import datetime as _dt
    ev = db._sb_get("evento_venta?canal=eq.%s&order_id=eq.%s&select=id,estado"
                    % (_q_(canal), _q_(order_id))) or []
    # idempotencia: si ya está procesado o en curso, no volver a descontar
    if ev and ev[0].get("estado") in ("procesado", "procesando"):
        return {"ok": True, "idempotente": True, "canal": canal,
                "order_id": order_id, "estado": ev[0].get("estado")}
    if not ev:
        db._sb_post("evento_venta", {"canal": canal, "order_id": str(order_id),
                                     "estado": "procesando", "payload": payload})
    else:
        db._sb_patch("evento_venta", "canal=eq.%s&order_id=eq.%s"
                     % (_q_(canal), _q_(order_id)), {"estado": "procesando"})
    resultados = []
    for raw in items:
        # items: (codigo, cantidad) o (codigo, cantidad, es_full). es_full=True
        # cuando el canal informa que el envío salió de Full (ML logistic_type
        # =fulfillment); None = desconocido → se decide por stock de bodega.
        codigo, cantidad = raw[0], int(raw[1])
        es_full = raw[2] if len(raw) >= 3 else None
        # COMBO: descuenta sus componentes en vez de la bodega del propio combo.
        if _combo_components(codigo):
            resultados.append(_combo_vender(canal, order_id, codigo, cantidad,
                                            es_full))
            continue
        prows = db._sb_get("inventory_products?code=eq.%s&select=id,code,name,"
                           "qty_bogota,qty_yopal" % _q_(codigo)) or []
        if not prows:
            resultados.append({"codigo": codigo, "skip": True,
                               "motivo": "no_en_inventario_central"})
            continue
        # Descuento de bodega + registro (helper reutilizable, también para combos)
        asignado, bog, yop = _descontar_una(canal, order_id, prows[0], cantidad,
                                            es_full)
        disp = max(0, bog + yop - _pending_cola(codigo))
        # Propagación a canales: escribe SOLO los de la lista blanca (si vacía,
        # sigue en DRY-RUN).
        plan, aplicado = _propagar(codigo, disp, order_id)
        resultados.append({"codigo": codigo, "cantidad": cantidad,
                           "es_full": bool(es_full), "bodega": asignado,
                           "disponible_tras_venta": disp,
                           "plan": plan, "aplicado": aplicado})
    db._sb_patch("evento_venta", "canal=eq.%s&order_id=eq.%s"
                 % (_q_(canal), _q_(order_id)),
                 {"estado": "procesado",
                  "procesado_at": _dt.datetime.now(
                      _dt.timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")})
    return {"ok": True, "dry_run": not bool(_sync_apply_channels()),
            "canal": canal, "order_id": order_id, "resultados": resultados}


def _sync_enabled() -> bool:
    """Interruptor maestro de la ingestión automática (poller + webhooks).
    Apagado por defecto: el motor queda desplegado pero no procesa ventas reales
    hasta activarlo (setting sync_enabled=1). /api/sync/simular siempre funciona."""
    return (db.get_setting("sync_enabled", "") == "1"
            or os.environ.get("SYNC_ENABLED", "") == "1")


def _ml_code_for_item(item_id: str) -> str:
    """código BOUN de una publicación ML (vía inventory_links → product code)."""
    links = db._sb_get("inventory_links?ml_item_id=eq.%s&%sselect=product_id"
                       % (_q_(item_id), db._ml_only_filter())) or []
    if not links:
        return ""
    p = db._sb_get("inventory_products?id=eq.%d&select=code"
                   % links[0]["product_id"]) or []
    return p[0]["code"] if p else ""


def _ml_item_full(item_id: str) -> bool:
    """¿La publicación ML vendida es Full? Usa el logistic_type guardado en
    inventory_links (=fulfillment). Si no hay dato, asume que NO (la lógica de
    bodega decide)."""
    links = db._sb_get("inventory_links?ml_item_id=eq.%s&%sselect=ml_logistic"
                       % (_q_(item_id), db._ml_only_filter())) or []
    return bool(links) and (links[0].get("ml_logistic") == "fulfillment")


def _process_async(canal, order_id, items, payload=None):
    threading.Thread(target=lambda: _safe(_process_sale, canal, order_id,
                                          items, payload), daemon=True).start()


def _safe(fn, *a, **k):
    try:
        fn(*a, **k)
    except Exception:
        pass


# ── Webhooks (responden 200 rápido; el pipeline corre en hilo) ───────────────

@app.post("/webhooks/shopify/{tienda}")
async def webhook_shopify(tienda: str, request: Request):
    shop = _SHOPIFY_SHOPS.get("shopify_%s" % tienda)
    if not shop:
        return Response(status_code=404)
    _cid, sec = _shopify_app_creds(shop)
    raw = await request.body()
    given = request.headers.get("x-shopify-hmac-sha256", "")
    calc = base64.b64encode(
        _hmac.new(sec.encode(), raw, _hashlib.sha256).digest()).decode()
    if not (sec and given and _hmac.compare_digest(calc, given)):
        return Response(status_code=401)
    try:
        data = json.loads(raw or b"{}")
        oid = str(data.get("id") or data.get("order_id") or "")
        items = [((li.get("sku") or "").strip(), int(li.get("quantity") or 0))
                 for li in data.get("line_items", [])
                 if (li.get("sku") or "").strip() and li.get("quantity")]
        if oid and items and _sync_enabled():
            _process_async("shopify_%s" % tienda, oid, items,
                           {"webhook": "shopify"})
    except Exception:
        pass
    return Response(status_code=200)


@app.post("/webhooks/mercadolibre")
async def webhook_ml(request: Request):
    try:
        data = await request.json()
    except Exception:
        data = {}
    resource = data.get("resource", "") or ""
    # solo órdenes
    if "/orders/" in resource and _sync_enabled():
        oid = resource.rstrip("/").split("/")[-1]

        def _work():
            r = _ml_request("GET", "/orders/%s" % oid)
            if not r or r.status_code != 200:
                return
            agg = {}
            for oi in r.json().get("order_items", []):
                iid = (oi.get("item") or {}).get("id")
                qty = int(oi.get("quantity") or 0)
                code = _ml_code_for_item(iid) if iid else ""
                if code and qty:
                    k = (code, _ml_item_full(iid))
                    agg[k] = agg.get(k, 0) + qty
            items = [(c, q, f) for (c, f), q in agg.items()]
            if items:
                _safe(_process_sale, "ml", oid, items, {"webhook": "ml"})
        threading.Thread(target=_work, daemon=True).start()
    return Response(status_code=200)


# ── Poller Falabella (cada SYNC_FALABELLA_POLL_SEC) ──────────────────────────

def _falabella_poll_once():
    """Procesa órdenes Falabella desde la activación, RE-ESCANEANDO los últimos
    días en cada ciclo para no perder ninguna por un 503 puntual o una orden que
    llega tarde/desordenada. La idempotencia (evento_venta=procesado) evita
    descontar dos veces. `sync_falabella_since` es solo el PISO (no se reprocesa
    lo anterior a la activación); ya NO avanza saltando órdenes. Las órdenes que
    aún no mapean a un código no se marcan → se reintentan solas al corregir el SKU."""
    import falabella as fb
    import sync as _sync
    if not _sync_enabled() or not fb.is_connected():
        return
    co = timezone(timedelta(hours=-5))
    floor = db.get_setting("sync_falabella_since", "")
    if not floor:
        # primera activación: marca "desde ahora", no reprocesa el histórico
        db.set_setting("sync_falabella_since",
                       datetime.now(co).strftime("%Y-%m-%d %H:%M:%S"))
        return
    _norm = lambda s: str(s or "").replace("T", " ")[:19]  # formatos comparables
    floor_n = _norm(floor)
    # Ventana acotada de re-escaneo: últimos N días (>= piso de activación).
    dias = int(os.environ.get("SYNC_FALABELLA_LOOKBACK_DAYS", "3") or 3)
    lookback = (datetime.now(co) - timedelta(days=max(1, dias))).isoformat()
    orders = fb.get_orders(lookback)
    pend = []
    for o in orders:
        oid = str(o.get("OrderId") or "")
        if not oid or _norm(o.get("CreatedAt", "")) < floor_n:
            continue                       # sin id o anterior a la activación
        ev = db._sb_get("evento_venta?canal=eq.falabella&order_id=eq.%s&"
                        "select=estado" % _q_(oid)) or []
        if ev and ev[0].get("estado") == "procesado":
            continue                       # idempotencia: ya descontada
        pend.append(o)
    if not pend:
        return
    items_map = fb._items_by_order([str(o["OrderId"]) for o in pend])
    for o in pend:
        oid = str(o["OrderId"]); ca = o.get("CreatedAt", "")
        agg = {}
        for nm, sku in items_map.get(oid, []):
            code = _sync.FAL_SKU_TO_BOUN.get(sku)
            if code:
                agg[code] = agg.get(code, 0) + 1
        if agg:
            _safe(_process_sale, "falabella", oid, list(agg.items()),
                  {"poller": "falabella", "created": ca})


def _falabella_poller():
    sec = int(os.environ.get("SYNC_FALABELLA_POLL_SEC", "180") or 180)
    while True:
        _safe(_falabella_poll_once)
        time.sleep(max(60, sec))


def _ml_poll_once():
    """Procesa SOLO órdenes ML nuevas (posteriores al watermark). Reemplaza al
    webhook (que requiere configurar callback tras 2FA en el panel ML)."""
    import sync as _sync
    if not _sync_enabled():
        return
    from ml_scraper import _ml_session_auth, ML_API
    s, uid = _ml_session_auth()
    if not s:
        return
    co = timezone(timedelta(hours=-5))
    fmt = "%Y-%m-%dT%H:%M:%S.000-05:00"
    wm = db.get_setting("sync_ml_since", "")
    if not wm:
        db.set_setting("sync_ml_since", datetime.now(co).strftime(fmt))
        return
    try:
        base = (datetime.strptime(wm[:19], "%Y-%m-%dT%H:%M:%S")
                .replace(tzinfo=co) - timedelta(hours=2))
    except Exception:
        base = datetime.now(co) - timedelta(hours=2)
    since = base.strftime(fmt)
    until = datetime.now(co).strftime(fmt)
    newest = wm
    offset = 0
    while True:
        r = s.get("%s/orders/search?seller=%s&order.date_created.from=%s"
                  "&order.date_created.to=%s&sort=date_desc&limit=50&offset=%d"
                  % (ML_API, uid, since, until, offset), timeout=20)
        if r.status_code != 200:
            break
        d = r.json()
        results = d.get("results", [])
        for od in results:
            dc = od.get("date_created", "")
            if dc <= wm:
                continue
            newest = max(newest, dc)
            oid = str(od.get("id"))
            ev = db._sb_get("evento_venta?canal=eq.ml&order_id=eq.%s&"
                            "select=estado" % _q_(oid)) or []
            if ev and ev[0].get("estado") in ("procesado", "procesando"):
                continue
            agg = {}
            for oi in od.get("order_items", []):
                iid = (oi.get("item") or {}).get("id")
                qty = int(oi.get("quantity") or 0)
                code = _ml_code_for_item(iid) if iid else ""
                if code and qty:
                    k = (code, _ml_item_full(iid))
                    agg[k] = agg.get(k, 0) + qty
            if agg:
                items = [(c, q, f) for (c, f), q in agg.items()]
                _safe(_process_sale, "ml", oid, items,
                      {"poller": "ml", "created": dc})
        total = d.get("paging", {}).get("total", 0)
        offset += 50
        if offset >= total or not results:
            break
    if newest > wm:
        db.set_setting("sync_ml_since", newest)


def _poller_loop():
    sec = int(os.environ.get("SYNC_FALABELLA_POLL_SEC", "180") or 180)
    sec = max(60, sec)
    while True:
        _safe(_falabella_poll_once)
        _safe(_ml_poll_once)
        time.sleep(sec)


@app.on_event("startup")
def _start_poller():
    threading.Thread(target=_poller_loop, daemon=True).start()


@app.post("/api/sync/falabella/reprocess")
def falabella_reprocess(order_id: str = "", dias: int = 30,
                        user: dict = Depends(_admin)):
    """Sana UNA orden Falabella puntual (p.ej. saltada por el piso/ventana del
    poller) reprocesándola por el MISMO camino que el poller (_process_sale).
    Acepta el OrderId interno O el OrderNumber visible al cliente: busca la
    orden en los últimos `dias` días y usa su OrderId canónico.
    Es idempotente: si ya está en evento_venta=procesado NO descuenta de nuevo,
    y al marcarla el poller jamás la volverá a tocar (cero doble descuento).
    Devuelve diagnóstico: piso, conexión, estado previo, orden hallada e items."""
    import falabella as fb
    import sync as _sync
    q = str(order_id or "").strip()
    out = {"buscado": q,
           "floor": db.get_setting("sync_falabella_since", ""),
           "connected": bool(fb.is_connected())}
    if not q:
        out["ok"] = False
        out["error"] = "order_id requerido"
        return out
    # 1) Ubicar la orden por OrderId o por OrderNumber en la ventana dada.
    co = timezone(timedelta(hours=-5))
    after = (datetime.now(co) - timedelta(days=max(1, int(dias)))).isoformat()
    canonical = q
    created = ""
    try:
        orders = fb.get_orders(after)
    except Exception as e:
        orders = []
        out["warn_get_orders"] = str(e)[:160]
    match = None
    for o in orders:
        if str(o.get("OrderId") or "") == q or \
           str(o.get("OrderNumber") or "") == q:
            match = o
            break
    if match:
        canonical = str(match.get("OrderId") or q)
        created = str(match.get("CreatedAt") or "")
        out["hallada"] = {"order_id": canonical,
                          "order_number": str(match.get("OrderNumber") or ""),
                          "created": created,
                          "status": str(match.get("Status") or "")}
    else:
        out["hallada"] = None
        out["nota"] = ("no aparece en los últimos %d días; puede ser más "
                       "antigua o un id distinto" % int(dias))
    out["order_id"] = canonical
    # 2) Idempotencia sobre el OrderId canónico.
    ev = db._sb_get("evento_venta?canal=eq.falabella&order_id=eq.%s&"
                    "select=estado" % _q_(canonical)) or []
    out["evento_previo"] = ev[0].get("estado") if ev else None
    if ev and ev[0].get("estado") == "procesado":
        out["ok"] = True
        out["ya_procesada"] = True
        return out
    # 3) Traer items de la orden canónica.
    try:
        items_map = fb._items_by_order([canonical])
    except Exception as e:
        out["ok"] = False
        out["error"] = "items: " + str(e)[:200]
        return out
    raw = items_map.get(canonical, [])
    out["items_raw"] = raw
    agg = {}
    for _nm, sku in raw:
        code = _sync.FAL_SKU_TO_BOUN.get(sku)
        if code:
            agg[code] = agg.get(code, 0) + 1
    out["agg"] = [[k, v] for k, v in agg.items()]
    if not agg:
        out["ok"] = False
        out["error"] = "sin items mapeables (revisar SKU/orden)"
        return out
    _process_sale("falabella", canonical, list(agg.items()),
                  {"poller": "reprocess-manual", "created": created})
    out["ok"] = True
    out["reprocesada"] = True
    return out


@app.get("/api/sync/falabella/find")
def falabella_find(sku: str = "", dias: int = 30,
                   user: dict = Depends(_admin)):
    """Diagnóstico: escanea las órdenes Falabella de los últimos `dias` días y
    lista las que contienen `sku` (SellerSku), con su OrderId/OrderNumber real,
    fecha, estado y si ya fueron procesadas en evento_venta. Sirve para ubicar
    una venta que no se descontó cuando no se conoce su OrderId interno."""
    import falabella as fb
    import sync as _sync
    sku = str(sku or "").strip()
    co = timezone(timedelta(hours=-5))
    after = (datetime.now(co) - timedelta(days=max(1, int(dias)))).isoformat()
    try:
        orders = fb.get_orders(after)
    except Exception as e:
        return {"ok": False, "error": "get_orders: " + str(e)[:200]}
    ids = [str(o.get("OrderId") or "") for o in orders if o.get("OrderId")]
    try:
        items_map = fb._items_by_order(ids)
    except Exception as e:
        items_map = {}
    rows = []
    for o in orders:
        oid = str(o.get("OrderId") or "")
        its = items_map.get(oid, [])
        skus = [s for _n, s in its]
        if sku and sku not in skus:
            continue
        ev = db._sb_get("evento_venta?canal=eq.falabella&order_id=eq.%s&"
                        "select=estado" % _q_(oid)) or []
        rows.append({"order_id": oid,
                     "order_number": str(o.get("OrderNumber") or ""),
                     "created": str(o.get("CreatedAt") or ""),
                     "status": str(o.get("Status") or ""),
                     "skus": skus,
                     "boun": [_sync.FAL_SKU_TO_BOUN.get(s) for s in skus],
                     "evento": ev[0].get("estado") if ev else None})
    return {"ok": True, "sku": sku, "dias": int(dias),
            "floor": db.get_setting("sync_falabella_since", ""),
            "n_orders_total": len(orders), "n_match": len(rows), "rows": rows}


@app.get("/api/sync/simular")
def sync_simular(key: str = "", canal: str = "test", order_id: str = "",
                 codigo: str = "", cantidad: int = 1, full: str = ""):
    """Simula una venta y corre el pipeline (DRY-RUN) para probar end-to-end.
    full=1 marca la venta como Full (no descuenta bodega)."""
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    if not token or key != token:
        return JSONResponse({"error": "unauthorized"}, status_code=401,
                            headers=_EXPORT_CORS)
    if not codigo or not order_id:
        return JSONResponse({"error": "bad_request",
                             "hint": "codigo y order_id requeridos"},
                            status_code=400, headers=_EXPORT_CORS)
    try:
        es_full = True if full == "1" else None
        res = _process_sale(canal, order_id,
                            [(codigo, int(cantidad), es_full)],
                            payload={"simulado": True})
        return JSONResponse(res, headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"error": str(e)[:200]}, status_code=502,
                            headers=_EXPORT_CORS)


# ── Motor de sincronización — PREVIEW de escritura (no envía nada) ────────────

@app.options("/api/sync/apply-preview")
def sync_apply_preview_preflight():
    return Response(status_code=204, headers=_EXPORT_CORS)


@app.get("/api/sync/apply-preview")
def sync_apply_preview(key: str = "", codigo: str = "", vendidos: int = 0,
                       canal: str = "mercadolibre"):
    """Muestra EXACTAMENTE lo que el motor escribiría en un canal para `codigo`
    (snapshot actual → objetivo por publicación), SIN enviar nada. Úsalo antes de
    poblar `sync_apply_channels`. Por defecto previsualiza MercadoLibre."""
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    if not token or key != token:
        return JSONResponse({"error": "unauthorized"}, status_code=401,
                            headers=_EXPORT_CORS)
    if not codigo:
        return JSONResponse({"error": "bad_request"}, status_code=400,
                            headers=_EXPORT_CORS)
    try:
        prows = db._sb_get("inventory_products?code=eq.%s&select=qty_bogota,"
                           "qty_yopal" % _q_(codigo)) or []
        if not prows:
            return JSONResponse({"error": "codigo_no_encontrado",
                                 "codigo": codigo}, status_code=404,
                                headers=_EXPORT_CORS)
        if _combo_components(codigo):
            disp = max(0, _combo_disponible(codigo) - int(vendidos or 0))
        else:
            bog = int(prows[0].get("qty_bogota") or 0)
            yop = int(prows[0].get("qty_yopal") or 0)
            disp = max(0, bog + yop - _pending_cola(codigo) - int(vendidos or 0))
        plan = _compute_plan(codigo, disp)
        prev = _apply_plan(codigo, plan, dry=True, force={canal})
        return JSONResponse({"ok": True, "codigo": codigo, "disponible": disp,
                             "plan": plan, "preview": prev,
                             "nota": "dry-run: no se escribió nada"},
                            headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"error": str(e)[:200]}, status_code=502,
                            headers=_EXPORT_CORS)


# ── Activación de la escritura real — toggle admin (sin SQL) ─────────────────

_APPLY_CHANNELS_VALIDOS = {"mercadolibre", "falabella",
                           "shopify_boun", "shopify_kat"}


@app.get("/api/sync/apply-status")
def sync_apply_status(user: dict = Depends(_admin)):
    """Estado actual de la propagación real: qué canales escriben, el tope de
    salto y si el motor está en DRY-RUN. Solo lectura (admin)."""
    canales = sorted(_sync_apply_channels())
    return {"ok": True, "channels": canales,
            "max_delta": _sync_apply_max_delta(),
            "dry_run": not bool(canales),
            "sync_enabled": _sync_enabled(),
            "scan_daily": _scan_daily_enabled(),
            "scan_daily_hour": _scan_daily_hour(),
            "scan_reactivate": _scan_reactivate_enabled(),
            "ml_solo_bogota": _ml_solo_bogota(),
            "validos": sorted(_APPLY_CHANNELS_VALIDOS)}


class ApplyConfigIn(BaseModel):
    channels: Optional[list] = None       # lista blanca de escritura ([] = kill-switch)
    max_delta: Optional[int] = None       # tope de salto por publicación (0/None = sin tope)
    scan_daily: Optional[bool] = None     # escaneo de reconciliación automático diario
    scan_daily_hour: Optional[int] = None # hora local Colombia (0-23) del escaneo diario
    scan_reactivate: Optional[bool] = None # reactivar publicaciones ML agotadas en el escaneo
    ml_solo_bogota: Optional[bool] = None  # regla temporal: ML solo vende bodega Bogotá


@app.post("/api/sync/apply-config")
def sync_apply_config(data: ApplyConfigIn, user: dict = Depends(_admin)):
    """Prende/apaga la escritura real por canal SIN tocar SQL ni redeploy.
    - channels=[] → DRY-RUN (kill-switch instantáneo).
    - channels=['mercadolibre'] → solo ML escribe.
    Valida contra la lista de canales soportados. Devuelve el estado resultante.
    """
    if data.channels is not None:
        pedidos = {str(c).strip() for c in data.channels if str(c).strip()}
        invalidos = pedidos - _APPLY_CHANNELS_VALIDOS
        if invalidos:
            raise HTTPException(400, "Canales no válidos: %s"
                                % ", ".join(sorted(invalidos)))
        db.set_setting("sync_apply_channels", ",".join(sorted(pedidos)))
    if data.max_delta is not None:
        if data.max_delta < 0:
            raise HTTPException(400, "max_delta no puede ser negativo")
        db.set_setting("sync_apply_max_delta", str(int(data.max_delta)))
    if data.scan_daily is not None:
        db.set_setting("sync_scan_daily", "1" if data.scan_daily else "0")
    if data.scan_daily_hour is not None:
        h = int(data.scan_daily_hour)
        if not (0 <= h <= 23):
            raise HTTPException(400, "hora debe estar entre 0 y 23")
        db.set_setting("sync_scan_daily_hour", str(h))
    if data.scan_reactivate is not None:
        db.set_setting("sync_scan_reactivate", "1" if data.scan_reactivate else "0")
    if data.ml_solo_bogota is not None:
        db.set_setting("ml_solo_bogota", "1" if data.ml_solo_bogota else "0")
    canales = sorted(_sync_apply_channels())
    return {"ok": True, "channels": canales,
            "max_delta": _sync_apply_max_delta(),
            "scan_daily": _scan_daily_enabled(),
            "scan_daily_hour": _scan_daily_hour(),
            "scan_reactivate": _scan_reactivate_enabled(),
            "ml_solo_bogota": _ml_solo_bogota(),
            "dry_run": not bool(canales)}


# ── CEREBRO — mapa de trabajo de la IA ───────────────────────────────────────
# Vista única del trabajo automatizado de BOUN: estado real del motor de
# sincronización (leído de settings), las tareas programadas (definición + el
# último reporte que cada una deja vía heartbeat) y los pendientes/fallas que
# necesitan decisión humana. La página /cerebro consume este JSON.

# Definición de las tareas programadas (corren en Cowork; aquí solo su ficha).
# `cron` es informativo; el front calcula la ventana de ejecución con `hours`/`days`.
_CEREBRO_TASKS = [
    {"id": "mercadolibre-boun-inventario-diario", "canal": "mercadolibre",
     "nombre": "Inventario diario", "icon": "box",
     "cad": "Diario · 8:00 AM", "hours": [8], "days": None,
     "desc": "Revisa publicación por publicación; en las agotadas quita el envío "
             "Full y deja el stock en 0 (primera medida autorizada).",
     "run": "Recorriendo publicaciones activas…",
     "done": "Revisión completa · agotadas marcadas",
     "idle": "Listo hasta mañana 8:00 AM"},
    {"id": "mercadolibre-boun-preguntas-reclamos", "canal": "mercadolibre",
     "nombre": "Preguntas, reclamos y facturación", "icon": "chat",
     "cad": "2× día · 9:00 AM y 4:00 PM", "hours": [9, 16], "days": None,
     "desc": "Responde compradores solo con datos de la ficha, gestiona reclamos "
             "y envía el RUT a Edgar por WhatsApp.",
     "run": "Leyendo preguntas y reclamos nuevos…",
     "done": "Bandeja respondida · facturación al día",
     "idle": "Próxima pasada a las 4:00 PM"},
    {"id": "mercadolibre-boun-promo-mensual", "canal": "mercadolibre",
     "nombre": "Campaña promo mensual (BOUN)", "icon": "tag",
     "cad": "Mensual · día 7, 1:00 PM", "hours": [13], "days": [7],
     "desc": "Crea la BOUN del mes y alinea las promociones a los precios del mes "
             "anterior. No sube ni confirma sin tu aprobación.",
     "run": "Armando la BOUN del mes…",
     "done": "BOUN del mes lista para tu revisión",
     "idle": "Programada para el día 7"},
    {"id": "denuncias-diario", "canal": "mercadolibre",
     "nombre": "Protección de marca (denuncias)", "icon": "alert",
     "cad": "Diario · 8:00 PM", "hours": [20], "days": None,
     "desc": "Busca tu marca en el Brand Protection Program, denuncia a los "
             "vendedores que se cuelgan de tus catálogos BOUN y hace seguimiento "
             "del estado de cada denuncia.",
     "run": "Detectando y denunciando infractores…",
     "done": "Infractores denunciados · estados actualizados",
     "idle": "Próxima corrida a las 8:00 PM"},
    {"id": "falabella-boun-inventario-diario", "canal": "falabella",
     "nombre": "Corrida diaria de contenido", "icon": "spark",
     "cad": "Diario · 11:00 PM", "hours": [23], "days": None,
     "desc": "Sube el puntaje de contenido a 100, pide reseñas y aplica el playbook "
             "de reseñas 1★. El inventario ya NO lo toca (lo hace el motor).",
     "run": "Optimizando fichas y reseñas…",
     "done": "Catálogo en puntaje 100 · reporte listo",
     "idle": "Próxima corrida a las 11:00 PM"},
    {"id": "falabella-boun-auditoria-quincenal", "canal": "falabella",
     "nombre": "Auditoría quincenal", "icon": "chart",
     "cad": "Días 1 y 15 · 9:00 AM", "hours": [9], "days": [1, 15],
     "desc": "Audita ventas, productos killers y ROAS contra la línea base. Solo "
             "lee y reporta, no modifica nada.",
     "run": "Auditando ventas y ROAS…",
     "done": "Auditoría lista vs. línea base",
     "idle": "Próxima auditoría el día 1"},
    {"id": "falabella-boun-ajuste-pauta-quincenal", "canal": "falabella",
     "nombre": "Ajuste de pauta quincenal", "icon": "target",
     "cad": "Días 1 y 16 · 5:00 PM", "hours": [17], "days": [1, 16],
     "desc": "Optimiza campañas con 15 días de resultados: pausa sin stock, recorta "
             "ACOS alto y escala el ROAS sano.",
     "run": "Recalculando campañas de Retail Media…",
     "done": "Pauta optimizada · sin stock pausado",
     "idle": "Próximo ajuste el día 1"},
]

_CEREBRO_SKILLS = [
    {"nombre": "Optimizador SEO Falabella", "icon": "search", "tag": "Chrome",
     "desc": "Optimiza títulos, descripciones y puntaje de contenido con tendencias "
             "reales de búsqueda en Colombia."},
    {"nombre": "Playbook reseñas 1★", "icon": "star", "tag": "Embebida",
     "desc": "Flujo duplicar → corregir → eliminar para neutralizar reseñas de una "
             "estrella en Falabella."},
    {"nombre": "Reportes (Word · Excel · PDF)", "icon": "file", "tag": "Documentos",
     "desc": "Genera auditorías, resúmenes de ventas y reportes operativos en "
             "formato profesional."},
]

# Pendientes MANUALES extra (opcional): se añaden vía el setting JSON
# `cerebro_alertas` (lista de {sev:'err'|'warn', title, txt}). Por defecto VACÍO:
# los pendientes ya NO se hardcodean. La fuente principal de "Pendientes de la IA"
# son ahora los heartbeats — ver `_cerebro_pendientes()`: cualquier tarea/skill que
# reporte `warn`/`err` aparece como pendiente y, cuando vuelve a correr en `ok`,
# desaparece sola. Así Sebastián solo ve fallas vigentes (regla 16-jun-2026).
_CEREBRO_ALERTAS_DEFAULT = []


def _cerebro_pendientes(tasks: list) -> list:
    """Construye la lista de "Pendientes de la IA" que ve Sebastián.

    Regla (16-jun-2026): toda tarea/skill reporta a Cerebro y los pendientes se
    auto-actualizan. Fuente:
      1) DERIVADOS de los heartbeats: cada tarea cuyo último estado sea `warn` o
         `err` se convierte en pendiente; al reportar `ok`/`run` deja de aparecer
         (auto-limpieza en la siguiente corrida, sin pendientes obsoletos).
      2) MANUALES: lo que haya en el setting `cerebro_alertas` (para pendientes que
         no estén atados a una tarea). Se pueden auto-limpiar si llevan `task_id`
         y esa tarea ya reportó `ok`.
    """
    pend, vistos = [], set()
    # 1) derivados de heartbeats
    for t in tasks:
        hb = t.get("heartbeat") or {}
        st = hb.get("status")
        if st in ("warn", "err"):
            tid = t.get("id")
            vistos.add(tid)
            cuando = hb.get("last_run") or ""
            txt = hb.get("msg") or t.get("desc") or "Sin detalle."
            if cuando:
                txt = f"{txt} · último reporte {cuando}"
            pend.append({"sev": st, "title": t.get("nombre") or tid,
                         "txt": txt, "task_id": tid, "auto": True})
    # 2) manuales del setting (si su task_id ya reportó ok/run, se omite = limpio)
    try:
        raw = db.get_setting("cerebro_alertas", "")
        manuales = json.loads(raw) if raw else []
        if not isinstance(manuales, list):
            manuales = []
    except Exception:
        manuales = []
    estados = {t.get("id"): (t.get("heartbeat") or {}).get("status") for t in tasks}
    for a in manuales:
        if not isinstance(a, dict):
            continue
        tid = a.get("task_id")
        if tid:
            if tid in vistos:        # ya está como derivado
                continue
            if estados.get(tid) in ("ok", "run"):  # resuelto → no mostrar
                continue
        pend.append(a)
    return pend


def _cerebro_estado() -> dict:
    """Heartbeats de las tareas: lo que cada skill reporta tras correr.
    Mapa {task_id: {status, msg, last_run, next_run}} guardado en el setting
    JSON `cerebro_estado`. Vacío si ninguna ha reportado todavía."""
    try:
        raw = db.get_setting("cerebro_estado", "") or "{}"
        d = json.loads(raw)
        return d if isinstance(d, dict) else {}
    except Exception:
        return {}


@app.get("/api/cerebro")
def cerebro(user: dict = Depends(_current_user)):
    """Mapa de trabajo de la IA: motor (estado real desde settings), tareas
    programadas (ficha + último heartbeat) y pendientes. Cualquier usuario."""
    canales = sorted(_sync_apply_channels())
    try:
        scan_last = json.loads(db.get_setting("sync_scan_last", "") or "{}")
    except Exception:
        scan_last = {}
    motor = {
        "sync_enabled": _sync_enabled(),
        "apply_channels": canales,
        "dry_run": not bool(canales),
        "max_delta": _sync_apply_max_delta(),
        "scan_daily": _scan_daily_enabled(),
        "scan_daily_hour": _scan_daily_hour(),
        "scan_daily_last": db.get_setting("sync_scan_daily_last", "") or None,
        "scan_reactivate": _scan_reactivate_enabled(),
        "ml_solo_bogota": _ml_solo_bogota(),
        "scan_last": scan_last,
    }
    estado = _cerebro_estado()
    ids_estaticos = {t["id"] for t in _CEREBRO_TASKS}
    tasks = []
    for t in _CEREBRO_TASKS:
        hb = estado.get(t["id"], {}) if isinstance(estado, dict) else {}
        tasks.append({**t, "heartbeat": hb})
    # AUTO-DESCUBRIMIENTO: cualquier proceso que haya reportado un heartbeat y no
    # esté en la ficha estática aparece solo (se auto-registra desde lo que envió).
    for tid, hb in (estado.items() if isinstance(estado, dict) else []):
        if tid in ids_estaticos or not isinstance(hb, dict):
            continue
        tasks.append({
            "id": tid,
            "canal": hb.get("canal") or "sistema",
            "nombre": hb.get("nombre") or tid,
            "icon": hb.get("icon") or "bolt",
            "cad": hb.get("cad") or "proceso dinámico",
            "hours": None, "days": None,
            "desc": hb.get("desc") or "Proceso reportado automáticamente al Cerebro.",
            "auto": True, "heartbeat": hb,
        })
    return {"ok": True,
            "now_utc": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
            "motor": motor, "tasks": tasks,
            "skills": _CEREBRO_SKILLS, "alertas": _cerebro_pendientes(tasks)}


class HeartbeatIn(BaseModel):
    task_id: str
    status: str = "ok"          # ok | run | warn | err | remove (borra el reporte)
    msg: str = ""               # qué hizo / qué está haciendo
    next_run: Optional[str] = None
    # Auto-registro de procesos NUEVOS (opcionales): si un proceso que no está en
    # la ficha estática manda estos campos, aparece solo en el Cerebro.
    nombre: Optional[str] = None
    canal: Optional[str] = None   # mercadolibre | falabella | shopify | sistema | …
    desc: Optional[str] = None
    cad: Optional[str] = None     # cadencia legible, ej. "Diario · 6:00 AM"
    icon: Optional[str] = None    # box|chat|tag|spark|chart|target|search|star|file|bolt|clock


@app.post("/api/cerebro/heartbeat")
def cerebro_heartbeat(data: HeartbeatIn, key: str = "",
                      authorization: Optional[str] = Header(None)):
    """Las tareas/procesos reportan aquí su estado real tras correr. Acepta DOS
    formas de auth: (1) ?key=<BOUN_EXPORT_TOKEN> para el planificador sin sesión, o
    (2) sesión admin (Authorization: Bearer) para que las tareas que ya operan la
    web logueadas reporten sin manejar el secreto. Un proceso NUEVO que mande
    nombre/canal/desc se auto-registra y aparece solo. status='remove' lo borra."""
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    authed = bool(token and key == token)
    if not authed:
        try:
            u = _current_user(authorization)
            authed = bool(u and u.get("role") == "admin")
        except Exception:
            authed = False
    if not authed:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    estado = _cerebro_estado()
    if data.status == "remove":
        estado.pop(data.task_id, None)
        db.set_setting("cerebro_estado", json.dumps(estado, ensure_ascii=False))
        return {"ok": True, "task_id": data.task_id, "removed": True}
    entry = dict(estado.get(data.task_id, {}))   # conserva meta previa
    entry.update({
        "status": data.status, "msg": data.msg, "next_run": data.next_run,
        "last_run": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ"),
    })
    for k, v in (("nombre", data.nombre), ("canal", data.canal),
                 ("desc", data.desc), ("cad", data.cad), ("icon", data.icon)):
        if v:
            entry[k] = v
    estado[data.task_id] = entry
    db.set_setting("cerebro_estado", json.dumps(estado, ensure_ascii=False))
    return {"ok": True, "task_id": data.task_id}


# ── MAPEO DE SKU — auditoría de sincronización publicación ↔ inventario ───────
# La fuente de verdad del stock es el inventario BOUN (inventory_products). Cada
# publicación viva (ML / Falabella / Shopify) debe estar vinculada
# (inventory_links) al SKU correcto. Aquí se audita por canal y se reporta lo
# que falta mapear o está cruzado.
#
# Visitar la sección /mapeo (o llamar /api/mapeo/scan) dispara la auditoría,
# persiste el snapshot en Supabase (mapeo_pendientes) y reporta a Cerebro
# (heartbeat `mapeo-sku-diario`). La skill diaria "mapeo de sku" la corre a las
# 8:00 PM. Diseñado para extenderse a tiendas futuras: cualquier canal que
# _fetch_channel_items() sepa leer se audita solo.

_MAPEO_TASK_ID = "mapeo-sku-diario"
_MAPEO_CACHE = {"ts": 0, "data": None}
_MAPEO_LOCK = threading.Lock()
_MAPEO_TTL = 5 * 60   # 5 min: varias cargas seguidas no re-escanean los 4 canales
_MAPEO_DESC = ("Audita ML, Falabella y Shopify y verifica que cada publicación "
               "esté vinculada al SKU correcto del inventario BOUN.")


def _mapeo_link(channel: str, ext_id: str, sku: str = "") -> str:
    """Mejor URL disponible hacia la publicación, por canal."""
    ext_id = (ext_id or "").strip()
    if channel == "mercadolibre":
        return _permalink_from_mco(ext_id)
    if channel == "falabella":
        # No tenemos la URL pública; el Seller Center la encuentra por SellerSku.
        return ("https://sellercenter.falabella.com.co/products/manage?search=%s"
                % _q_(sku or ext_id))
    # Shopify: ext_id = gid de la variante; sin handle no hay URL pública directa.
    return ""


def _cerebro_set_heartbeat(task_id, status, msg, *, nombre=None, canal=None,
                           desc=None, cad=None, icon=None, next_run=None):
    """Escribe un heartbeat al Cerebro desde el backend (sin pasar por HTTP)."""
    try:
        estado = _cerebro_estado()
        entry = dict(estado.get(task_id, {}))
        entry.update({"status": status, "msg": msg, "next_run": next_run,
                      "last_run": datetime.now(timezone.utc).strftime(
                          "%Y-%m-%dT%H:%M:%SZ")})
        for k, v in (("nombre", nombre), ("canal", canal), ("desc", desc),
                     ("cad", cad), ("icon", icon)):
            if v:
                entry[k] = v
        estado[task_id] = entry
        db.set_setting("cerebro_estado", json.dumps(estado, ensure_ascii=False))
    except Exception:
        pass


def _mapeo_audit() -> dict:
    """Audita los canales contra inventory_links con DOBLE verificación:

      A) Por publicación viva → detecta `sin_mapear` y `mal_mapeado` (SKU cruzado).
      B) Por vínculo guardado → detecta `huerfano` (el vínculo apunta a una
         publicación que el canal ya NO muestra viva: id cambiado, pausada o
         eliminada). Este es el segundo método independiente.

    Además calcula una RECONCILIACIÓN por canal que debe cuadrar por partida
    doble (vivas = mapeadas + sin_mapear) y un veredicto de COHERENCIA global:
    solo es verde si todos los canales respondieron y sus conteos cuadran sin
    pendientes. Así un "0 pendientes" queda respaldado, no asumido.
    NO escribe nada (eso lo hace _mapeo_persist)."""
    from collections import defaultdict as _dd
    items, ch_status = _fetch_channel_items()
    links = db.inv_get_links()
    prods = db._sb_get("inventory_products?select=id,code,name") or []
    id2code = {p["id"]: (p.get("code") or "") for p in prods}
    code2id = {(p.get("code") or "").strip().upper(): p["id"]
               for p in prods if p.get("code")}

    # Índice de publicaciones vivas por (canal, id externo).
    live, live_by_ch = {}, _dd(int)
    for it in items:
        ch = it.get("channel") or "mercadolibre"
        ext = it.get("item_id") or ""
        if not ext:
            continue
        live[(ch, ext)] = it
        live_by_ch[ch] += 1

    # Índice de vínculos guardados por (canal, id externo).
    linkmap, links_by_ch = {}, _dd(int)
    for l in links:
        ext = l.get("ml_item_id") or ""
        if not ext:
            continue
        ch = l.get("channel") or "mercadolibre"
        linkmap[(ch, ext)] = l
        links_by_ch[ch] += 1

    pend = []

    # ── A) Recorrido por publicación viva ──
    for (ch, ext), it in live.items():
        sku = (it.get("sku") or "").strip()
        motivo, detalle = None, ""
        if (ch, ext) not in linkmap:
            motivo = "sin_mapear"
            detalle = "La publicación no está vinculada a ningún SKU del inventario."
        else:
            linked_code = (id2code.get(linkmap[(ch, ext)].get("product_id"))
                           or "").strip().upper()
            if sku.upper() and sku.upper() in code2id and sku.upper() != linked_code:
                motivo = "mal_mapeado"
                detalle = ("SKU cruzado: la publicación declara «%s» pero está "
                           "vinculada a «%s»." % (sku, linked_code or "—"))
        if not motivo:
            continue
        sugerido = id2code.get(code2id.get(sku.upper())) if sku else ""
        pend.append({
            "channel": ch, "ext_id": ext, "title": it.get("title") or ext,
            "thumb": it.get("thumbnail") or "", "link": _mapeo_link(ch, ext, sku),
            "sku_canal": sku, "qty": int(it.get("inventory") or 0),
            "price": float(it.get("price") or 0), "motivo": motivo,
            "sugerido_code": sugerido or "", "detalle": detalle,
            "_meta": {
                "title": it.get("title") or "", "thumb": it.get("thumbnail") or "",
                "qty": it.get("inventory") or 0, "price": it.get("price") or 0,
                "logistic": it.get("logistic_type") or "",
                "inv_id": it.get("inventory_id") or "", "upid": it.get("upid") or "",
            },
        })

    # ── B) Recorrido por vínculo guardado → huérfanos ──
    # Solo se juzga en canales que respondieron OK (si la API no cargó, no se
    # puede afirmar que la publicación "ya no existe").
    for (ch, ext), l in linkmap.items():
        if not ch_status.get(ch, {}).get("ok"):
            continue
        if (ch, ext) in live:
            continue
        linked_code = id2code.get(l.get("product_id")) or ""
        pend.append({
            "channel": ch, "ext_id": ext,
            "title": l.get("ml_title") or ext,
            "thumb": l.get("ml_thumb") or "", "link": _mapeo_link(ch, ext, linked_code),
            "sku_canal": "", "qty": 0, "price": 0, "motivo": "huerfano",
            "sugerido_code": linked_code,
            "detalle": ("Vínculo huérfano: el inventario lo cree vinculado al SKU "
                        "«%s», pero el canal ya no muestra esa publicación viva "
                        "(id cambiado, pausada o eliminada)." % (linked_code or "—")),
            "_meta": {"title": l.get("ml_title") or "", "thumb": l.get("ml_thumb") or ""},
        })

    _orden = {"sin_mapear": 0, "mal_mapeado": 1, "huerfano": 2}
    pend.sort(key=lambda x: (x["channel"], _orden.get(x["motivo"], 9),
                             x["title"].lower()))

    # ── Reconciliación por canal (partida doble) ──
    recon = {}
    canales = set(list(live_by_ch) + list(links_by_ch) + list(ch_status))
    for ch in canales:
        st = ch_status.get(ch, {})
        respondio = bool(st.get("ok"))
        n_sin = sum(1 for p in pend if p["channel"] == ch and p["motivo"] == "sin_mapear")
        n_cru = sum(1 for p in pend if p["channel"] == ch and p["motivo"] == "mal_mapeado")
        n_hue = sum(1 for p in pend if p["channel"] == ch and p["motivo"] == "huerfano")
        vivas = live_by_ch.get(ch, 0)
        mapeadas = sum(1 for (c, e) in linkmap if c == ch and (c, e) in live)
        nlinks = links_by_ch.get(ch, 0)
        # Cuadre por partida doble cruzando las DOS fuentes:
        #   · feed vivo:  vivas == mapeadas + sin_mapear        (toda viva está o no mapeada)
        #   · tabla links: links == mapeadas + huérfanos         (todo vínculo apunta o no a una viva)
        # La 2ª es la no trivial: si no cuadra, hay vínculos duplicados o basura.
        cuadra = (vivas == mapeadas + n_sin) and (
            not respondio or nlinks == mapeadas + n_hue)
        ok_ch = respondio and cuadra and n_sin == 0 and n_cru == 0 and n_hue == 0
        recon[ch] = {
            "respondio": respondio, "error": st.get("error") or "",
            "vivas": vivas, "mapeadas": mapeadas, "links": links_by_ch.get(ch, 0),
            "sin_mapear": n_sin, "cruzados": n_cru, "huerfanos": n_hue,
            "cuadra": cuadra, "ok": ok_ch,
        }

    coherencia = bool(recon) and all(r["ok"] for r in recon.values())
    return {
        "pendientes": pend, "channels": ch_status, "reconciliacion": recon,
        "coherencia": coherencia,
        "n_sin_mapear": sum(1 for p in pend if p["motivo"] == "sin_mapear"),
        "n_mal_mapeado": sum(1 for p in pend if p["motivo"] == "mal_mapeado"),
        "n_huerfano": sum(1 for p in pend if p["motivo"] == "huerfano"),
        "generado": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")}


def _mapeo_persist(audit: dict):
    """Upsert del snapshot en mapeo_pendientes (por channel+ext_id) y marca como
    resueltas las que ya no aparecen. Best-effort (si la tabla no existe, no
    rompe: la sección sigue funcionando con la auditoría en vivo)."""
    try:
        now = datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")
        vivos = set()
        for p in audit.get("pendientes", []):
            vivos.add((p["channel"], p["ext_id"]))
            db._sb_post("mapeo_pendientes?on_conflict=channel,ext_id", {
                "channel": p["channel"], "ext_id": p["ext_id"],
                "title": p["title"], "thumb": p["thumb"], "link": p["link"],
                "sku_canal": p["sku_canal"], "qty": p["qty"], "price": p["price"],
                "motivo": p["motivo"], "sugerido_code": p["sugerido_code"],
                "detalle": p["detalle"], "visto_at": now, "resuelto": False,
                "resuelto_at": None, "resuelto_code": None}, upsert=True)
        prev = db._sb_get("mapeo_pendientes?resuelto=eq.false&"
                          "select=id,channel,ext_id") or []
        for r in prev:
            if (r.get("channel"), r.get("ext_id")) not in vivos:
                db._sb_patch("mapeo_pendientes", "id=eq.%d" % r["id"],
                             {"resuelto": True, "resuelto_at": now,
                              "resuelto_code": "(auto: ya mapeada)"})
    except Exception:
        pass


def _mapeo_run(force=False, report=True) -> dict:
    """Corre la auditoría (caché de 5 min), persiste y reporta a Cerebro."""
    with _MAPEO_LOCK:
        if (not force and _MAPEO_CACHE["data"]
                and (time.time() - _MAPEO_CACHE["ts"]) < _MAPEO_TTL):
            return _MAPEO_CACHE["data"]
        audit = _mapeo_audit()
        _mapeo_persist(audit)
        _MAPEO_CACHE["ts"] = time.time()
        _MAPEO_CACHE["data"] = audit
    if report:
        recon = audit.get("reconciliacion") or {}
        caidos = [c for c, r in recon.items() if not r.get("respondio")]
        # Resumen de cuadre por canal para el mensaje de Cerebro.
        ok_ch = sorted(c for c, r in recon.items() if r.get("ok"))
        if audit.get("coherencia"):
            _cerebro_set_heartbeat(
                _MAPEO_TASK_ID, "ok",
                "Coherencia verificada: cada publicación viva cuadra con su SKU "
                "(0 sin mapear · 0 cruzados · 0 huérfanos) en %s." % (
                    ", ".join(ok_ch) or "los canales activos"),
                nombre="Mapeo de SKU", canal="sistema", icon="bolt",
                cad="Diario · 8:00 PM", desc=_MAPEO_DESC)
        elif caidos:
            # No es verde porque un canal no respondió → verificación PARCIAL.
            _cerebro_set_heartbeat(
                _MAPEO_TASK_ID, "warn",
                "Verificación parcial: no respondió %s. %d pendiente(s) en lo "
                "auditado (%d sin mapear · %d cruzado · %d huérfano). Reintentar." % (
                    ", ".join(sorted(caidos)), len(audit["pendientes"]),
                    audit["n_sin_mapear"], audit["n_mal_mapeado"],
                    audit["n_huerfano"]),
                nombre="Mapeo de SKU", canal="sistema", icon="bolt",
                cad="Diario · 8:00 PM", desc=_MAPEO_DESC)
        else:
            _cerebro_set_heartbeat(
                _MAPEO_TASK_ID, "warn",
                "%d publicación(es) por resolver: %d sin mapear · %d SKU cruzado · "
                "%d huérfano. Revisa la sección Mapeo." % (
                    len(audit["pendientes"]), audit["n_sin_mapear"],
                    audit["n_mal_mapeado"], audit["n_huerfano"]),
                nombre="Mapeo de SKU", canal="sistema", icon="bolt",
                cad="Diario · 8:00 PM", desc=_MAPEO_DESC)
    return audit


@app.get("/api/mapeo")
def mapeo_get(force: int = 0, user: dict = Depends(_current_user)):
    """Reporte de publicaciones pendientes de mapear + productos para el
    desplegable. Visitar la sección audita los canales, persiste el snapshot y
    reporta a Cerebro. Cualquier usuario."""
    audit = _mapeo_run(force=bool(force))
    prods = db._sb_get("inventory_products?select=id,code,name&order=code.asc") or []
    # Foto por producto: la miniatura de su publicacion mas vendida (para el
    # selector con imagenes de la seccion Mapeo).
    thumb_by_pid = {}
    try:
        for l in (db._sb_get("inventory_links?select=product_id,ml_thumb,ml_sold") or []):
            th = l.get("ml_thumb") or ""
            if not th:
                continue
            pid = l.get("product_id"); sold = float(l.get("ml_sold") or 0)
            cur = thumb_by_pid.get(pid)
            if cur is None or sold > cur[0]:
                thumb_by_pid[pid] = (sold, th)
    except Exception:
        thumb_by_pid = {}
    productos = [{"id": p["id"], "code": p.get("code") or "",
                  "name": p.get("name") or "",
                  "thumb": (thumb_by_pid.get(p["id"]) or (0, ""))[1]} for p in prods]
    return {"ok": True, "generado": audit["generado"],
            "channels": audit["channels"], "reconciliacion": audit["reconciliacion"],
            "coherencia": audit["coherencia"], "n_sin_mapear": audit["n_sin_mapear"],
            "n_mal_mapeado": audit["n_mal_mapeado"], "n_huerfano": audit["n_huerfano"],
            "pendientes": audit["pendientes"], "productos": productos}


@app.get("/api/mapeo/count")
def mapeo_count(user: dict = Depends(_current_user)):
    """Pendientes sin resolver (badge del menú). Lee la tabla persistida; si no
    existe aún, devuelve 0."""
    try:
        rows = db._sb_get("mapeo_pendientes?resuelto=eq.false&select=id") or []
        return {"count": len(rows)}
    except Exception:
        return {"count": 0}


@app.post("/api/mapeo/scan")
def mapeo_scan(key: str = "", authorization: Optional[str] = Header(None)):
    """Fuerza la auditoría, persiste y reporta a Cerebro. Auth: ?key=
    BOUN_EXPORT_TOKEN (planificador sin sesión) o sesión admin. La usa la skill
    diaria de mapeo si corre sin abrir la web."""
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    authed = bool(token and key == token)
    if not authed:
        try:
            u = _current_user(authorization)
            authed = bool(u and u.get("role") == "admin")
        except Exception:
            authed = False
    if not authed:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    audit = _mapeo_run(force=True)
    return {"ok": True, "generado": audit["generado"], "coherencia": audit["coherencia"],
            "total": len(audit["pendientes"]), "n_sin_mapear": audit["n_sin_mapear"],
            "n_mal_mapeado": audit["n_mal_mapeado"], "n_huerfano": audit["n_huerfano"],
            "reconciliacion": audit["reconciliacion"], "channels": audit["channels"]}


class MapeoAsociarIn(BaseModel):
    channel: str
    ext_id: str
    product_id: int
    title: Optional[str] = ""
    thumb: Optional[str] = ""
    qty: Optional[float] = 0
    price: Optional[float] = 0
    logistic: Optional[str] = ""
    inv_id: Optional[str] = ""
    upid: Optional[str] = ""


@app.post("/api/mapeo/asociar")
def mapeo_asociar(data: MapeoAsociarIn, user: dict = Depends(_current_user)):
    """Asocia UNA publicación pendiente al SKU elegido y la marca como resuelta."""
    ch = (data.channel or "").strip()
    ext = (data.ext_id or "").strip()
    if not ext or not data.product_id:
        raise HTTPException(400, "Faltan channel/ext_id/product_id")
    prows = db._sb_get("inventory_products?id=eq.%d&select=id,code"
                       % data.product_id) or []
    if not prows:
        raise HTTPException(404, "Producto no existe")
    ok = db.inv_link_add(data.product_id, ch, ext, {
        "title": data.title, "thumb": data.thumb, "qty": data.qty,
        "price": data.price, "logistic": data.logistic,
        "inv_id": data.inv_id, "upid": data.upid})
    if not ok:
        raise HTTPException(400, "No se pudo crear el vínculo")
    code = prows[0].get("code") or ""
    try:
        db._sb_patch("mapeo_pendientes",
                     "channel=eq.%s&ext_id=eq.%s" % (_q_(ch), _q_(ext)),
                     {"resuelto": True, "resuelto_code": code,
                      "resuelto_at": datetime.now(timezone.utc).strftime(
                          "%Y-%m-%dT%H:%M:%SZ")})
    except Exception:
        pass
    _MAPEO_CACHE["ts"] = 0   # invalida la caché para reflejar el cambio
    return {"ok": True, "channel": ch, "ext_id": ext, "code": code}


class MapeoDesvincularIn(BaseModel):
    channel: str
    ext_id: str


@app.post("/api/mapeo/desvincular")
def mapeo_desvincular(data: MapeoDesvincularIn, user: dict = Depends(_admin)):
    """Quita un vínculo huérfano (apunta a una publicación que el canal ya no
    muestra viva). Solo admin: borra la fila de inventory_links."""
    ch = (data.channel or "").strip()
    ext = (data.ext_id or "").strip()
    if not ext:
        raise HTTPException(400, "Falta ext_id")
    ok = db.inv_link_delete(ch, ext)
    if not ok:
        raise HTTPException(400, "No se pudo quitar el vínculo")
    try:
        db._sb_patch("mapeo_pendientes",
                     "channel=eq.%s&ext_id=eq.%s" % (_q_(ch), _q_(ext)),
                     {"resuelto": True, "resuelto_code": "(huérfano quitado)",
                      "resuelto_at": datetime.now(timezone.utc).strftime(
                          "%Y-%m-%dT%H:%M:%SZ")})
    except Exception:
        pass
    _MAPEO_CACHE["ts"] = 0
    return {"ok": True, "channel": ch, "ext_id": ext}


# ── DENUNCIAS — Brand Protection Program (protección de marca BOUN) ───────────
# La skill diaria "detectar-denunciar-marca" (8:00 PM) busca "boun" en el BPP,
# detecta catálogos Marca:BOUN donde participa un vendedor ≠ BOUN COLOMBIA, los
# denuncia (marca registrada → producto falsificado) y los envía; luego revisa
# el estado de las denuncias de días anteriores. Cada corrida hace POST a
# /api/denuncias/report con el snapshot; esta sección lo persiste en `denuncias`
# (upsert por seller_nick+catalog_id), arma la traza en `historial`, y reporta a
# Cerebro. La sección /denuncias de la web lo muestra.

_DENUNCIAS_TASK_ID = "denuncias-diario"
_DENUNCIAS_DESC = ("Protege la marca registrada BOUN: detecta vendedores que se "
                   "cuelgan de tus catálogos y los denuncia en el Brand Protection "
                   "Program de MercadoLibre, y hace seguimiento del estado.")
# Estados que cierran el caso (la publicación cayó o ML resolvió).
_DENUNCIA_ESTADOS_RESUELTO = {"procedente", "rechazada", "publicacion_inactiva"}


def _denuncia_now() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


class DenunciaIn(BaseModel):
    seller_nick: str
    catalog_id: str
    seller_link: Optional[str] = ""
    pub_id: Optional[str] = ""
    pub_title: Optional[str] = ""
    pub_link: Optional[str] = ""
    pub_price: Optional[float] = 0
    thumb: Optional[str] = ""
    catalog_title: Optional[str] = ""
    catalog_link: Optional[str] = ""
    motivo: Optional[str] = "marca_registrada"
    tipo_infraccion: Optional[str] = "Es un producto falsificado"
    texto: Optional[str] = ""
    estado: Optional[str] = "en_proceso"
    nota: Optional[str] = ""           # nota libre para el historial de esta corrida


class DenunciasReportIn(BaseModel):
    denuncias: List[DenunciaIn] = []   # snapshot de la corrida (nuevas + revisadas)
    resumen: Optional[dict] = None     # {nuevas, en_proceso, procedentes, ...} opcional


def _denuncia_upsert(d: dict) -> dict:
    """Upsert de UNA denuncia por (seller_nick, catalog_id). Si ya existe y cambió
    el estado, agrega una entrada al historial. Best-effort."""
    nick = (d.get("seller_nick") or "").strip()
    cat = (d.get("catalog_id") or "").strip()
    if not nick or not cat:
        return {"ok": False, "error": "faltan seller_nick/catalog_id"}
    now = _denuncia_now()
    estado = (d.get("estado") or "en_proceso").strip()
    pub_link = (d.get("pub_link") or "").strip() or _permalink_from_mco(d.get("pub_id") or "")
    cat_link = (d.get("catalog_link") or "").strip() or _permalink_from_mco(cat)
    try:
        prev = db._sb_get("denuncias?seller_nick=eq.%s&catalog_id=eq.%s&select=*"
                          % (_q_(nick), _q_(cat))) or []
    except Exception:
        prev = []
    hist_nota = (d.get("nota") or "").strip()
    if prev:
        row = prev[0]
        hist = row.get("historial") or []
        if not isinstance(hist, list):
            hist = []
        cambio_estado = (row.get("estado") != estado)
        if cambio_estado or hist_nota:
            hist.append({"fecha": now, "estado": estado,
                         "nota": hist_nota or ("Estado: %s" % estado)})
        patch = {
            "seller_link": d.get("seller_link") or row.get("seller_link"),
            "pub_id": d.get("pub_id") or row.get("pub_id"),
            "pub_title": d.get("pub_title") or row.get("pub_title"),
            "pub_link": pub_link or row.get("pub_link"),
            "pub_price": d.get("pub_price") or row.get("pub_price") or 0,
            "thumb": d.get("thumb") or row.get("thumb"),
            "catalog_title": d.get("catalog_title") or row.get("catalog_title"),
            "catalog_link": cat_link or row.get("catalog_link"),
            "motivo": d.get("motivo") or row.get("motivo") or "marca_registrada",
            "tipo_infraccion": d.get("tipo_infraccion") or row.get("tipo_infraccion"),
            "texto": d.get("texto") or row.get("texto"),
            "estado": estado, "revisado_at": now, "historial": hist,
            "resuelto": estado in _DENUNCIA_ESTADOS_RESUELTO,
            "resuelto_at": now if estado in _DENUNCIA_ESTADOS_RESUELTO else row.get("resuelto_at"),
        }
        db._sb_patch("denuncias", "id=eq.%d" % row["id"], patch)
        return {"ok": True, "id": row["id"], "nuevo": False, "cambio_estado": cambio_estado}
    # Nueva denuncia
    hist = [{"fecha": now, "estado": estado,
             "nota": hist_nota or "Denuncia presentada en el BPP."}]
    db._sb_post("denuncias", {
        "seller_nick": nick, "seller_link": d.get("seller_link") or "",
        "pub_id": d.get("pub_id") or "", "pub_title": d.get("pub_title") or "",
        "pub_link": pub_link, "pub_price": d.get("pub_price") or 0,
        "thumb": d.get("thumb") or "", "catalog_id": cat,
        "catalog_title": d.get("catalog_title") or "", "catalog_link": cat_link,
        "motivo": d.get("motivo") or "marca_registrada",
        "tipo_infraccion": d.get("tipo_infraccion") or "Es un producto falsificado",
        "texto": d.get("texto") or "", "estado": estado,
        "denunciado_at": now, "revisado_at": now,
        "resuelto": estado in _DENUNCIA_ESTADOS_RESUELTO,
        "resuelto_at": now if estado in _DENUNCIA_ESTADOS_RESUELTO else None,
        "historial": hist})
    return {"ok": True, "nuevo": True}


def _denuncias_counts() -> dict:
    try:
        rows = db._sb_get("denuncias?select=estado,resuelto") or []
    except Exception:
        rows = []
    c = {"total": len(rows), "activas": 0, "en_proceso": 0, "pendiente": 0,
         "procedente": 0, "rechazada": 0, "publicacion_inactiva": 0}
    for r in rows:
        e = r.get("estado") or ""
        c[e] = c.get(e, 0) + 1
        if not r.get("resuelto"):
            c["activas"] += 1
    return c


@app.get("/api/denuncias")
def denuncias_get(user: dict = Depends(_current_user)):
    """Lista de denuncias para la sección /denuncias (más recientes primero)."""
    try:
        rows = db._sb_get("denuncias?select=*&order=denunciado_at.desc") or []
    except Exception:
        rows = []
    return {"ok": True, "generado": _denuncia_now(),
            "counts": _denuncias_counts(), "denuncias": rows}


@app.get("/api/denuncias/count")
def denuncias_count(user: dict = Depends(_current_user)):
    """Denuncias activas (en proceso / pendientes) para el badge del menú."""
    try:
        return {"count": _denuncias_counts().get("activas", 0)}
    except Exception:
        return {"count": 0}


@app.post("/api/denuncias/report")
def denuncias_report(data: DenunciasReportIn, key: str = "",
                     authorization: Optional[str] = Header(None)):
    """Recibe el snapshot diario de la skill y persiste cada denuncia (upsert por
    seller_nick+catalog_id, con traza en historial). Reporta a Cerebro.
    Auth: ?key=BOUN_EXPORT_TOKEN (planificador sin sesión) o sesión admin."""
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    authed = bool(token and key == token)
    if not authed:
        try:
            u = _current_user(authorization)
            authed = bool(u and u.get("role") == "admin")
        except Exception:
            authed = False
    if not authed:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    nuevas = cambios = 0
    errores = []
    for d in (data.denuncias or []):
        try:
            r = _denuncia_upsert(d.dict())
            if r.get("ok"):
                if r.get("nuevo"):
                    nuevas += 1
                elif r.get("cambio_estado"):
                    cambios += 1
            else:
                errores.append(r.get("error"))
        except Exception as e:
            errores.append(str(e)[:120])
    c = _denuncias_counts()
    # Heartbeat a Cerebro: resumen del día y seguimiento de las activas.
    msg = ("%d denuncia(s) presentada(s) hoy · %d con cambio de estado · "
           "%d activa(s) en seguimiento (%d en proceso · %d procedentes · "
           "%d publicaciones caídas)." % (
               nuevas, cambios, c.get("activas", 0), c.get("en_proceso", 0),
               c.get("procedente", 0), c.get("publicacion_inactiva", 0)))
    _cerebro_set_heartbeat(_DENUNCIAS_TASK_ID, "ok", msg,
                           nombre="Protección de marca (denuncias)",
                           canal="mercadolibre", icon="alert",
                           cad="Diario · 8:00 PM", desc=_DENUNCIAS_DESC)
    return {"ok": True, "nuevas": nuevas, "cambios": cambios,
            "counts": c, "errores": [e for e in errores if e]}


# ── MARÍA JOSÉ — Liquidación de productos propios ────────────────────────────
# Separa de las ventas totales lo que corresponde a los productos propios de
# María José (inventory_products.owner = 'MARIA_JOSE'), descuenta los costos
# reales de venta (comisión real + retención, envío y publicidad real por ítem)
# y lleva el SALDO que se le debe menos los abonos ya pagados, además de CUÁNDO
# libera cada plataforma el dinero.  Persiste en mj_ventas / mj_abonos.

_MJ_TASK_ID = "maria-jose-liquidacion"
_MJ_DESC = ("Liquidación de los productos propios de María José: separa sus "
            "ventas de cada plataforma, descuenta los costos reales (comisión, "
            "envío y publicidad) y lleva el saldo a pagar y la liberación del "
            "dinero por plataforma.")
_MJ_CACHE = {"ts": 0.0}
_MJ_TTL = 10 * 60
_MJ_LOCK = threading.Lock()


def _mj_setting_f(key, default):
    try:
        return float(db.get_setting(key, "") or default)
    except Exception:
        return float(default)


def _mj_now():
    return datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")


def _mj_targets() -> dict:
    """Objetivos de María José: publicaciones e identificadores por canal.

    Devuelve {ml:{item_id:info}, fa:{sku:info}, codes:{CODIGO:info}, has:bool}
    donde info = {product_id, code, name, thumb}.  Las ventas se atribuyen a MJ
    por item_id (ML), por SellerSku (Falabella) o por SKU/código BOUN (Shopify).
    """
    try:
        prods = db._sb_get("inventory_products?owner=eq.MARIA_JOSE"
                           "&select=id,code,name,mj_qty,mj_anchor") or []
    except Exception:
        prods = []
    by_id = {p["id"]: p for p in prods}
    codes, ml, fa, meta = {}, {}, {}, {}
    for p in prods:
        # mj_qty <= 0 o null = "todas las unidades" (sin tope).
        try:
            q = float(p.get("mj_qty")) if p.get("mj_qty") is not None else None
        except (TypeError, ValueError):
            q = None
        meta[p["id"]] = {"qty": (q if (q is not None and q > 0) else None),
                         "anchor": (str(p.get("mj_anchor"))[:10]
                                    if p.get("mj_anchor") else None),
                         "code": p.get("code", ""), "name": p.get("name", "")}
        c = str(p.get("code") or "").strip().upper()
        if c:
            codes[c] = {"product_id": p["id"], "code": p.get("code", ""),
                        "name": p.get("name", ""), "thumb": ""}
    try:
        links = db.inv_get_links() or []
    except Exception:
        links = []
    for l in links:
        pid = l.get("product_id")
        if pid not in by_id:
            continue
        ext = str(l.get("ml_item_id") or "")
        if not ext:
            continue
        ch = l.get("channel") or "mercadolibre"
        info = {"product_id": pid, "code": by_id[pid].get("code", ""),
                "name": l.get("ml_title") or by_id[pid].get("name", ""),
                "thumb": l.get("ml_thumb") or ""}
        if ch == "mercadolibre":
            ml[ext] = info
        elif ch == "falabella":
            fa[ext] = info
        c = str(by_id[pid].get("code") or "").strip().upper()
        if c in codes and not codes[c]["thumb"] and info["thumb"]:
            codes[c]["thumb"] = info["thumb"]
    return {"ml": ml, "fa": fa, "codes": codes, "meta": meta,
            "has": bool(prods)}


def _mj_ml_ad_cost(s, item_ids, since_d, until_d) -> dict:
    """{item_id → gasto REAL de Product Ads (Product Ads) en el rango}.

    Usa el reporte de ítems del anunciante:
      /advertising/advertisers/{adv}/product_ads/items?metrics=cost  (paginado).
    Devuelve el costo de TODOS los ítems anunciados (no solo los links de MJ),
    porque la publicación vendida puede diferir del link actual. La API de Ads
    sólo admite rangos <= 90 días, así que se capa.  Header Api-Version:1 SOLO
    en estas peticiones (no contaminar orders/search ni el detalle de orden)."""
    out = {}
    try:
        from ml_scraper import ML_API
        import datetime as _dt
        _adh = {"Api-Version": "1"}
        a = s.get(f"{ML_API}/advertising/advertisers?product_id=PADS",
                  headers=_adh, timeout=12)
        adv_id = None
        if a.status_code == 200:
            arr = a.json().get("advertisers", [])
            if arr:
                adv_id = arr[0].get("advertiser_id")
        if not adv_id:
            return out
        # Rango <= 90 días (la API rechaza más).
        try:
            d2 = _dt.date.fromisoformat(until_d)
            d1 = _dt.date.fromisoformat(since_d)
        except Exception:
            d2 = _dt.date.today(); d1 = d2 - _dt.timedelta(days=89)
        if (d2 - d1).days > 89:
            d1 = d2 - _dt.timedelta(days=89)
        df, dtt = d1.isoformat(), d2.isoformat()
        offset = 0
        while True:
            r = s.get(f"{ML_API}/advertising/advertisers/{adv_id}/product_ads/"
                      f"items?date_from={df}&date_to={dtt}&metrics=cost"
                      f"&limit=50&offset={offset}", headers=_adh, timeout=25)
            if r.status_code != 200:
                break
            j = r.json()
            res = j.get("results", []) or []
            for it in res:
                iid = it.get("item_id")
                c = float((it.get("metrics") or {}).get("cost") or 0)
                if iid and c:
                    out[str(iid)] = out.get(str(iid), 0.0) + c
            total = j.get("paging", {}).get("total", 0)
            offset += 50
            if offset >= total or not res:
                break
    except Exception:
        pass
    return out


def _mj_ml_release(s, oid, od, co):
    """Fecha en que ML libera el dinero de una orden y si YA está liberado.

    OJO: ni /orders ni /orders/{id} traen money_release_date.  El dato vive en
    el pago.  El endpoint de MercadoPago (/v1/payments/{id}) devuelve la fecha
    Y el estado real `money_release_status` ('released' = ya disponible); el de
    ML /collections/{id} da la fecha pero deja el estado en null.  Se usa MP
    primero (fecha + estado) y /collections como respaldo de fecha.
    Devuelve (fecha|None, liberado_bool)."""
    import datetime as _dt
    from ml_scraper import ML_API
    pids = [p.get("id") for p in (od.get("payments") or []) if p.get("id")]
    if not pids:
        try:
            r = s.get(f"{ML_API}/orders/{oid}", timeout=12)
            if r.status_code == 200:
                pids = [p.get("id") for p in (r.json().get("payments") or [])
                        if p.get("id")]
        except Exception:
            pass
    best, released = None, False
    for pid in pids:
        mr, st = "", ""
        try:
            r = s.get(f"https://api.mercadopago.com/v1/payments/{pid}",
                      timeout=12)
            if r.status_code == 200:
                b = r.json()
                mr = b.get("money_release_date") or ""
                st = b.get("money_release_status") or ""
        except Exception:
            pass
        if not mr:
            try:
                r = s.get(f"{ML_API}/collections/{pid}", timeout=12)
                if r.status_code == 200:
                    b = r.json()
                    mr = b.get("money_release_date") or ""
                    st = st or (b.get("money_release_status") or "")
            except Exception:
                pass
        if st == "released":
            released = True
        if mr:
            try:
                # Fecha TAL CUAL la reporta ML/MP (su propia zona, -04:00), sin
                # convertir a Colombia: convertir corría la fecha 1 día (una
                # liberación a las 00:02 -04:00 caía el día anterior en -05:00),
                # haciéndola no coincidir con lo que ML le muestra al vendedor.
                d = _dt.date.fromisoformat(mr[:10])
                if best is None or d > best:
                    best = d
            except Exception:
                pass
    return best, released


def _mj_ml_shipping(s, shipment_id) -> float:
    """Costo de envío que asume el vendedor para un envío de ML (best-effort)."""
    if not shipment_id:
        return 0.0
    try:
        from ml_scraper import ML_API
        r = s.get(f"{ML_API}/shipments/{shipment_id}/costs", timeout=12)
        if r.status_code == 200:
            j = r.json()
            tot = 0.0
            for sd in (j.get("senders") or []):
                tot += float(sd.get("cost") or 0)
                tot -= float(sd.get("compensation") or 0)
            return round(max(tot, 0.0), 2)
    except Exception:
        pass
    return 0.0


def _mj_upsert_row(row: dict) -> bool:
    """Upsert idempotente de una fila de mj_ventas por (plataforma, orden, ítem)."""
    flt = ("mj_ventas?plataforma=eq.%s&order_id=eq.%s&item_id=eq.%s&select=id"
           % (_q_(row["plataforma"]), _q_(str(row["order_id"])),
              _q_(str(row.get("item_id") or ""))))
    try:
        prev = db._sb_get(flt) or []
    except Exception:
        prev = []
    payload = dict(row)
    payload["updated_at"] = _mj_now()
    if prev:
        return db._sb_patch("mj_ventas", "id=eq.%d" % prev[0]["id"], payload)
    return bool(db._sb_post("mj_ventas", payload))


def _mj_sync(window_days=None) -> dict:
    """Recolecta las ventas de los productos de María José en los 3 canales,
    calcula los costos reales y hace upsert en mj_ventas.  Devuelve resumen."""
    from config import RETENCION_FUENTE
    import datetime as _dt
    tg = _mj_targets()
    # Purga de huérfanas: borra de la liquidación las ventas de productos que YA
    # no están marcados como de María José (p. ej. se desmarcaron). Así la
    # sección refleja exactamente los productos marcados ahora.
    try:
        mj_pids = [int(p) for p in (tg.get("meta", {}) or {}).keys()]
        if mj_pids:
            db._sb_delete("mj_ventas", "product_id=not.in.(%s)"
                          % ",".join(str(p) for p in mj_pids))
        else:
            db._sb_delete("mj_ventas", "id=gte.0")   # ninguno marcado → limpiar todo
    except Exception:
        pass
    if not tg["has"]:
        return {"ok": True, "n": 0,
                "note": "Aún no hay productos marcados como de María José."}
    co = _dt.timezone(_dt.timedelta(hours=-5))
    wd = int(window_days or _mj_setting_f("mj_window_days", 120))
    # Extender la ventana para cubrir desde la marca (anchor) más antigua de un
    # producto con cupo, así no se subcuenta lo vendido desde que se asignó.
    try:
        anchors = [m["anchor"] for m in tg.get("meta", {}).values()
                   if m.get("qty") and m.get("anchor")]
        if anchors:
            old = min(anchors)
            days_since = (_dt.date.today() - _dt.date.fromisoformat(old)).days
            wd = min(max(wd, days_since + 2), 400)
    except Exception:
        pass
    to_d = _dt.datetime.now(co)
    from_d = (to_d - _dt.timedelta(days=wd)).replace(
        hour=0, minute=0, second=0, microsecond=0)
    since = from_d.strftime("%Y-%m-%dT%H:%M:%S.000-05:00")
    until = to_d.strftime("%Y-%m-%dT%H:%M:%S.000-05:00")
    since_d, until_d = from_d.date().isoformat(), to_d.date().isoformat()
    today = to_d.date()
    rows, errores = [], []

    # ── MercadoLibre (comisión real = sale_fee · liberación = money_release_date) ──
    ml_set = tg["ml"]
    if ml_set:
        try:
            from ml_scraper import _ml_session_auth, ML_API as _MLAPI
            s, uid = _ml_session_auth()
            if not s:
                errores.append("ML sin conexión")
            else:
                # Asegurar versión por defecto (otras secciones pueden dejar
                # "Api-Version: 1" pegado y eso oculta money_release_date).
                try:
                    s.headers.pop("Api-Version", None)
                except Exception:
                    pass
                ad_cost = _mj_ml_ad_cost(s, list(ml_set.keys()),
                                         since_d, until_d)
                _NO = {"cancelled", "invalid", "payment_required",
                       "payment_in_process"}
                tmp, units_by_item = [], {}
                offset, seen = 0, set()
                while True:
                    r = s.get(f"{_MLAPI}/orders/search?seller={uid}"
                              f"&order.date_created.from={since}"
                              f"&order.date_created.to={until}"
                              f"&sort=date_desc&limit=50&offset={offset}",
                              timeout=20)
                    if r.status_code != 200:
                        break
                    d = r.json()
                    res = d.get("results", [])
                    for od in res:
                        oid = str(od.get("id") or "")
                        if not oid or oid in seen:
                            continue
                        seen.add(oid)
                        if (od.get("status") or "") in _NO:
                            continue
                        oi_mj = [oi for oi in od.get("order_items", [])
                                 if str((oi.get("item") or {}).get("id")
                                        or "") in ml_set]
                        if not oi_mj:
                            continue
                        dc = od.get("date_created") or ""
                        try:
                            fecha = _dt.datetime.fromisoformat(
                                dc.replace("Z", "+00:00")).astimezone(
                                co).date()
                        except Exception:
                            fecha = today
                        rel, rel_done = _mj_ml_release(s, oid, od, co)
                        ship_id = (od.get("shipping") or {}).get("id")
                        env = _mj_ml_shipping(s, ship_id)
                        tot_u = sum(int(oi.get("quantity") or 0)
                                    for oi in oi_mj) or 1
                        for oi in oi_mj:
                            it = oi.get("item") or {}
                            iid = str(it.get("id"))
                            q = int(oi.get("quantity") or 0)
                            unit = float(oi.get("unit_price") or 0)
                            gross = unit * q
                            fee = float(oi.get("sale_fee") or 0) * q
                            units_by_item[iid] = units_by_item.get(iid, 0) + q
                            tmp.append({
                                "iid": iid, "oid": oid, "q": q, "gross": gross,
                                "fee": fee, "ret": gross * RETENCION_FUENTE,
                                "env": env * (q / tot_u), "fecha": fecha,
                                "rel": rel, "rel_done": rel_done,
                                "info": ml_set.get(iid, {}),
                                "nombre": (it.get("title")
                                           or ml_set.get(iid, {}).get(
                                               "name", ""))})
                    total = d.get("paging", {}).get("total", 0)
                    offset += 50
                    if offset >= total or not res:
                        break
                for t in tmp:
                    iid = t["iid"]
                    pub = 0.0
                    tot_ad = ad_cost.get(iid, 0.0)
                    if tot_ad and units_by_item.get(iid):
                        pub = tot_ad * (t["q"] / units_by_item[iid])
                    neto = t["gross"] - t["fee"] - t["ret"] - t["env"] - pub
                    rel = t["rel"]
                    libre = bool(t.get("rel_done") or (rel and rel <= today))
                    rows.append({
                        "plataforma": "mercadolibre", "order_id": t["oid"],
                        "item_id": iid,
                        "product_id": t["info"].get("product_id"),
                        "codigo": t["info"].get("code", ""),
                        "nombre": t["nombre"], "thumb": t["info"].get("thumb", ""),
                        "unidades": t["q"],
                        "fecha_venta": t["fecha"].isoformat(),
                        "precio_venta": round(t["gross"], 2), "descuentos": 0,
                        "comision": round(t["fee"], 2),
                        "retencion": round(t["ret"], 2),
                        "costo_envio": round(t["env"], 2),
                        "costo_publicidad": round(pub, 2),
                        "neto_mj": round(neto, 2),
                        "release_date": (rel.isoformat() if rel else None),
                        "liberado": libre,
                        "estado_pago": "liberado" if libre else "pendiente"})
        except Exception as e:
            errores.append("ML: %s" % str(e)[:90])

    # ── Falabella (comisión y liberación estimadas por configuración) ──
    fa_set = tg["fa"]
    if fa_set or tg["codes"]:
        try:
            import falabella as fb
            fee_pct = _mj_setting_f("mj_falabella_fee_pct", 0.18)
            rel_days = int(_mj_setting_f("mj_falabella_release_days", 30))
            orders = fb.get_orders(from_d.isoformat())
            oids = [str(o.get("OrderId")) for o in orders if o.get("OrderId")]
            items = fb._all_order_items(oids) if oids else []
            agg = {}
            for it in items:
                sku = str(it.get("SellerSku") or it.get("Sku") or "")
                info = fa_set.get(sku) or tg["codes"].get(sku.strip().upper())
                if not info:
                    continue
                oid = str(it.get("OrderId") or "")
                cre = it.get("CreatedAt") or ""
                key = (oid, sku)
                a = agg.setdefault(key, {"oid": oid, "sku": sku, "q": 0,
                                         "gross": 0.0, "cre": cre,
                                         "info": info,
                                         "name": it.get("Name", "")})
                a["q"] += 1
                try:
                    a["gross"] += float(it.get("PaidPrice")
                                        or it.get("ItemPrice") or 0)
                except (TypeError, ValueError):
                    pass
            for (oid, sku), a in agg.items():
                try:
                    fv = _dt.datetime.fromisoformat(
                        a["cre"].replace("Z", "+00:00")).astimezone(co).date()
                except Exception:
                    fv = today
                fee = a["gross"] * fee_pct
                ret = a["gross"] * RETENCION_FUENTE
                neto = a["gross"] - fee - ret
                rel = fv + _dt.timedelta(days=rel_days)
                libre = rel <= today
                rows.append({
                    "plataforma": "falabella", "order_id": oid,
                    "item_id": sku, "product_id": a["info"].get("product_id"),
                    "codigo": a["info"].get("code", ""),
                    "nombre": a["name"] or a["info"].get("name", ""),
                    "thumb": a["info"].get("thumb", ""), "unidades": a["q"],
                    "fecha_venta": fv.isoformat(),
                    "precio_venta": round(a["gross"], 2), "descuentos": 0,
                    "comision": round(fee, 2), "retencion": round(ret, 2),
                    "costo_envio": 0, "costo_publicidad": 0,
                    "neto_mj": round(neto, 2), "release_date": rel.isoformat(),
                    "liberado": libre,
                    "estado_pago": "liberado" if libre else "pendiente"})
        except Exception as e:
            errores.append("Falabella: %s" % str(e)[:90])

    # ── Shopify (tienda propia: se atribuye por SKU = código BOUN) ──
    if tg["codes"]:
        fee_pct = _mj_setting_f("mj_shopify_fee_pct", 0.0)
        rel_days = int(_mj_setting_f("mj_shopify_release_days", 0))
        for ch, shop in _SHOPIFY_SHOPS.items():
            tok = db.get_setting("shopify_token::%s" % shop, "")
            if not tok:
                continue
            try:
                orders = _shopify_orders(shop, tok, from_d.isoformat(),
                                         to_d.isoformat())
            except Exception as e:
                errores.append("%s: %s" % (ch, str(e)[:70]))
                continue
            for od in orders:
                ca = od.get("created_at") or ""
                try:
                    fv = _dt.datetime.fromisoformat(
                        ca.replace("Z", "+00:00")).astimezone(co).date()
                except Exception:
                    fv = today
                for li in od.get("line_items", []):
                    sku = str(li.get("sku") or "").strip().upper()
                    info = tg["codes"].get(sku)
                    if not info:
                        continue
                    q = int(li.get("quantity") or 0)
                    gross = float(li.get("price") or 0) * q
                    fee = gross * fee_pct
                    neto = gross - fee
                    rel = fv + _dt.timedelta(days=rel_days)
                    libre = rel <= today
                    rows.append({
                        "plataforma": ch, "order_id": str(od.get("id")),
                        "item_id": str(li.get("id")),
                        "product_id": info.get("product_id"),
                        "codigo": info.get("code", ""),
                        "nombre": li.get("title") or info.get("name", ""),
                        "thumb": info.get("thumb", ""), "unidades": q,
                        "fecha_venta": fv.isoformat(),
                        "precio_venta": round(gross, 2), "descuentos": 0,
                        "comision": round(fee, 2), "retencion": 0,
                        "costo_envio": 0, "costo_publicidad": 0,
                        "neto_mj": round(neto, 2), "release_date": rel.isoformat(),
                        "liberado": libre,
                        "estado_pago": "liberado" if libre else "pendiente"})

    # ── Asignación por cupo (productos compartidos María José / BOUN) ──
    # María vende PRIMERO sus unidades; cuando se agotan, las siguientes ventas
    # son de BOUN.  Para los productos con cupo (mj_qty) se ordenan sus ventas
    # desde la marca (anchor) de más antigua a más nueva y se le asignan hasta
    # completar el cupo; la venta que cruza el tope se reparte proporcional.
    # Al consumir el cupo, el producto se DESMARCA solo (owner → BOUN).
    meta = tg.get("meta", {})
    from collections import defaultdict as _dd
    by_prod = _dd(list)
    for r in rows:
        by_prod[r.get("product_id")].append(r)
    final_rows, consumed_map, clear_pids = [], {}, []
    _MONEY_K = ("precio_venta", "descuentos", "comision", "retencion",
                "costo_envio", "costo_publicidad", "neto_mj")
    for pid, rs in by_prod.items():
        m = meta.get(pid, {})
        qty = m.get("qty")
        if not qty:                      # "todas las unidades" (sin tope)
            consumed_map[pid] = sum(float(x.get("unidades") or 0) for x in rs)
            final_rows.extend(rs)
            continue
        anchor = m.get("anchor")
        elig = [x for x in rs if (not anchor or x.get("fecha_venta", "") >= anchor)]
        elig.sort(key=lambda x: (x.get("fecha_venta", ""), str(x.get("order_id"))))
        rem, consumed = float(qty), 0.0
        for x in elig:
            if rem <= 0:
                break
            u = float(x.get("unidades") or 0)
            if u <= 0:
                continue
            if u <= rem:
                final_rows.append(x)
                consumed += u
                rem -= u
            else:                         # esta venta cruza el tope → repartir
                frac = rem / u
                xx = dict(x)
                xx["unidades"] = round(rem, 2)
                for kk in _MONEY_K:
                    xx[kk] = round(float(x.get(kk) or 0) * frac, 2)
                final_rows.append(xx)
                consumed += rem
                rem = 0
        consumed_map[pid] = consumed
        if consumed >= float(qty):
            clear_pids.append(pid)

    n = 0
    for row in final_rows:
        try:
            if _mj_upsert_row(row):
                n += 1
        except Exception:
            pass

    # Guardar lo consumido por producto. IMPORTANTE: cuando se agota el cupo NO
    # se cambia el dueño a BOUN. El producto sigue marcado como de María José
    # (queda "agotado", 0 restantes) por dos motivos:
    #   1) sus ventas pasadas YA contadas siguen siendo de ella (se le deben), y
    #      si lo desmarcáramos la limpieza de huérfanas borraría esas ventas;
    #   2) como el cupo está lleno, las ventas NUEVAS ya no se le atribuyen y
    #      pasan a BOUN automáticamente. El desmarcado real lo hace Sebastián a
    #      mano con el botón «MJ» (eso sí borra sus ventas de la liquidación).
    for pid, cons in consumed_map.items():
        try:
            db._sb_patch("inventory_products", "id=eq.%s" % pid,
                         {"mj_consumed": round(cons, 2)})
        except Exception:
            pass

    _MJ_CACHE["ts"] = time.time()
    return {"ok": True, "n": n, "filas": len(final_rows),
            "agotados": len(clear_pids),
            "errores": [e for e in errores if e]}


def _mj_summary(date_from: str = None, date_to: str = None) -> dict:
    """KPIs + ventas + abonos para la sección /maria-jose (lee de Supabase).

    Los KPIs de deuda (saldo, abonos) son SIEMPRE globales (toda la historia).
    Si se pasa date_from/date_to, la TABLA de ventas, el resumen por plataforma
    y el subtotal `periodo` se filtran por fecha_venta; los costos/neto del
    encabezado también reflejan el periodo filtrado."""
    try:
        ventas = db._sb_get("mj_ventas?select=*&order=fecha_venta.desc") or []
    except Exception:
        ventas = []
    try:
        abonos = db._sb_get("mj_abonos?select=*&order=fecha.desc") or []
    except Exception:
        abonos = []
    df = (date_from or "")[:10] or None
    dt = (date_to or "")[:10] or None
    if df and dt and dt < df:
        df, dt = dt, df

    def _in_range(v):
        f = (v.get("fecha_venta") or "")[:10]
        if df and f < df:
            return False
        if dt and f > dt:
            return False
        return True

    vfilt = [v for v in ventas if _in_range(v)] if (df or dt) else ventas

    # Deuda GLOBAL (no depende del filtro).
    g_neto = sum(float(v.get("neto_mj") or 0) for v in ventas)
    g_libre = sum(float(v.get("neto_mj") or 0) for v in ventas
                  if v.get("liberado"))
    tot_abonos = sum(float(a.get("monto") or 0) for a in abonos)

    # Costos/neto del PERIODO mostrado (todo si no hay filtro).
    p_bruto = sum(float(v.get("precio_venta") or 0) for v in vfilt)
    p_com = sum(float(v.get("comision") or 0) for v in vfilt)
    p_ret = sum(float(v.get("retencion") or 0) for v in vfilt)
    p_env = sum(float(v.get("costo_envio") or 0) for v in vfilt)
    p_pub = sum(float(v.get("costo_publicidad") or 0) for v in vfilt)
    p_neto = sum(float(v.get("neto_mj") or 0) for v in vfilt)
    p_libre = sum(float(v.get("neto_mj") or 0) for v in vfilt
                  if v.get("liberado"))

    por_plat = {}
    for v in vfilt:
        p = v.get("plataforma") or "?"
        b = por_plat.setdefault(p, {"unidades": 0, "bruto": 0.0, "neto": 0.0})
        b["unidades"] += int(v.get("unidades") or 0)
        b["bruto"] += float(v.get("precio_venta") or 0)
        b["neto"] += float(v.get("neto_mj") or 0)
    return {
        "ok": True, "generado": _mj_now(),
        "filtro": {"date_from": df or "", "date_to": dt or "",
                   "activo": bool(df or dt)},
        "kpis": {
            "bruto": round(p_bruto, 2), "comision": round(p_com, 2),
            "retencion": round(p_ret, 2), "envio": round(p_env, 2),
            "publicidad": round(p_pub, 2), "neto": round(p_neto, 2),
            "neto_liberado": round(p_libre, 2),
            "neto_pendiente": round(p_neto - p_libre, 2),
            "abonos": round(tot_abonos, 2),
            "neto_global": round(g_neto, 2),
            "saldo": round(g_neto - tot_abonos, 2),
            "saldo_liberado": round(g_libre - tot_abonos, 2),
            "ventas": len(vfilt), "ventas_global": len(ventas)},
        "por_plataforma": por_plat, "ventas": vfilt, "abonos": abonos}


@app.get("/api/mj")
def mj_get(force: bool = False, date_from: str = "", date_to: str = "",
           user: dict = Depends(_current_user)):
    """Liquidación de María José.  Refresca desde las plataformas si el caché
    venció (10 min) o si force=1, y devuelve KPIs + ventas + abonos.
    Filtro opcional por fecha: ?date_from=YYYY-MM-DD&date_to=YYYY-MM-DD."""
    if force or (time.time() - _MJ_CACHE["ts"]) > _MJ_TTL:
        if _MJ_LOCK.acquire(blocking=False):
            try:
                _mj_sync()
            except Exception:
                pass
            finally:
                _MJ_LOCK.release()
    out = _mj_summary(date_from or None, date_to or None)
    out["cache_age_min"] = int((time.time() - _MJ_CACHE["ts"]) / 60)
    return out


@app.get("/api/mj/count")
def mj_count(user: dict = Depends(_current_user)):
    """Saldo pendiente de pago a María José (para el badge del menú)."""
    try:
        s = _mj_summary()["kpis"]
        return {"count": 1 if s.get("saldo", 0) > 0 else 0,
                "saldo": s.get("saldo", 0)}
    except Exception:
        return {"count": 0, "saldo": 0}


@app.post("/api/mj/sync")
def mj_sync_post(key: str = "", authorization: Optional[str] = Header(None)):
    """Fuerza una resincronización de las ventas de María José.
    Auth: ?key=BOUN_EXPORT_TOKEN (planificador) o sesión de usuario."""
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    authed = bool(token and key == token)
    if not authed:
        try:
            authed = bool(_current_user(authorization))
        except Exception:
            authed = False
    if not authed:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    r = {}
    with _MJ_LOCK:
        r = _mj_sync()
    s = _mj_summary()["kpis"]
    _cerebro_set_heartbeat(
        _MJ_TASK_ID, "ok",
        "Liquidación María José: %d venta(s) · saldo a pagar %s "
        "(liberado %s · pendiente de liberación %s)." % (
            s.get("ventas", 0), _money(s.get("saldo", 0)),
            _money(s.get("saldo_liberado", 0)), _money(s.get("neto_pendiente", 0))),
        nombre="Liquidación María José", canal="mercadolibre", icon="chart",
        cad="Diario", desc=_MJ_DESC)
    return {"ok": True, "sync": r, "kpis": s}




class MJAbonoIn(BaseModel):
    monto: float
    fecha: Optional[str] = None
    metodo: Optional[str] = ""
    nota: Optional[str] = ""


@app.post("/api/mj/abono")
def mj_abono_add(data: MJAbonoIn, user: dict = Depends(_current_user)):
    """Registra un abono (pago) hecho a María José.  Resta del saldo."""
    if not data.monto or data.monto <= 0:
        raise HTTPException(400, "El monto debe ser mayor a 0.")
    payload = {"monto": round(float(data.monto), 2),
               "metodo": (data.metodo or "").strip(),
               "nota": (data.nota or "").strip(),
               "created_by": user.get("username", "")}
    if data.fecha:
        payload["fecha"] = data.fecha[:10]
    row = db._sb_post("mj_abonos", payload)
    if row is None:
        raise HTTPException(400, "No se pudo registrar el abono.")
    return {"ok": True, "id": row.get("id")}


@app.delete("/api/mj/abono/{aid}")
def mj_abono_del(aid: int, user: dict = Depends(_current_user)):
    """Elimina un abono (corrige un registro)."""
    return {"ok": db._sb_delete("mj_abonos", "id=eq.%d" % aid)}


def _money(v) -> str:
    try:
        return "$%s" % format(int(round(float(v))), ",d")
    except Exception:
        return "$0"


# ── COMBOS (kits) — definición vía UI ────────────────────────────────────────

@app.get("/api/combos")
def combos_get(user: dict = Depends(_current_user)):
    """Devuelve el mapa de combos: {codigo_combo: [{codigo, cant}, …]}.
    Lectura para cualquier usuario (para marcar combos en Inventario)."""
    return {"ok": True, "combos": _combos_def()}


class CombosIn(BaseModel):
    combos: dict = {}     # {codigo_combo: [{codigo, cant}, …]} — reemplaza todo


@app.post("/api/combos")
def combos_set(data: CombosIn, user: dict = Depends(_admin)):
    """Reemplaza el mapa completo de combos (valida y limpia la entrada)."""
    clean = {}
    for combo, comps in (data.combos or {}).items():
        combo = str(combo).strip()
        if not combo or not isinstance(comps, list):
            continue
        lst = []
        for c in comps:
            cc = str((c or {}).get("codigo") or "").strip()
            rc = (c or {}).get("cant")
            try:
                cant = int(rc) if rc not in (None, "") else 1
            except Exception:
                cant = 1
            if cc and cant > 0:    # cant<=0 (p.ej. 0) → componente inválido, se omite
                lst.append({"codigo": cc, "cant": cant})
        if lst:
            clean[combo] = lst
    db.set_setting("combos_def", json.dumps(clean, ensure_ascii=False))
    return {"ok": True, "combos": clean}


# ── Motor de sincronización — ESCANEO de reconciliación (Web → canales) ───────
# Recorre TODO el inventario central y alinea cada publicación de los canales
# elegidos al disponible vendible (Bogotá+Yopal−pendientes) de la web BOUN, que
# es la fuente de verdad. Salta Full y catálogo (guardas del escritor). Sin tope
# de salto (ignore_delta) para poder rellenar agotadas grandes 0→N. Corre en
# segundo plano (thread) para no chocar con timeouts del proxy; el avance y el
# resultado se consultan con /api/sync/scan-status.

_SCAN_STATE = {"status": "idle"}        # idle | running | done | error
_SCAN_LOCK = threading.Lock()


def _scan_row(canal: str, r: dict, dry: bool) -> dict:
    """Normaliza el resultado de un escritor (ML/Falabella/Shopify) a una fila
    legible: actual → objetivo + acción."""
    ref = (r.get("item_id") or r.get("sku") or r.get("ref")
           or r.get("inv_item") or "")
    actual, objetivo = r.get("actual"), r.get("set")
    if r.get("skip"):
        accion = "saltado:%s" % (r.get("reason") or "")
    elif not r.get("ok"):
        accion = "error"
    elif r.get("reactivar"):
        accion = "reactivaria" if dry else "reactivada"
    elif actual is not None and objetivo is not None:
        accion = "sin_cambio" if int(actual) == int(objetivo) else \
                 ("cambiaria" if dry else "escrito")
    else:
        # Falabella no expone snapshot del stock actual → escribe el absoluto.
        accion = "escribiria" if dry else "escrito"
    return {"canal": canal, "ref": str(ref), "actual": actual,
            "objetivo": objetivo, "accion": accion,
            "detalle": str(r.get("error") or "")[:160]}


def _scan_reconcile(channels: set, dry: bool, tag: str = "scan"):
    """Núcleo del escaneo. Recorre todos los productos del inventario central,
    calcula el disponible y alinea las publicaciones de `channels`. Va volcando
    progreso y filas en `_SCAN_STATE`. No relanza: deja el error en el estado."""
    prods = db._sb_get("inventory_products?select=code,qty_bogota,qty_yopal"
                       "&order=code") or []
    total = len(prods)
    reactivate = _scan_reactivate_enabled()
    counts = {"cambios": 0, "sin_cambio": 0, "saltados": 0,
              "errores": 0, "escritos": 0, "reactivadas": 0}
    rows = []
    _SCAN_STATE.clear()
    _SCAN_STATE.update({
        "status": "running", "mode": "preview" if dry else "apply",
        "channels": sorted(channels), "done": 0, "total": total,
        "counts": counts, "rows": rows,
        "started_at": datetime.now(timezone.utc).strftime("%Y-%m-%dT%H:%M:%SZ")})
    for i, p in enumerate(prods, 1):
        codigo = p.get("code") or ""
        if _combo_components(codigo):
            # Combo: disponible = cuántos se pueden armar de los componentes.
            disp = _combo_disponible(codigo)
        else:
            bog = int(p.get("qty_bogota") or 0)
            yop = int(p.get("qty_yopal") or 0)
            disp = max(0, bog + yop - _pending_cola(codigo))
        try:
            plan = _compute_plan(codigo, disp, reactivate=reactivate)
            res = _apply_plan(codigo, plan, order_id=tag, dry=dry,
                              force=channels, ignore_delta=True,
                              reactivate=reactivate)
        except Exception as e:
            counts["errores"] += 1
            rows.append({"codigo": codigo, "canal": "-", "ref": "", "actual": None,
                         "objetivo": None, "accion": "error",
                         "detalle": str(e)[:160], "disponible": disp})
            _SCAN_STATE["done"] = i
            continue
        for canal, lst in (res.get("canales") or {}).items():
            for r in (lst or []):
                row = _scan_row(canal, r, dry)
                row["codigo"] = codigo
                row["disponible"] = disp
                rows.append(row)
                a = row["accion"]
                if a == "escrito":
                    counts["cambios"] += 1
                    counts["escritos"] += 1
                elif a == "reactivada":
                    counts["cambios"] += 1; counts["escritos"] += 1
                    counts["reactivadas"] += 1
                elif a == "reactivaria":
                    counts["cambios"] += 1; counts["reactivadas"] += 1
                elif a in ("cambiaria", "escribiria"):
                    counts["cambios"] += 1
                elif a == "sin_cambio":
                    counts["sin_cambio"] += 1
                elif a.startswith("saltado"):
                    counts["saltados"] += 1
                else:
                    counts["errores"] += 1
        _SCAN_STATE["done"] = i
    _SCAN_STATE["status"] = "done"
    _SCAN_STATE["finished_at"] = datetime.now(timezone.utc).strftime(
        "%Y-%m-%dT%H:%M:%SZ")
    _safe(db.set_setting, "sync_scan_last", json.dumps({
        "mode": _SCAN_STATE["mode"], "channels": _SCAN_STATE["channels"],
        "counts": counts, "total": total,
        "finished_at": _SCAN_STATE["finished_at"]}))


def _scan_run_bg(channels: set, dry: bool):
    try:
        _scan_reconcile(channels, dry)
    except Exception as e:
        _SCAN_STATE.update({"status": "error", "error": str(e)[:200]})
    finally:
        if _SCAN_LOCK.locked():
            _SCAN_LOCK.release()


class ScanStartIn(BaseModel):
    mode: str = "preview"                 # "preview" (no escribe) | "apply"
    channels: Optional[list] = None       # default: los 4 canales


@app.post("/api/sync/scan-start")
def sync_scan_start(data: ScanStartIn, user: dict = Depends(_admin)):
    """Inicia un escaneo de reconciliación en segundo plano. mode='preview' no
    escribe nada (solo calcula actual→objetivo); mode='apply' escribe el
    disponible real en cada publicación de los canales elegidos (salta Full y
    catálogo, sin tope de salto). Avanza con /api/sync/scan-status."""
    dry = (data.mode or "preview") != "apply"
    chans = ({str(c).strip() for c in (data.channels or []) if str(c).strip()}
             or set(_APPLY_CHANNELS_VALIDOS))
    invalidos = chans - _APPLY_CHANNELS_VALIDOS
    if invalidos:
        raise HTTPException(400, "Canales no válidos: %s"
                            % ", ".join(sorted(invalidos)))
    if not _SCAN_LOCK.acquire(blocking=False):
        raise HTTPException(409, "Ya hay un escaneo en curso")
    threading.Thread(target=_scan_run_bg, args=(chans, dry),
                     daemon=True).start()
    return {"ok": True, "started": True, "mode": "preview" if dry else "apply",
            "channels": sorted(chans)}


@app.get("/api/sync/scan-status")
def sync_scan_status(user: dict = Depends(_admin)):
    """Estado y avance del último escaneo (admin). Para no inflar la respuesta,
    devuelve solo las filas con cambio o incidencia (omite las 'sin_cambio').

    Mientras el escaneo está corriendo NO se devuelven las filas: el front solo
    pinta la barra de progreso (done/total/mode), y copiar/serializar la lista
    de filas —que crece— en cada sondeo dispara picos de memoria justo en la
    ventana del escaneo diario (causa de los OOM >512MB). Las filas se entregan
    solo cuando el escaneo termina (done/error)."""
    s = dict(_SCAN_STATE)
    status = s.get("status")
    if status == "running":
        # Respuesta liviana: solo lo que necesita la barra de progreso.
        return {"status": "running", "mode": s.get("mode"),
                "channels": s.get("channels"), "done": s.get("done", 0),
                "total": s.get("total", 0), "counts": s.get("counts"),
                "started_at": s.get("started_at")}
    # snapshot defensivo: el thread sigue haciendo append a la misma lista, y
    # copiar una lista que crece puede lanzar RuntimeError en CPython → reintenta.
    src = s.get("rows") or []
    for _ in range(3):
        try:
            rows = list(src)
            break
        except RuntimeError:
            time.sleep(0.05)
    else:
        rows = []
    s["rows"] = [r for r in rows if r.get("accion") != "sin_cambio"]
    s["rows_total"] = len(rows)
    return s


# ── Escaneo de reconciliación — AUTOMÁTICO diario (servidor) ─────────────────
# Corre el escaneo en modo Aplicar una vez al día a la hora configurada (hora
# Colombia, UTC-5). Idempotente por fecha (setting sync_scan_daily_last) para no
# repetir aunque el proceso reinicie por un cold-start. Respeta _SCAN_LOCK.

def _scan_daily_enabled() -> bool:
    return (db.get_setting("sync_scan_daily", "") == "1"
            or os.environ.get("SYNC_SCAN_DAILY", "") == "1")


def _scan_reactivate_enabled() -> bool:
    """Si el escaneo reactiva las publicaciones ML pausadas por falta de stock.
    Default ON (solo se desactiva poniendo el setting en '0')."""
    return db.get_setting("sync_scan_reactivate", "1") != "0"


def _ml_solo_bogota() -> bool:
    """Regla TEMPORAL: si está activa, MercadoLibre solo vende/refleja el stock de
    la bodega Bogotá (Yopal no cuenta para ML). Los demás canales usan ambas."""
    return db.get_setting("ml_solo_bogota", "") == "1"


def _scan_daily_hour() -> int:
    try:
        h = int(db.get_setting("sync_scan_daily_hour", "") or "4")
        return h if 0 <= h <= 23 else 4
    except Exception:
        return 4


def _scan_daily_loop():
    import datetime as _dt
    while True:
        try:
            if _scan_daily_enabled():
                # Hora local Colombia (UTC-5, sin horario de verano).
                now_co = _dt.datetime.now(_dt.timezone.utc) - _dt.timedelta(hours=5)
                hoy = now_co.strftime("%Y-%m-%d")
                if (now_co.hour == _scan_daily_hour()
                        and db.get_setting("sync_scan_daily_last", "") != hoy):
                    # Marca la fecha ANTES de correr → si reinicia a media corrida
                    # no la repite hoy. El escaneo es idempotente igual.
                    db.set_setting("sync_scan_daily_last", hoy)
                    if _SCAN_LOCK.acquire(blocking=False):
                        try:
                            _scan_reconcile(set(_APPLY_CHANNELS_VALIDOS),
                                            dry=False, tag="scan-auto")
                        finally:
                            if _SCAN_LOCK.locked():
                                _SCAN_LOCK.release()
        except Exception:
            pass
        time.sleep(60)   # revisa cada minuto


@app.on_event("startup")
def _start_scan_daily():
    threading.Thread(target=_scan_daily_loop, daemon=True).start()


# ── Motor de sincronización — PLAN en DRY-RUN (no escribe en ningún canal) ────

@app.options("/api/sync/plan")
def sync_plan_preflight():
    return Response(status_code=204, headers=_EXPORT_CORS)


@app.get("/api/sync/plan")
def sync_plan(key: str = "", codigo: str = "", vendidos: int = 0):
    """Calcula el disponible de un código BOUN y cómo se repartiría entre las
    publicaciones activas de cada canal, simulando una venta de `vendidos`.
    NO escribe nada. Sirve para validar la lógica antes de activar la propagación.
    """
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    if not token or key != token:
        return JSONResponse({"error": "unauthorized"}, status_code=401,
                            headers=_EXPORT_CORS)
    if not codigo:
        return JSONResponse({"error": "bad_request"}, status_code=400,
                            headers=_EXPORT_CORS)
    try:
        from urllib.parse import quote as _q
        import sync as _sync
        prows = db._sb_get(
            "inventory_products?code=eq.%s&select=id,code,name,qty_bogota,"
            "qty_yopal,qty_transit" % _q(codigo, safe="")) or []
        if not prows:
            return JSONResponse({"error": "codigo_no_encontrado",
                                 "codigo": codigo}, status_code=404,
                                headers=_EXPORT_CORS)
        p = prows[0]
        bog = int(p.get("qty_bogota") or 0)
        yop = int(p.get("qty_yopal") or 0)
        disponible = max(0, bog + yop - int(vendidos or 0))

        # ── MercadoLibre: publicaciones activas no-Full del código ──
        links = db._sb_get(
            "inventory_links?product_id=eq.%d&%sselect=ml_item_id,ml_sold60,"
            "ml_logistic" % (p["id"], db._ml_only_filter())) or []
        ml_pubs, ml_excluidas = [], []
        for l in links:
            iid = l.get("ml_item_id")
            if not iid:
                continue
            r = _ml_request("GET", "/items/%s?attributes=id,status,shipping"
                                   % iid)
            st = lg = None
            if r is not None and r.status_code == 200:
                jd = r.json(); st = jd.get("status")
                lg = (jd.get("shipping") or {}).get("logistic_type")
            if lg == "fulfillment":
                ml_excluidas.append({"item_id": iid, "motivo": "full"})
            elif st != "active":
                ml_excluidas.append({"item_id": iid, "motivo": st or "?"})
            else:
                ml_pubs.append({"key": iid,
                                "ventas": int(l.get("ml_sold60") or 0)})
        ml_reparto = _sync.reparto(disponible, ml_pubs)

        # ── Falabella: SKUs mapeados del código (desde el CSV) ──
        fal = _sync.falabella_skus(codigo)
        fal_pubs = [{"key": f["seller_sku"], "ventas": 0} for f in fal]
        fal_reparto = _sync.reparto(disponible, fal_pubs)

        out = {
            "codigo": codigo, "producto": p.get("name"),
            "stock": {"bogota": bog, "yopal": yop,
                      "transit": int(p.get("qty_transit") or 0)},
            "vendidos_simulado": int(vendidos or 0),
            "disponible_vendible": disponible,
            "canales": {
                "mercadolibre": {"activas": list(ml_reparto.keys()),
                                 "reparto": ml_reparto,
                                 "excluidas": ml_excluidas},
                "falabella": {"skus": [f["seller_sku"] for f in fal],
                              "reparto": fal_reparto,
                              "mapeado_en_csv": bool(fal)},
                "shopify_boun": {"pendiente": "credenciales_tienda"},
                "shopify_kat": {"pendiente": "credenciales_tienda"},
            },
            "dry_run": True,
            "nota": "Cálculo sin escribir en canales (Fase 1).",
        }
        return JSONResponse(out, headers=_EXPORT_CORS)
    except Exception as e:
        return JSONResponse({"error": str(e)[:200]}, status_code=502,
                            headers=_EXPORT_CORS)


# ── Pendientes de bodega (¿de qué bodega salió cada venta?) ──────────────────
# El disponible TOTAL ya se descontó en la venta; esto solo cuadra de cuál
# bodega salió. Los casos de una sola bodega con stock se auto-asignan en el
# motor; aquí quedan únicamente los ambiguos (ambas bodegas con stock).

class ConfirmBodegaIn(BaseModel):
    bodega: str


def _stock_map(codigos: list) -> dict:
    """{codigo: {name, qty_bogota, qty_yopal, img}} para una lista de códigos.
    La foto sale del thumb de ML (inventory_links.ml_thumb, subido a alta
    resolución) y, si no hay, de image_path cuando es una URL pública."""
    out = {}
    cods = [c for c in dict.fromkeys(codigos) if c]
    if not cods:
        return out
    # Una consulta por código con filtro code=eq.<código> (patrón probado del
    # motor). OJO: inventory_products NO tiene image_path → pedirla rompía TODO
    # el select (400). La foto sale de inventory_links.ml_thumb.
    for c in cods:
        rows = db._sb_get("inventory_products?code=eq.%s&select=id,code,name,"
                          "qty_bogota,qty_yopal&limit=1" % _q_(c)) or []
        if not rows:
            continue
        r = rows[0]
        img = ""
        links = db._sb_get("inventory_links?product_id=eq.%d&select=ml_thumb"
                           "&limit=10" % r.get("id")) or []
        for l in links:
            thumb = l.get("ml_thumb") or ""
            if thumb:
                img = _img_full(thumb)
                break
        out[c] = {"name": r.get("name"),
                  "qty_bogota": int(r.get("qty_bogota") or 0),
                  "qty_yopal": int(r.get("qty_yopal") or 0),
                  "img": img}
    return out


@app.get("/api/cola-bodega")
def cola_bodega_list(user: dict = Depends(_current_user)):
    rows = db._sb_get("cola_bodega?estado=eq.pendiente&select=id,codigo_boun,"
                      "nombre,cantidad,canal,order_id,created_at&"
                      "order=created_at.asc") or []
    sm = _stock_map([r.get("codigo_boun") for r in rows])
    out = []
    for r in rows:
        s = sm.get(r.get("codigo_boun"), {})
        out.append({
            "id": r.get("id"), "codigo_boun": r.get("codigo_boun"),
            "nombre": r.get("nombre") or s.get("name") or "",
            "cantidad": int(r.get("cantidad") or 0), "canal": r.get("canal"),
            "order_id": r.get("order_id"), "created_at": r.get("created_at"),
            "img": s.get("img", ""),
            "stock_bogota": s.get("qty_bogota", 0),
            "stock_yopal": s.get("qty_yopal", 0)})
    return {"pendientes": out, "total": len(out)}


@app.get("/api/cola-bodega/count")
def cola_bodega_count(user: dict = Depends(_current_user)):
    rows = db._sb_get("cola_bodega?estado=eq.pendiente&select=id") or []
    return {"count": len(rows)}


@app.post("/api/cola-bodega/{cid}/confirmar")
def cola_bodega_confirmar(cid: int, data: ConfirmBodegaIn,
                          user: dict = Depends(_current_user)):
    bodega = (data.bodega or "").strip().lower()
    if bodega not in ("bogota", "yopal"):
        raise HTTPException(400, "bodega debe ser 'bogota' o 'yopal'")
    rows = db._sb_get("cola_bodega?id=eq.%d&select=id,codigo_boun,cantidad,"
                      "estado,canal,order_id" % cid) or []
    if not rows:
        raise HTTPException(404, "Pendiente no encontrado")
    c = rows[0]
    if c.get("estado") != "pendiente":
        raise HTTPException(409, "Ese pendiente ya fue procesado")
    prows = db._sb_get("inventory_products?code=eq.%s&select=id,name,qty_bogota,"
                       "qty_yopal" % _q_(c["codigo_boun"])) or []
    if not prows:
        raise HTTPException(404, "Producto no está en el inventario central")
    p = prows[0]
    col = "qty_bogota" if bodega == "bogota" else "qty_yopal"
    actual = int(p.get(col) or 0)
    cant = int(c.get("cantidad") or 0)
    if actual < cant:
        nb = "Bogotá" if bodega == "bogota" else "Yopal"
        otra = "Yopal" if bodega == "bogota" else "Bogotá"
        raise HTTPException(400, "%s solo tiene %d; ¿salió de %s?"
                            % (nb, actual, otra))
    nuevo = actual - cant
    db._sb_patch("inventory_products", "id=eq.%d" % p["id"], {col: nuevo})
    db._sb_post("movimiento_stock", {
        "codigo_boun": c["codigo_boun"], "delta": -cant,
        "motivo": "asignacion_bodega_%s" % bodega, "canal": c.get("canal"),
        "order_id": c.get("order_id")})
    db._sb_patch("cola_bodega", "id=eq.%d" % cid, {
        "estado": "confirmado", "bodega_asignada": bodega,
        "confirmado_at": datetime.now(timezone.utc).strftime(
            "%Y-%m-%dT%H:%M:%SZ")})
    return {"ok": True, "id": cid, "bodega": bodega, "saldo_nuevo": nuevo}


@app.post("/api/cola-bodega/{cid}/full")
def cola_bodega_full(cid: int, user: dict = Depends(_current_user)):
    rows = db._sb_get("cola_bodega?id=eq.%d&select=id,estado" % cid) or []
    if not rows:
        raise HTTPException(404, "Pendiente no encontrado")
    if rows[0].get("estado") != "pendiente":
        raise HTTPException(409, "Ese pendiente ya fue procesado")
    db._sb_patch("cola_bodega", "id=eq.%d" % cid, {
        "estado": "full", "confirmado_at": datetime.now(timezone.utc).strftime(
            "%Y-%m-%dT%H:%M:%SZ")})
    return {"ok": True, "id": cid, "estado": "full"}


# ── Shopify OAuth (captura del Admin API token de cada tienda) ───────────────
# Instala la app "BOUN Sync Stock" en una tienda y guarda su token en Supabase
# (app_settings), sin que nadie tenga que copiar el token a mano.

_SHOP_SCOPES = ("read_products,write_products,read_inventory,write_inventory,"
                "read_orders,read_locations")
_SHOP_REDIRECT = "https://boun-web-deploy.onrender.com/shopify/callback"


def _shopify_app_creds(shop: str = ""):
    """Credenciales de la app Shopify. Por-tienda (shopify_api_key::{shop}) si
    existen, si no las globales (la app de BOUN). KAT está en otra organización
    y usa su propia app (creds por-tienda)."""
    cid = db.get_setting("shopify_api_key::%s" % shop, "") if shop else ""
    sec = db.get_setting("shopify_api_secret::%s" % shop, "") if shop else ""
    if not cid:
        cid = (db.get_setting("shopify_api_key", "")
               or os.environ.get("SHOPIFY_API_KEY", ""))
    if not sec:
        sec = (db.get_setting("shopify_api_secret", "")
               or os.environ.get("SHOPIFY_API_SECRET", ""))
    return cid, sec


@app.get("/shopify/install")
def shopify_install(shop: str = "", key: str = ""):
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    if not token or key != token:
        return JSONResponse({"error": "unauthorized"}, status_code=401)
    shop = (shop or "").strip()
    cid, _sec = _shopify_app_creds(shop)
    if not cid:
        return JSONResponse({"error": "missing_shopify_api_key"},
                            status_code=500)
    if not shop.endswith(".myshopify.com"):
        return JSONResponse({"error": "bad_shop",
                             "hint": "usa el dominio .myshopify.com"},
                            status_code=400)
    import urllib.parse as _u
    url = ("https://%s/admin/oauth/authorize?client_id=%s&scope=%s"
           "&redirect_uri=%s&state=boun" % (
               shop, cid, _SHOP_SCOPES, _u.quote(_SHOP_REDIRECT, safe="")))
    return RedirectResponse(url)


@app.get("/shopify/callback")
async def shopify_callback(request: Request):
    params = dict(request.query_params)
    shop = params.get("shop", "")
    code = params.get("code", "")
    cid, sec = _shopify_app_creds(shop)
    given_hmac = params.get("hmac", "")
    if not (cid and sec and shop and code):
        return HTMLResponse("<h3>Falta configuración o parámetros.</h3>",
                            status_code=400)
    # Verificar HMAC del callback
    msg = "&".join("%s=%s" % (k, params[k]) for k in sorted(params)
                   if k not in ("hmac", "signature"))
    calc = _hmac.new(sec.encode(), msg.encode(), _hashlib.sha256).hexdigest()
    if not _hmac.compare_digest(calc, given_hmac):
        return HTMLResponse("<h3>HMAC inválido.</h3>", status_code=401)
    # Intercambiar code por access_token
    try:
        import requests as _rq
        r = _rq.post("https://%s/admin/oauth/access_token" % shop,
                     json={"client_id": cid, "client_secret": sec,
                           "code": code}, timeout=20)
        if r.status_code != 200:
            return HTMLResponse("<h3>Error al obtener token: %s</h3>"
                                % r.text[:200], status_code=502)
        tok = r.json().get("access_token", "")
        db.set_setting("shopify_token::%s" % shop, tok)
        # Guardar también el location principal
        try:
            lr = _rq.get("https://%s/admin/api/2025-01/locations.json" % shop,
                         headers={"X-Shopify-Access-Token": tok}, timeout=15)
            locs = lr.json().get("locations", []) if lr.status_code == 200 else []
            if locs:
                db.set_setting("shopify_location::%s" % shop,
                               str(locs[0].get("id")))
        except Exception:
            pass
        return HTMLResponse(
            "<h2>✅ Tienda conectada: %s</h2>"
            "<p>El token quedó guardado de forma segura. Ya puedes cerrar "
            "esta pestaña.</p>" % shop)
    except Exception as e:
        return HTMLResponse("<h3>Error: %s</h3>" % str(e)[:200],
                            status_code=502)


@app.get("/shopify/status")
def shopify_status(key: str = ""):
    token = os.environ.get("BOUN_EXPORT_TOKEN", "")
    if not token or key != token:
        return JSONResponse({"error": "unauthorized"}, status_code=401,
                            headers=_EXPORT_CORS)
    rows = db._sb_get("app_settings?key=like.shopify_*&select=key,value") or []
    out = {}
    for r in rows:
        k = r.get("key", "")
        out[k] = "***" if ("token" in k or "secret" in k) else r.get("value")
    return JSONResponse({"configurado": out}, headers=_EXPORT_CORS)


# ── Frontend estático ────────────────────────────────────────────────────────

_FRONT = os.path.join(os.path.dirname(os.path.dirname(__file__)), "frontend")


@app.get("/")
def index():
    return FileResponse(os.path.join(_FRONT, "index.html"))


app.mount("/", StaticFiles(directory=_FRONT), name="static")
